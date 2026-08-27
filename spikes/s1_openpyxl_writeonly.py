# S1: openpyxl write_only で「金額のみ数値型・他は文字列型・〓に条件付き書式・COUNTIF」
# が同時に成立するかの検証（実装レビュー M3・設計 §6.6）
import os

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PATH = os.path.join(ROOT, "workdir", "s1_writeonly.xlsx")

wb = Workbook(write_only=True)
ws = wb.create_sheet("output")

# 条件付き書式はワークシートレベル: write_only でも成立するかが論点のひとつ
fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ws.conditional_formatting.add("B1:C10", CellIsRule(operator="equal", formula=['"〓"'], fill=fill))

def text_cell(v):
    c = WriteOnlyCell(ws, value=v)
    c.number_format = "@"  # 文字列型: 先頭ゼロ保持
    return c

def num_cell(v):
    return WriteOnlyCell(ws, value=v)  # 数値型（金額列相当）

row_no = 0  # 行カウンタを自前で保持（COUNTIF の行番号がズレないため）

# ヘッダ
row_no += 1
ws.append([text_cell("要確認セル数"), text_cell("電話番号"), text_cell("品目"), text_cell("金額")])

# データ3行: 先頭ゼロ・〓・数値の混在
rows = [("0471234567", "〓", 10000), ("〓", "家内安全", 100), ("09012345678", "〓", 1000000)]
for tel, item, amount in rows:
    row_no += 1
    countif = WriteOnlyCell(ws, value=f'=COUNTIF(B{row_no}:C{row_no},"〓")')
    ws.append([countif, text_cell(tel), text_cell(item), num_cell(amount)])

wb.save(PATH)

# --- 読み戻して検証 ---
wb2 = load_workbook(PATH)
ws2 = wb2["output"]
ok = True

def check(label, cond):
    global ok
    print(f"{'OK' if cond else 'NG'}  {label}")
    ok = ok and cond

check("B2 先頭ゼロ保持（値）", ws2["B2"].value == "0471234567")
check("B2 number_format='@'", ws2["B2"].number_format == "@")
check("D2 金額は数値型のまま", isinstance(ws2["D2"].value, int))
check("A2 COUNTIF 数式が残る", ws2["A2"].value == '=COUNTIF(B2:C2,"〓")')
check("A4 行カウンタ追従", ws2["A4"].value == '=COUNTIF(B4:C4,"〓")')
cf = list(ws2.conditional_formatting)
check("条件付き書式が保存される", len(cf) == 1 and str(cf[0].sqref) == "B1:C10")
print("\n結論:", "write_only で全要件成立" if ok else "不成立 → 通常モードへ切替")
