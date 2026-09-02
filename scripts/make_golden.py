"""golden（replay 経由の run→render 出力）を再現可能な手順で作り直す。

背景（2026-09-02・#70 前提作業 §10.2-3 の再検証）: 最初に作った
`testdata/golden_manifest.json`（HEAD 71384a4）は、実行に使った Config の
`unclear_threshold` を manifest に残していなかった。この値は既定の 0.85 では
なく **0.4**（`core/tests/test_e2e_replay.py` の `outputs` fixture と同じ値・
配置の正しさを検証する「第1層検証」用の意図的な低め設定）を使っていた。
第三者（おかゆ）がこの手順を「デフォルト設定だろう」と推測して 0.85 で
再現した結果、同一の入力から出力の sha256 が一致しなくなった
（実測: 差分18セル・うち17セルが 0.4 では読める値なのに 0.85 では〓化。
コーディネーター報告の「19セル中18セルが〓」とほぼ一致）。

このスクリプトは同じ divergence を二度と生まないよう、
1. **すべての入力（テンプレート・画像・S2応答・Config値）を引数で明示**し、
2. manifest に **procedure（手順の全パラメータ）** を必ず記録し、
3. 出力の同一性判定を「ファイルの sha256」ではなく
   **セル値を正規化した配列のハッシュ**（openpyxl のブック作成メタデータ・
   csv の改行/BOM 差に左右されない）で行う。

使い方（既定値はすべてこのリポジトリの実データを指す。実 API は一切呼ばない
——常に ReplayClient 経由）:
    .venv\\Scripts\\python.exe -X utf8 scripts\\make_golden.py

出力: `workdir/golden/<head_short>/output.{xlsx,csv}`（.gitignore 対象・
記入値を含むためコミットしない）と `testdata/golden_manifest.json`
（記入値を含まない・コミット対象）。
"""
from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "core"))

from chouhyo_ocr.config import Config  # noqa: E402
from chouhyo_ocr.pipeline import render, run  # noqa: E402
from chouhyo_ocr.vision_client import ReplayClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

# --- 既定パス（すべて .gitignore 対象・このマシン限定のローカルデータ）---
DEFAULT_TEMPLATE = REPO_ROOT / "templates" / "chouhyo-v1.json"
DEFAULT_PAGE_PNG = REPO_ROOT / "workdir" / "pages" / "sample-1.png"
DEFAULT_RESPONSE = REPO_ROOT / "workdir" / "s2" / "resp_DOCUMENT_TEXT_DETECTION.json"
# 位置合わせ後のファイル名規約（ingest の既定命名）: "<入力ファイル名(拡張子なし)>_p0001"
DEFAULT_PAGE_ID = "sample-1_p0001"

# 「配置の正しさ（第1層検証）」用の意図的な低め閾値。既定 0.85 ではない——
# この値を変えると〓化するセル数が大きく変わる（本スクリプトの docstring の
# 実測参照）。golden の再現性はこの値の一致に依存するため、変更する場合は
# 必ず procedure.config.unclear_threshold として manifest に残ること
# （このスクリプトは常に残す）。
DEFAULT_UNCLEAR_THRESHOLD = 0.4


def _normalized_xlsx_hash(xlsx_path: Path) -> tuple[str, int, int, int]:
    """セル値配列の正規化ハッシュ・(行数[ヘッダ除く]・列数・〓の個数)。

    ファイルの sha256 ではなくセル値を対象にする——openpyxl が書き出す
    ブックのメタデータ（作成日時等）は本リポジトリの環境では固定されており
    実際には決定論的だったが、環境差でメタデータが変わりうる将来のために
    「セル値が同じなら同じハッシュ」を保証する設計にする。
    """
    wb = load_workbook(xlsx_path)
    ws = wb["output"]
    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    n_cols = len(header)
    n_rows = len(data_rows)
    n_masked = sum(1 for row in data_rows for v in row if v == "〓")
    # \x1f（セル区切り）・\x1e（行区切り）は通常のセル内容に出現しない制御文字
    blob = "\x1e".join(
        "\x1f".join("" if v is None else str(v) for v in row)
        for row in rows
    ).encode("utf-8")
    return hashlib.sha256(blob).hexdigest(), n_rows, n_cols, n_masked


def _normalized_csv_hash(csv_path: Path) -> str:
    """BOM・改行コードを正規化してからの内容ハッシュ。"""
    text = csv_path.read_text(encoding="utf-8-sig")
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    ap.add_argument("--page-png", type=Path, default=DEFAULT_PAGE_PNG)
    ap.add_argument("--response", type=Path, default=DEFAULT_RESPONSE)
    ap.add_argument("--page-id", default=DEFAULT_PAGE_ID)
    ap.add_argument("--unclear-threshold", type=float, default=DEFAULT_UNCLEAR_THRESHOLD)
    args = ap.parse_args()

    for label, p in (("template", args.template), ("page_png", args.page_png),
                     ("response", args.response)):
        if not p.exists():
            print(f"[SKIP] {label} が見つからない: {p}（このマシン限定のローカル素材が無い環境）",
                 file=sys.stderr)
            sys.exit(1)

    head = subprocess.check_output(["git", "rev-parse", "HEAD"], text=True, cwd=REPO_ROOT).strip()
    head_short = subprocess.check_output(["git", "rev-parse", "--short", "HEAD"], text=True, cwd=REPO_ROOT).strip()
    dirty = subprocess.run(["git", "diff", "--quiet"], cwd=REPO_ROOT).returncode != 0
    dirty_names = []
    if dirty:
        out = subprocess.check_output(["git", "status", "--porcelain"], text=True, cwd=REPO_ROOT)
        dirty_names = [ln[3:] for ln in out.splitlines() if ln.strip()]

    cwd_used = str(REPO_ROOT)

    golden_dir = REPO_ROOT / "workdir" / "golden" / head_short
    if golden_dir.exists():
        shutil.rmtree(golden_dir)
    golden_dir.mkdir(parents=True)

    with tempfile.TemporaryDirectory() as tmp_s:
        tmp = Path(tmp_s)
        input_dir = tmp / "input"; input_dir.mkdir()
        # ページIDの導出は ingest の既定命名規約（"<ファイル名(拡張子なし)>_p0001"）に
        # 従う。入力側のファイル名は必ず page_id の頭部（"_p0001" を除いた部分）
        # に合わせる——これがずれると ReplayClient が対応する応答ファイルを
        # 見つけられず SendError("REPLAY_MISSING") になる
        stem = args.page_id.rsplit("_p", 1)[0]
        input_name = f"{stem}{args.page_png.suffix}"
        shutil.copy(args.page_png, input_dir / input_name)
        replay_dir = tmp / "responses"; replay_dir.mkdir()
        shutil.copy(args.response, replay_dir / f"{args.page_id}.json")

        cfg = Config(unclear_threshold=args.unclear_threshold,
                    output_dir=str(tmp / "out"), workdir=str(tmp / "wd"), log_dir=str(tmp / "logs"))
        summary = run(input_dir, args.template, cfg, ReplayClient(replay_dir))
        xlsx, csvp, rows = render(args.template, cfg, timestamp="golden")

        dst_xlsx = golden_dir / "output.xlsx"
        dst_csv = golden_dir / "output.csv"
        shutil.copy(xlsx, dst_xlsx)
        shutil.copy(csvp, dst_csv)

    xlsx_hash, n_rows, n_cols, n_masked = _normalized_xlsx_hash(dst_xlsx)
    csv_hash = _normalized_csv_hash(dst_csv)

    manifest = {
        "head": head,
        "head_short": head_short,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "procedure": {
            "script": "scripts/make_golden.py",
            "cwd": cwd_used,
            "worktree": "メインの作業ツリー（git worktree で分離していない）",
            "working_tree_dirty_at_generation": dirty,
            "dirty_files_at_generation": dirty_names,
            "python": sys.executable,
            "template": str(args.template.relative_to(REPO_ROOT)),
            "page_png": str(args.page_png.relative_to(REPO_ROOT)) + "（.gitignore 対象・ローカル限定）",
            "s2_response": str(args.response.relative_to(REPO_ROOT)) + "（.gitignore 対象・ローカル限定）",
            "page_id": args.page_id,
            "replay": True,
            "api_calls": getattr(summary, "api_calls", None),
            "config": {
                "unclear_threshold": args.unclear_threshold,
                "note": ("既定値 0.85 ではなく 0.4 を使用。理由: "
                        "core/tests/test_e2e_replay.py の outputs fixture と同じ、"
                        "配置の正しさ（第1層検証）を見るための意図的な低め設定。"),
            },
            "root_cause_of_2026-09-02_divergence": (
                "最初の golden_manifest.json（HEAD 71384a4）はファイルの sha256 のみを"
                "記録し、unclear_threshold=0.4 を使ったことを procedure として残して"
                "いなかった。第三者が既定値 0.85 で再現した結果、〓化するセルが"
                "大幅に増え（実測: 同一入力で threshold=0.4→masked_cells=4 に対し"
                "0.85→masked_cells=21・両者の差分18セルのうち17セルが0.4では読める"
                "値・0.85では〓）、出力が一致しなくなった。原因はコードの版差では"
                "なく Config の unclear_threshold の取り違えだった——本スクリプトは"
                "この値を procedure.config として必ず manifest に残すことで再発を防ぐ。"
            ),
        },
        "outputs": {
            "output.xlsx": {
                "path": f"workdir/golden/{head_short}/output.xlsx",
                "cell_value_sha256": xlsx_hash,
                "rows": n_rows,
                "cols": n_cols,
                "masked_cells": n_masked,
            },
            "output.csv": {
                "path": f"workdir/golden/{head_short}/output.csv",
                "normalized_sha256": csv_hash,
            },
        },
        "note": ("記入値（帳票の読取結果そのもの）はこのファイルに含めない。"
                "output.xlsx/output.csv 自体は workdir/ 配下（.gitignore 対象）に"
                "置きコミットしない。ハッシュはファイルバイト列ではなく"
                "セル値・正規化テキストを対象にしている（procedure 参照）。"),
    }

    out_manifest = REPO_ROOT / "testdata" / "golden_manifest.json"
    out_manifest.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {out_manifest}")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
