"""出力（.xlsx / .csv 同時生成・設計 §6.6）。

- .xlsx: write_only（M0-S1 で成立確認）。金額のみ数値型・他は文字列型（'@'）・
  〓（欄全体・一部）へ条件付き書式（U-12）・参照先採用セルへ静的な由来色
  （U-04・FILL_ORIGIN_FALLBACK）・要確認セル数は COUNTIF 数式（U-13・ワイルドカード）
- .csv: 全列クォート・BOM 付き UTF-8・要確認セル数は静的な数値（U-13・「〓を含む」で数える）
- 再現性: 同一中間データ・同一設定 → バイト一致（要件 §6.2）。xlsx は
  docProps の日時を固定し、zip エントリの日時も正規化する

2026-08-31（5巡目 第2段・docs/design/chouhyo-ocr/04_unclear_policy.md §8）:
文字単位〓（#62）で「〓」を含むが完全一致ではないセルが生じるため、要確認
セル数の数え方を完全一致から「含む」へ統一した（xlsx の COUNTIF・条件付き
書式・csv の3経路）。

2026-08-31（QA 再判定・T-16 ブロッカーの解消）: 上記の「含む」化は
unclear_char_level でゲートする。COUNTIF のワイルドカード（"*〓*"）が
Excel 実機で期待どおり動くかは T-16 として未検証のため、既定 OFF の経路
（大半の利用者）にはこの未検証の仮定を載せない——OFF では COUNTIF・
条件付き書式・csv 判定のすべてが機能追加前と同じ完全一致に戻る。ON に
した場合のみ「含む」化が効く（詳細は 04_unclear_policy.md §8.3/§8.5）。
"""
from __future__ import annotations

import csv
import io
import os
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill

from .columns import META_COLUMNS, excel_column_letter
from .render_rows import Row

_FIXED_DT = datetime(2000, 1, 1)
_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
# 一部〓（U-12）: セル値が "〓" を含み、かつ長さ2以上（完全一致の _FILL とは
# 別ルール・2026-08-31）。橙は debug-images で「信頼度不足」を表す色と系統を揃える
_FILL_PARTIAL = PatternFill(start_color="FFE8CC", end_color="FFE8CC", fill_type="solid")
# 参照先採用セルの由来色（U-04）。値は書き換えない（転記主義）——静的な背景色のみ。
# 同じセルに〓の条件付き書式が乗った場合は条件付き書式が優先して表示される
# （設計 §3 U-04「由来色は〓でなければ」——起きても実害はない多重防御）
FILL_ORIGIN_FALLBACK = PatternFill(start_color="E8F4FA", end_color="E8F4FA", fill_type="solid")


def _normalize_zip(path: Path) -> None:
    """zip エントリの日時・順序と docProps の日時を固定してバイト決定性を得る。

    openpyxl は save 時に dcterms:modified を現在時刻で上書きするため
    （wb.properties への設定は無視される）、ここで固定値へ書き戻す。
    """
    src = zipfile.ZipFile(path, "r")
    entries = {n: src.read(n) for n in src.namelist()}
    src.close()
    core = "docProps/core.xml"
    if core in entries:
        import re
        entries[core] = re.sub(
            rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
            rb"\g<1>2000-01-01T00:00:00Z\g<2>", entries[core])
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for name in sorted(entries):
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            z.writestr(info, entries[name])
    path.write_bytes(buf.getvalue())


def write_xlsx(path: Path, columns: list[str], rows: list[Row],
               unclear_char_level: bool = False) -> None:
    """unclear_char_level=False（既定）では COUNTIF・条件付き書式とも機能追加
    前（完全一致・1本）のまま。True のときだけ「含む」判定・2本目の条件付き
    書式（一部〓用）が有効になる（QA 再判定・T-16 ブロッカーの解消）。
    """
    n_extract = len(columns) - len(META_COLUMNS)
    first = excel_column_letter(len(META_COLUMNS) + 1)          # G
    last = excel_column_letter(len(columns))                    # 220列なら HL

    # Q-MH: 行の値数＝抽出列数は出力の中核不変条件。write_outputs 側の検査
    # （issue #27）は write_xlsx/write_csv を呼ぶ唯一の経路だが、この2関数は
    # 公開関数として直接も呼ばれうる（write_outputs を経由しない呼び出しは
    # 検査を素通りし、列がずれた xlsx が出力されうる）——多重防御としてここでも
    # 検査する。文言は write_outputs と同じ（page_id のみ・値は出さない・§8.1）。
    # 書き込み開始前（write_only の Workbook を作る前）に全行を検査する——
    # 開始後に raise すると、write_only のストリーミング writer が未完了のまま
    # GC 回収され「I/O operation on closed file」の無害だが煩い警告が出るため
    for r in rows:
        if len(r.values) != n_extract:
            raise ValueError(
                f"行の値数({len(r.values)})が抽出列数({n_extract})と一致しない"
                f"（帳票ID: {r.page_id}）。テンプレートと中間データの整合を確認する")

    wb = Workbook(write_only=True)
    wb.properties.created = _FIXED_DT
    wb.properties.modified = _FIXED_DT
    ws = wb.create_sheet("output")
    rng = f"{first}1:{last}{len(rows) + 1}"
    # 欄全体〓（完全一致）: 既存の色 FFF2CC を維持する（見た目を変えない・U-12）。
    # この1本は unclear_char_level に関わらず常に有効（機能追加前からの既存挙動）
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"〓"'], fill=_FILL))
    if unclear_char_level:
        # 一部〓（U-12・#62）: "〓" を含み、かつ長さ2以上。相対参照は範囲の
        # 左上セル（{first}1）を基準に Excel 側が各セルへ自動調整する
        ws.conditional_formatting.add(
            rng, FormulaRule(
                formula=[f'AND(ISNUMBER(FIND("〓",{first}1)),LEN({first}1)>1)'],
                fill=_FILL_PARTIAL))

    def text(v):
        c = WriteOnlyCell(ws, value=v)
        c.number_format = "@"
        if isinstance(v, str) and v:
            # data_type を明示的に文字列へ固定する（issue #34・D-28）。openpyxl の
            # _bind_value は "=" 始まりの値を無条件に数式型へ昇格させ、
            # number_format="@" はこの判定に関与しない。結果、記入値 "=SUM(A1:A9)"
            # が <f> タグとして書かれて Excel で実行され、さらに同じ行の COUNTIF
            # （要確認セル数）を計算グラフへ巻き込んで無言で壊す（実測）。
            # **値そのものは書き換えない**——型宣言を正すだけなので転記主義に反しない
            c.data_type = "s"
        return c

    ws.append([text(c) for c in columns])
    row_no = 1
    for r in rows:
        row_no += 1
        # U-13: unclear_char_level=True のときだけ COUNTIF をワイルドカード化
        # する（完全一致のままだと文字単位〓の部分置換セルを数え損なう・
        # 設計 §8.3）。openpyxl は数式を評価しないため、この文字列一致までが
        # 自動テストで固定できる範囲（T-16 は実機確認・ON 切替の前提条件）。
        # False（既定）では機能追加前と同じ完全一致のまま
        pattern = '"*〓*"' if unclear_char_level else '"〓"'
        formula = WriteOnlyCell(
            ws, value=f'=COUNTIF({first}{row_no}:{last}{row_no},{pattern})')
        meta = [formula, text(r.min_conf), text(r.page_id),
                text(r.source_file), text(str(r.page_no)), text(r.status)]
        origins = r.origins if len(r.origins) == len(r.values) else ("",) * len(r.values)
        body = []
        for v, origin in zip(r.values, origins):
            cell = WriteOnlyCell(ws, value=v) if isinstance(v, int) else text(v)
            if origin == "fallback":
                # 由来色（U-04）: 参照先採用セルへ静的な背景色を付ける。値は
                # 書き換えない。〓の条件付き書式が同じセルに乗る場合は
                # 条件付き書式側が表示上優先される（Excel の描画順）
                cell.fill = FILL_ORIGIN_FALLBACK
            body.append(cell)
        ws.append(meta + body)
    wb.save(path)
    _normalize_zip(path)


def write_csv(path: Path, columns: list[str], rows: list[Row]) -> None:
    # Q-MH: write_xlsx と同じ多重防御（write_outputs を経由しない直接呼び出し
    # でも列ズレを検知する）。文言も揃える
    n_extract = len(columns) - len(META_COLUMNS)
    # N-5: write_xlsx と同じく open() の手前で全行を一括検査する。以前は
    # open() の内側（1行ずつ書きながら）で検査していたため、途中行まで
    # 書いたファイルを残したまま raise していた（header と一部行だけの
    # 中途半端な csv が path に残留する）
    for r in rows:
        if len(r.values) != n_extract:
            raise ValueError(
                f"行の値数({len(r.values)})が抽出列数({n_extract})と一致しない"
                f"（帳票ID: {r.page_id}）。テンプレートと中間データの整合を確認する")
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
        w.writerow(columns)
        for r in rows:
            w.writerow([str(r.unclear_count), r.min_conf, r.page_id,
                        r.source_file, str(r.page_no), r.status]
                       + [str(v) for v in r.values])


def write_columns_txt(path: Path, columns: list[str]) -> None:
    """列名一覧を1行1列名で書く（issue #66 第2弾・05 Q-30③・Could）。

    取り込み先システムとの列構成突き合わせ用。内容は derive_columns の結果
    そのまま（管理6列を含む・列構成の唯一の正・FR-0.1 と同じ思想）——ここで
    列を選び直したり並べ替えたりしない。

    BOM 付き UTF-8・CRLF は write_csv と同じ流儀に揃える（根拠）: 列名は
    日本語を含むため、BOM を落とすと Windows のメモ帳・一部の取り込み先で
    csv と同じ文字化けが起きる。改行を \r\n に固定するのは、os.linesep
    （OS 既定）に委ねると実行環境が変わったときにバイトが変わり、§6.2 の
    決定性（同一テンプレ・同一設定での再出力バイト一致）が環境依存になって
    しまうため——csv 側が lineterminator="\r\n" を明示しているのと同じ理由。
    """
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        f.write("\r\n".join(columns))
        f.write("\r\n")


# Excel が数式として解釈しうる接頭文字（D-28）。@ とタブは実測で無害だったが、
# 検出は無害化ではないので偽陽性コストが低く、Excel のバージョン差・LibreOffice・
# 別の取り込み先まで含めて広く取る
RISKY_PREFIXES = "=+-@\t\r\n"


def scan_risky_prefixes(columns: list[str],
                        rows: list[Row]) -> list[tuple[str, str]]:
    """Excel が数式として解釈しうる接頭文字を持つセルを列挙する（D-28）。

    返すのは (page_id, 列名) のみ——**値は返さない**（設計 §8.1: 記入値を
    ログ・イベントへ出さない方針を、型で守る）。値の書き換えも〓化もしない
    （要件 §5.5 転記主義）。CSV の内容はこの検出の有無で1バイトも変わらない。
    """
    extract_cols = columns[6:]
    hits: list[tuple[str, str]] = []
    for r in rows:
        for name, v in zip(extract_cols, r.values):
            if isinstance(v, str) and v and v[0] in RISKY_PREFIXES:
                hits.append((r.page_id, name))
    return hits


def _rollback(backups: list[tuple[Path, Path]], replaced: list[Path]) -> list[Path]:
    """差し替え途中の失敗を「どちらも変わらない」状態へ戻す（issue #51）。

    退避（.bak）があるものは書き戻す（差し替え済みなら新しい方を捨てる）。
    退避が無いのに差し替わったもの＝元は存在しなかったファイルなので消す。
    戻せなかった分をここで例外にすると元の失敗原因が消えるため、呼び出し元が
    元の例外を __cause__ に付けて投げる形にしてある。

    **戻せなかったパスの一覧を返す**。呼び出し元はこれを見て文言を変える——
    巻き戻しに失敗しているのに「元の内容に戻した」と断言すると、xlsx=新・
    csv=旧という不整合な組み合わせが残っているのに利用者が安心してしまう
    （レビュー4巡目）。
    """
    failed: list[Path] = []
    had_backup: set[Path] = set()
    for final, bak in backups:
        had_backup.add(final)
        if bak.exists():
            try:
                os.replace(bak, final)
            except OSError:
                failed.append(final)
    for final in replaced:
        if final not in had_backup and final.exists():
            try:
                final.unlink()
            except OSError:
                failed.append(final)
    return failed


def write_outputs(
    out_dir: str | Path, timestamp: str, columns: list[str], rows: list[Row],
    unclear_char_level: bool = False,
) -> tuple[Path, Path, list[tuple[str, str]]]:
    """xlsx/csv を書き、危険接頭セルの一覧（page_id, 列名）を併せて返す。

    検出をここに置くのは呼び忘れを構造で防ぐため（issue #27 の行長検査と同じ
    置き方）。検出は出力内容に一切影響しない（D-28）。

    unclear_char_level は write_xlsx の COUNTIF・条件付き書式のゲートへ
    そのまま渡す（既定 False＝機能追加前と同じ完全一致。QA 再判定・T-16）。
    csv 側の要確認セル数は render_rows.build_row が cfg.unclear_char_level を
    見て既に確定させた値（Row.unclear_count）をそのまま書くだけなので、
    ここでの分岐は不要。
    """
    # 行の値数＝抽出列数は出力の中核不変条件。assert（-O で消える）でなく
    # 明示例外で、xlsx/csv 両形式が必ず通るこの一箇所で検査する（issue #27）
    n_extract = len(columns) - 6
    for r in rows:
        if len(r.values) != n_extract:
            raise ValueError(
                f"行の値数({len(r.values)})が抽出列数({n_extract})と一致しない"
                f"（帳票ID: {r.page_id}）。テンプレートと中間データの整合を確認する")
    d = Path(out_dir)
    d.mkdir(parents=True, exist_ok=True)
    xlsx = d / f"output_{timestamp}.xlsx"
    csvp = d / f"output_{timestamp}.csv"
    # 列名一覧の併記（issue #66 第2弾・05 Q-30③・Could）。xlsx/csv と同じ
    # timestamp 系の命名（取り込み先との突き合わせで3ファイルの対応が
    # 一目で分かるように）。アトミック差し替えの対象に加える（下記）——
    # 併記のためだけに #36/#51 の保護から外れた経路を作らない
    colp = d / f"output_{timestamp}_columns.txt"
    # 一時ファイルへ書いてから os.replace で差し替える（issue #36）。
    # 直接書くと、出力を Excel で開いたまま再実行したときに
    # 「open→truncate→ロック違反」の順で **既存の正常な成果物が 0 バイトに
    # 破壊される**（実測: 9246→0）。置換が完了するまで既存ファイルは無傷に保つ
    tmp_x = xlsx.with_suffix(".xlsx.tmp")
    tmp_c = csvp.with_suffix(".csv.tmp")
    tmp_col = colp.with_suffix(".txt.tmp")
    try:
        write_xlsx(tmp_x, columns, rows, unclear_char_level=unclear_char_level)
        write_csv(tmp_c, columns, rows)
        write_columns_txt(tmp_col, columns)
        # 既存を先に退避してから差し替える（レビュー M-5）。片方ずつ replace
        # すると、csv だけ開かれている場合に **xlsx は新・csv は旧**となり
        # §8-12（xlsx↔csv の値一致）が破れる。退避に失敗した時点で中断して
        # 戻すので、「両方更新される」か「どちらも変わらない」かに限定される。
        # 開かれているファイルは rename できない＝これが最も確実な事前確認
        # （追記オープンでの判定は共有モードで通ってしまい役に立たなかった）。
        # columns.txt も同じ3点セットとして扱う（3つとも更新される／3つとも
        # 変わらない、以外の中間状態を作らない）
        backups: list[tuple[Path, Path]] = []
        try:
            for final in (xlsx, csvp, colp):
                if final.exists():
                    bak = final.parent / (final.name + ".bak")
                    if bak.exists():
                        bak.unlink()
                    os.replace(final, bak)
                    backups.append((final, bak))
        except OSError as e:
            for final, bak in backups:          # 退避済みを元へ戻す
                os.replace(bak, final)
            raise PermissionError(
                f"出力ファイルを更新できない（{xlsx.name}・{csvp.name}・"
                f"{colp.name} のいずれかが Excel などで開かれている可能性）。"
                "閉じてからやり直す。既存のファイルは壊れていない") from e
        # 差し替え本体も try で守る（issue #51）。裸で置くと、ここでの失敗
        # （xlsx を Excel で開いたままの PermissionError 等）で finally が tmp を
        # 消す一方、退避した .bak は戻らず削除もされない——手元に残るのは .bak
        # だけで **正規名のファイルが両方消える**。上のコメントが約束する
        # 「両方更新される／どちらも変わらない」は、差し替え中の失敗まで
        # 巻き戻して初めて成立する
        replaced: list[Path] = []
        try:
            os.replace(tmp_x, xlsx)
            replaced.append(xlsx)
            os.replace(tmp_c, csvp)
            replaced.append(csvp)
            os.replace(tmp_col, colp)
            replaced.append(colp)
        except OSError as e:
            not_restored = _rollback(backups, replaced)
            if not_restored:
                names = "・".join(sorted(p.name for p in not_restored))
                tail = (f"巻き戻しにも失敗した（{names}）。"
                        "出力ファイルで内容が食い違っている可能性があるため、"
                        "同フォルダの .bak と突き合わせて確認する")
            else:
                tail = "既存のファイルは元の内容に戻した"
            raise PermissionError(
                f"出力ファイルを更新できない（{xlsx.name}・{csvp.name}・"
                f"{colp.name} のいずれかの差し替えに失敗した。Excel などで"
                f"開かれている可能性）。閉じてからやり直す。{tail}") from e
        for _final, bak in backups:
            try:
                bak.unlink()
            except OSError:
                pass
    finally:
        for t in (tmp_x, tmp_c, tmp_col):
            if t.exists():
                try:
                    t.unlink()
                except OSError:
                    pass
    return xlsx, csvp, scan_risky_prefixes(columns, rows)
