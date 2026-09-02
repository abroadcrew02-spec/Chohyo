"""E2E: 入力フォルダ → run（Replay・課金ゼロ）→ .xlsx/.csv → 配置検証。

保存済み S2 応答と展開済みサンプル画像に依存する（.gitignore 配下・
このマシン限定）。無い環境では skip。

検証は要件 §8 第1層（配置の正しさ）と再現性（§6.2 バイト一致）。
"""
import json
import shutil

import pytest
from openpyxl import load_workbook

from chouhyo_ocr.config import Config
from chouhyo_ocr.paths import app_root
from chouhyo_ocr.pipeline import remap, render, run
from chouhyo_ocr.vision_client import ReplayClient

RESP = app_root() / "workdir" / "s2" / "resp_DOCUMENT_TEXT_DETECTION.json"
PAGE_PNG = app_root() / "workdir" / "pages" / "sample-1.png"
TPL = app_root() / "templates" / "chouhyo-v1.json"

pytestmark = pytest.mark.skipif(
    not (RESP.exists() and PAGE_PNG.exists()), reason="保存済み応答・展開画像が無い環境")


@pytest.fixture(scope="module")
def outputs(tmp_path_factory):
    tmp = tmp_path_factory.mktemp("e2e")
    input_dir = tmp / "input"
    input_dir.mkdir()
    shutil.copy(PAGE_PNG, input_dir / "sample-1.png")

    replay_dir = tmp / "responses"
    replay_dir.mkdir()
    shutil.copy(RESP, replay_dir / "sample-1_p0001.json")

    # 配置（第1層）の検証が目的のため〓閾値を下げる。既定閾値での〓化は
    # test_default_threshold_masks_low_conf で別途検証する。
    cfg = Config(unclear_threshold=0.4,
                 output_dir=str(tmp / "out"), workdir=str(tmp / "wd"), log_dir=str(tmp / "logs"))
    summary = run(input_dir, TPL, cfg, ReplayClient(replay_dir))
    xlsx, csvp, rows = render(TPL, cfg, timestamp="t1")
    return {"cfg": cfg, "summary": summary, "xlsx": xlsx, "csv": csvp, "rows": rows}


def sheet(outputs):
    wb = load_workbook(outputs["xlsx"])
    ws = wb["output"]
    header = [c.value for c in ws[1]]
    data = [c.value for c in ws[2]]
    return header, data


def val(outputs, col):
    header, data = sheet(outputs)
    return data[header.index(col)]


def test_one_page_one_row(outputs):
    assert outputs["summary"].pages == 1
    assert len(outputs["rows"]) == 1
    assert val(outputs, "ステータス") == "正常"
    assert val(outputs, "帳票ID") == "sample-1_p0001"


def test_placement_person(outputs):
    # 郵便番号は独立列（2026-08-31）。主（印字ボックス）が空のサンプルなので
    # 参照先（住所行の先頭ゾーン）から拾われ、住所列には混ざらない
    assert val(outputs, "person_郵便番号1") == "262-0032"
    assert val(outputs, "person_住所1").startswith("千葉県千葉市")
    assert "北海道旭川市" in val(outputs, "person_住所2")


def test_placement_family_date_split(outputs):
    """複合セル「7.7.20」が render で 年/月/日 3列へ分かれる（D-23）。"""
    assert val(outputs, "family_01_生年月日_年") == "7"
    assert val(outputs, "family_01_生年月日_月") == "7"
    assert val(outputs, "family_01_生年月日_日") == "20"


def test_placement_detail_row4_no_cross_contamination(outputs):
    """word 結合で混線していた行が正しい列に入る（symbol 割付の E2E 実証）。"""
    assert val(outputs, "detail_04_来店年月日_年") == "10"
    assert val(outputs, "detail_04_来店年月日_月") == "8"
    assert val(outputs, "detail_04_来店年月日_日") == "31"
    assert val(outputs, "detail_04_金額") == 1000        # 数値型
    assert val(outputs, "detail_04_品目") == "合格祈願"


def test_amount_types(outputs):
    assert val(outputs, "detail_02_金額") == 100000
    assert isinstance(val(outputs, "detail_02_金額"), int)
    assert val(outputs, "detail_03_金額") == 100


def test_empty_rows_are_empty_strings(outputs):
    assert val(outputs, "family_05_続柄") in ("", None)   # 空行 → 空文字
    assert val(outputs, "detail_10_品目") in ("", None)


def test_countif_formula_and_csv_static_agree(outputs):
    header, data = sheet(outputs)
    # 範囲の終端はテンプレート由来の列数から導出する（決め打ち廃止・2026-08-31。
    # 現行 220列なら HL）。リテラルで固定すると列の増減のたびにここが割れる
    from chouhyo_ocr.columns import derive_columns, excel_column_letter
    from chouhyo_ocr.template import load_template
    last = excel_column_letter(len(derive_columns(load_template(TPL))))
    # U-13（5巡目 第2段・#62・2026-08-31）→ QA 再判定（2026-08-31・T-16
    # ブロッカーの解消）で unclear_char_level ゲート化。この fixture の cfg は
    # unclear_char_level=False（既定）のため、COUNTIF は機能追加前と同じ
    # 完全一致のまま（docs/design/chouhyo-ocr/04_unclear_policy.md §8.3/§8.5）。
    # ON 時にワイルドカードへ切り替わることは test_e2e_replay.py の
    # test_char_level_on_produces_partial_unclear_in_real_output 系で別途固定
    assert data[0] == f'=COUNTIF(G2:{last}2,"〓")'
    lines = outputs["csv"].read_text(encoding="utf-8-sig").splitlines()
    csv_row = next(csv_line for csv_line in lines[1:] if csv_line)
    static = int(csv_row.split('","')[0].strip('"'))
    unclear_in_xlsx = sum(1 for v in data[6:] if v == "〓")
    assert static == unclear_in_xlsx == outputs["rows"][0].unclear_count


def test_min_conf_present_and_format(outputs):
    v = val(outputs, "最低信頼度")
    assert v and len(v.split(".")[1]) == 3


def test_era_decisions_are_valid_values(outputs):
    for col in ("person_生年月日_元号", "family_01_生年月日_元号"):
        assert val(outputs, col) in ("昭", "平", "令", "〓")


def test_default_threshold_masks_low_conf(outputs):
    """既定閾値 0.85 では低信頼セル（住所1 conf 0.83）が〓化される（§8-3 実証）。"""
    import dataclasses
    cfg85 = dataclasses.replace(outputs["cfg"], unclear_threshold=0.85)
    xlsx, _csv, _rows = render(TPL, cfg85, timestamp="t85")
    wb = load_workbook(xlsx)
    ws = wb["output"]
    header = [c.value for c in ws[1]]
    data = [c.value for c in ws[2]]
    assert data[header.index("person_住所1")] == "〓"   # 低信頼 symbol を含むため
    assert data[header.index("detail_01_品目")] == "家内安全"  # 高信頼は残る


def test_render_determinism_bytes(outputs):
    """同一中間データ・同一設定 → xlsx/csv ともバイト一致（要件 §6.2）。"""
    cfg = outputs["cfg"]
    x1, c1, _ = render(TPL, cfg, timestamp="d1")
    x2, c2, _ = render(TPL, cfg, timestamp="d2")
    assert x1.read_bytes() == x2.read_bytes()
    assert c1.read_bytes() == c2.read_bytes()


def test_remap_is_idempotent_without_template_change(outputs):
    cfg = outputs["cfg"]
    before = render(TPL, cfg, timestamp="r1")[0].read_bytes()
    assert remap(TPL, cfg) == 1
    after = render(TPL, cfg, timestamp="r2")[0].read_bytes()
    assert before == after


def test_remap_repopulates_extras_and_reports_fallback_counts(outputs):
    """remap() も run() と同じ変換（_extras_rows）で char_confs/origin を
    作り直すこと（設計 §12「remap にも同じ変更が要る」）。直し忘れると
    remap のたびに origin が既定値 '' へ巻き戻り、由来色が消える。
    """
    from chouhyo_ocr.store import Store
    from pathlib import Path
    cfg = outputs["cfg"]
    events = []
    n = remap(TPL, cfg, progress=events.append)
    assert n == 1
    remap_ev = next(e for e in events if e.get("event") == "remap_summary")
    assert remap_ev["pages"] == 1
    assert remap_ev["fallback_used"] == 2
    assert remap_ev["fallback_discarded"] == 0
    assert remap_ev["carve_hole"] == 0

    store = Store(Path(cfg.workdir) / "intermediate.sqlite")
    try:
        pid = store.pages()[0]["page_id"]
        extras = store.cell_extras(pid)
    finally:
        store.close()
    assert extras["person_郵便番号1"][1] == "fallback"
    assert extras["person_郵便番号1"][0] != ""  # char_confs も再構築されている


# ---------- 5巡目 第2段: pipeline.py 配線の実効性検証（2026-08-31） ----------
#
# run() は char_confs/origin を常に store.cell_extras() へ保存する（設定に
# 関わらず）。unclear_char_level・unclear_threshold は render 段だけの判断
# （設計 §14 不変条件3）なので、既に run 済みの outputs fixture を使い回して
# render() を条件違いで呼び直すだけで検証できる（test_default_threshold_
# masks_low_conf と同じ手筋・余分な run を増やさない）。

def _render_with(outputs, ts, **overrides):
    import dataclasses
    cfg = dataclasses.replace(outputs["cfg"], **overrides)
    xlsx, csvp, rows = render(TPL, cfg, timestamp=ts)
    wb = load_workbook(xlsx)
    ws = wb["output"]
    header = [c.value for c in ws[1]]
    data = [c.value for c in ws[2]]
    return {"xlsx": xlsx, "csv": csvp, "rows": rows, "header": header,
           "data": data, "ws": ws}


def test_char_level_on_produces_partial_unclear_in_real_output(outputs):
    """配線後の実効性検証（コーディネーター指示 5.）: unclear_char_level ON で
    render すると、実データ（S2 replay）に含まれる混在信頼度セルが本当に
    部分〓（〓が値の一部にだけ現れる形）になる。閾値は既定 0.85 に揃える（配置検証用に
    下げた fixture の 0.4 ではなく）。
    """
    on = _render_with(outputs, "char_on", unclear_threshold=0.85,
                      unclear_char_level=True)
    off = _render_with(outputs, "char_off", unclear_threshold=0.85,
                       unclear_char_level=False)
    name_on = on["data"][on["header"].index("family_01_氏名")]
    name_off = off["data"][off["header"].index("family_01_氏名")]
    assert name_off == "〓"          # OFF: 従来どおり欄全体〓
    assert name_on not in ("", "〓")  # ON: 単独の〓ではなく実在の部分値
    assert "〓" in name_on
    assert len(name_on) >= 2
    # 畳み込み不変条件（設計 §14 不変条件1）: 全文字が〓の複数文字列
    # （"〓〓〓" 等）を作らない。全文字未満なら1文字の "〓" へ畳む
    assert all(not (isinstance(v, str) and len(v) >= 2 and set(v) == {"〓"})
              for v in on["data"][6:])


def test_char_level_off_matches_pre_feature_output(outputs):
    """OFF 回帰: 文字単位〓が無効なとき、抽出列の値は完全一致のみ〓になり、
    ON/OFF で意味のある差が「一部だけ低信頼な既知セル」以外に出ないこと。
    """
    on = _render_with(outputs, "regress_on", unclear_threshold=0.85,
                      unclear_char_level=True)
    off = _render_with(outputs, "regress_off", unclear_threshold=0.85,
                       unclear_char_level=False)
    # 抽出列のみを比較する（META_COLUMNS の6列は除く）。管理列は
    # unclear_char_level 自体で内容が変わる（要確認セル数の COUNTIF 式・
    # 最低信頼度）ため、ここでの「抽出値の回帰」比較には含めない
    from chouhyo_ocr.columns import META_COLUMNS
    n_meta = len(META_COLUMNS)
    diffs = [(off["header"][i], off["data"][i], on["data"][i])
             for i in range(n_meta, len(off["header"]))
             if off["data"][i] != on["data"][i]]
    # 差分は「OFF側が欄全体〓で、ON側がそれを部分置換した値」のケースのみ
    for _name, off_v, on_v in diffs:
        assert off_v == "〓", f"OFF側は欄全体〓のはずが: {off_v!r}"
        assert on_v != "〓" and "〓" in str(on_v)
    assert len(diffs) > 0, "実データにOFF/ONの差が一切出ない＝検証になっていない"


def test_origin_fill_present_regardless_of_char_level_flag(outputs):
    """由来色（U-04）は unclear_char_level の設定に関わらず一貫して出る
    （2つの機能は独立——origin は run 時に常に保存され、char-level は render
    時のみに効く）。
    """
    on = _render_with(outputs, "origin_on", unclear_threshold=0.85,
                      unclear_char_level=True)
    off = _render_with(outputs, "origin_off", unclear_threshold=0.85,
                       unclear_char_level=False)
    idx = on["header"].index("person_郵便番号1")
    cell_on = list(on["ws"].rows)[1][idx]
    cell_off = list(off["ws"].rows)[1][idx]
    assert cell_on.fill.start_color.rgb == "00E8F4FA"
    assert cell_off.fill.start_color.rgb == "00E8F4FA"


def test_run_summary_event_reports_fallback_counts(tmp_path):
    """配線の実効性検証: run() の summary イベントに fallback_used/
    fallback_discarded/carve_hole が実データで載ること（設計 §10.3）。
    """
    input_dir = tmp_path / "input"; input_dir.mkdir()
    shutil.copy(PAGE_PNG, input_dir / "sample-1.png")
    replay_dir = tmp_path / "responses"; replay_dir.mkdir()
    shutil.copy(RESP, replay_dir / "sample-1_p0001.json")
    cfg = Config(unclear_threshold=0.85,
                output_dir=str(tmp_path / "out"), workdir=str(tmp_path / "wd"),
                log_dir=str(tmp_path / "logs"))
    events = []
    run(input_dir, TPL, cfg, ReplayClient(replay_dir), events.append)
    summary_ev = next(e for e in events if e.get("event") == "summary")
    # 実データ: 郵便番号1・2 とも主が空で参照先採用（§1.3 の実測どおり）
    assert summary_ev["fallback_used"] == 2
    assert summary_ev["fallback_discarded"] == 0
    assert summary_ev["carve_hole"] == 0
    page_ev = next(e for e in events if e.get("event") == "page" and e.get("status") == "done")
    assert page_ev["fallback_used"] == 2


def test_second_run_reports_reused_pages_without_api_calls(tmp_path):
    """コーディネーター指示（2026-09-02）: 実機の通し確認で見つかった詰まり所。

    同じ workdir で2回 run すると、2回目は前回 done になったページが todo から
    外れて無言で再利用される（API へ送らない）。summary の reused_pages で
    件数が読め、api_calls は 0 のまま——「なぜ送信されていないか」がサマリだけで
    分かることを確認する。
    """
    input_dir = tmp_path / "input"; input_dir.mkdir()
    shutil.copy(PAGE_PNG, input_dir / "sample-1.png")
    replay_dir = tmp_path / "responses"; replay_dir.mkdir()
    shutil.copy(RESP, replay_dir / "sample-1_p0001.json")
    cfg = Config(unclear_threshold=0.85,
                output_dir=str(tmp_path / "out"), workdir=str(tmp_path / "wd"),
                log_dir=str(tmp_path / "logs"))

    events1 = []
    run(input_dir, TPL, cfg, ReplayClient(replay_dir), events1.append)
    summary_ev1 = next(e for e in events1 if e.get("event") == "summary")
    # 1回目は新規処理なので送信あり・再利用なし
    assert summary_ev1["reused_pages"] == 0
    assert summary_ev1["api_calls"] == 1

    events2 = []
    run(input_dir, TPL, cfg, ReplayClient(replay_dir), events2.append)
    summary_ev2 = next(e for e in events2 if e.get("event") == "summary")
    assert summary_ev2["reused_pages"] == summary_ev2["pages"] == 1
    assert summary_ev2["api_calls"] == 0
    reused_page_ev = next(e for e in events2
                          if e.get("event") == "page" and e.get("reused") is True)
    assert reused_page_ev["status"] == "done"
    assert reused_page_ev["page_id"] == "sample-1_p0001"
