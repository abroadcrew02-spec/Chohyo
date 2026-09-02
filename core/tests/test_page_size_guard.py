"""Q-H1: 入力ページの寸法検査（align.page_size_verdict・align_page への配線）。

着手前実測（2026-09-02）: 現行コード（検査追加前）に対し、出荷サンプル
（testdata/formB/formB-1.png・1800x1200）を縦5%伸ばした入力（1800x1260）を
`align_page` にそのまま通したところ、例外を一切出さずに成功していた
（faces[0].angle=0.0・dx=0・dy=0・matched=10）。`align_page` 冒頭の
`page_img.resize((W, H))` が無検証で歪みを消してしまうため、後段の平行移動
推定が「一致」と誤判定し、歪んだ画像がそのまま送信対象になっていた。

  $ ../.venv/Scripts/python.exe -c "from PIL import Image; from chouhyo_ocr.template
    import load_template; from chouhyo_ocr.align import align_page; ..."
  → SUCCESS: no exception raised / face front angle 0.0 dx 0 dy 0 matched 10

このファイルは (1) page_size_verdict の純関数テスト（合成寸法・実画像不要）
(2) run 経由の統合テスト（出荷サンプルを5%伸ばした入力→様式不一致・
行数維持・api_calls=0）を確認する。
"""
from types import SimpleNamespace

import pytest

from chouhyo_ocr.align import PageSizeMismatch, align_page, page_size_verdict

# ---------- unit: page_size_verdict（合成寸法・整数px境界を厳密に作る） ----------
# 実テンプレート（2490x3510）は px 粒度が粗く 0.99%/1.01% ちょうどを作れない
# ため、境界値がきれいに整数化できる合成テンプレートを使う（image_size のみ
# 参照する純関数なので SimpleNamespace で足りる）。

_SQUARE = SimpleNamespace(image_size=(10000, 10000))
_PORTRAIT = SimpleNamespace(image_size=(7000, 10000))


def test_within_0_99_percent_is_ok():
    """縦を+0.99%だけ伸ばした入力（9901x10000 相当）はアスペクト比が通る。"""
    assert page_size_verdict((9901, 10000), _SQUARE) is None


def test_over_1_01_percent_is_rejected():
    """縦横比の相対差が1.01%になる入力は拒否される。"""
    assert page_size_verdict((9899, 10000), _SQUARE) == "aspect_mismatch"


def test_swapped_orientation_is_rejected():
    """縦横反転（portrait のテンプレートに landscape の入力）は拒否される。"""
    assert page_size_verdict((10000, 7000), _PORTRAIT) == "aspect_mismatch"


def test_uniform_double_scale_is_ok():
    """等比2倍（dpi違いに相当）はアスペクト比が変わらないので通る。"""
    assert page_size_verdict((14000, 20000), _PORTRAIT) is None


def test_exact_match_is_ok():
    assert page_size_verdict((7000, 10000), _PORTRAIT) is None


def test_zero_dimension_is_rejected():
    """0px は比の計算自体が無意味なので拒否側へ倒す（ZeroDivisionError にしない）。"""
    assert page_size_verdict((0, 10000), _PORTRAIT) == "aspect_mismatch"


# ---------- align_page への配線: 不一致で PageSizeMismatch（AlignError のサブクラス） ----------

def test_align_page_raises_page_size_mismatch_on_distorted_input():
    """5%縦伸ばしの入力は align_page が resize する前に弾かれる。

    着手前実測（本ファイル冒頭）では同じ入力が無検証で「成功」していた。
    """
    pytest.importorskip("numpy")
    from PIL import Image

    from chouhyo_ocr.paths import app_root
    from chouhyo_ocr.template import load_template

    tpl_path = app_root() / "testdata" / "formB" / "formB-v1.json"
    png_path = app_root() / "testdata" / "formB" / "formB-1.png"
    if not (tpl_path.exists() and png_path.exists()):
        pytest.skip("formB の検証資産が無い環境")
    template = load_template(tpl_path)
    im = Image.open(png_path).convert("RGB")
    w, h = im.size
    distorted = im.resize((w, int(round(h * 1.05))))
    with pytest.raises(PageSizeMismatch):
        align_page(distorted, template)

    # 対照: 無加工の同じ画像は従来どおり通る（過剰検出でないことの確認）
    faces, _composite = align_page(im, template)
    assert len(faces) >= 1


# ---------- 統合: run 経由で「様式不一致」に倒れ、行数維持・送信ゼロを確認 ----------

def test_run_with_distorted_input_reports_format_mismatch_without_sending(tmp_path):
    """縦5%伸ばした出荷サンプルを run に通すと、送信前（api_calls=0）に
    「様式不一致」へ倒れ、TR-G2（入力ページ数＝出力行数）どおり行は1行残る。

    cli.cmd_run の exit 判定（全滅なら exit 1）の母集団に format_mismatch を
    含めたことも、この経路の summary.format_mismatch==1 で裏付ける。
    """
    from PIL import Image
    from openpyxl import load_workbook

    from chouhyo_ocr.config import Config
    from chouhyo_ocr.paths import app_root
    from chouhyo_ocr.pipeline import render, run
    from chouhyo_ocr.vision_client import ReplayClient

    page_png = app_root() / "workdir" / "pages" / "sample-1.png"
    tpl = app_root() / "templates" / "chouhyo-v1.json"
    if not page_png.exists():
        pytest.skip("展開済みサンプル画像が無い環境")

    input_dir = tmp_path / "input"
    input_dir.mkdir()
    im = Image.open(page_png).convert("RGB")
    w, h = im.size
    im.resize((w, int(round(h * 1.05)))).save(input_dir / "sample-1.png")

    replay_dir = tmp_path / "responses"  # 送信に到達しないため中身は不要
    cfg = Config(output_dir=str(tmp_path / "out"), workdir=str(tmp_path / "wd"),
                log_dir=str(tmp_path / "logs"))
    summary = run(input_dir, tpl, cfg, ReplayClient(replay_dir))

    assert summary.api_calls == 0          # 送信前に落ちている（課金なし）
    assert summary.format_mismatch == 1    # PageSizeMismatch 起因を計上
    assert summary.align_failed == 0       # 「位置合わせ失敗」ではなく様式不一致側
    assert summary.rows == 1               # TR-G2: 入力1ページ→出力1行を維持

    _xlsx, _csvp, rows = render(tpl, cfg, timestamp="t1")
    assert len(rows) == 1
    assert rows[0].status == "様式不一致"

    wb = load_workbook(_xlsx)
    ws = wb["output"]
    header = [c.value for c in ws[1]]
    data = [c.value for c in ws[2]]
    assert data[header.index("ステータス")] == "様式不一致"
