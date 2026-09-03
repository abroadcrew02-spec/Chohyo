"""レビュー4巡目・入出力まわりの修正（issue #51 / #47 / #52 M-4）。

ここで固定する不変条件:
1. 出力の差し替えが途中で失敗しても、正規名の xlsx/csv は「両方とも前回の
   内容のまま」になる（#51。旧実装は .bak だけ残して両方消しうる）
2. 月次上限は refused 契約（JSON Lines ＋ exit 0）で伝える（#47）
3. 送信リトライは1回ごとに月次カウンタへ乗る（#52 M-4）
"""
import csv as csvmod
import json
import os

import pytest
from openpyxl import load_workbook

from chouhyo_ocr import api_budget
from chouhyo_ocr.pipeline_errors import OperationRefused
from chouhyo_ocr.render_out import write_outputs
from chouhyo_ocr.render_rows import Row

COLS = ["要確認セル数", "最低信頼度", "帳票ID", "入力ファイル名",
        "ページ番号", "ステータス", "a", "b"]


def _row(first: str) -> Row:
    return Row(page_id="p1", source_file="s.pdf", page_no=1, status="正常",
               values=[first, "x"], unclear_count=0, min_conf="0.9")


def _read_back(xlsx, csvp):
    """xlsx/csv の 1 行目・抽出1列目の値を返す（新旧の判別用）。"""
    x = load_workbook(xlsx)["output"].cell(row=2, column=7).value
    with open(csvp, encoding="utf-8-sig", newline="") as f:
        c = list(csvmod.reader(f))[1][6]
    return x, c


def _fail_replace_of(monkeypatch, tmp_suffix: str):
    """一時ファイル→正規名の os.replace だけを失敗させる。

    退避（正規名→.bak）と巻き戻し（.bak→正規名）は素通しにしないと、
    「巻き戻しも道連れで壊れた」のか「巻き戻しが効いた」のか区別できない。
    """
    real = os.replace

    def fake(src, dst, *a, **kw):
        if str(src).endswith(tmp_suffix):
            raise PermissionError(f"別プロセスが使用中（テスト擬似）: {dst}")
        return real(src, dst, *a, **kw)

    monkeypatch.setattr(os, "replace", fake)


# ---------- #51: 差し替えの失敗で正規名のファイルを消さない ----------

@pytest.mark.parametrize("target,label", [(".csv.tmp", "csv"), (".xlsx.tmp", "xlsx")])
def test_replace_failure_leaves_both_outputs_unchanged(tmp_path, monkeypatch,
                                                       target, label):
    """差し替えの片方が失敗しても「どちらも変わらない」に着地する。

    旧実装では os.replace が try の外にあり、失敗すると finally が tmp を消す
    一方で退避した .bak は戻らず、正規名の xlsx/csv が両方消えていた。
    """
    xlsx, csvp, _r = write_outputs(tmp_path, "t", COLS, [_row("OLD")])
    before = (xlsx.read_bytes(), csvp.read_bytes())

    _fail_replace_of(monkeypatch, target)
    with pytest.raises(OSError):            # 握りつぶさず呼び出し元へ伝える
        write_outputs(tmp_path, "t", COLS, [_row("NEW")])

    assert xlsx.exists(), f"{label} 失敗時に xlsx が消えた"
    assert csvp.exists(), f"{label} 失敗時に csv が消えた"
    assert (xlsx.read_bytes(), csvp.read_bytes()) == before, "旧内容に戻っていない"
    assert _read_back(xlsx, csvp) == ("OLD", "OLD")
    assert not list(tmp_path.glob("*.tmp")), "一時ファイルが残った"
    assert not list(tmp_path.glob("*.bak")), "退避ファイルが残った"


def test_replace_failure_on_first_run_leaves_no_partial_output(tmp_path, monkeypatch):
    """初回（退避対象が無い）実行で csv の差し替えが失敗したら xlsx も残さない。

    退避が無い＝元は存在しなかったので、xlsx だけ置き去りにすると
    「xlsx はあるのに csv が無い」半端な状態になる。
    """
    _fail_replace_of(monkeypatch, ".csv.tmp")
    with pytest.raises(OSError):
        write_outputs(tmp_path, "t", COLS, [_row("NEW")])
    assert list(tmp_path.glob("output_t.*")) == []


def test_normal_write_is_unaffected(tmp_path):
    """正常系: 2回目の書き込みで両方が新しい内容へ更新される。"""
    xlsx, csvp, _r = write_outputs(tmp_path, "t", COLS, [_row("OLD")])
    xlsx2, csvp2, _r2 = write_outputs(tmp_path, "t", COLS, [_row("NEW")])
    assert (xlsx2, csvp2) == (xlsx, csvp)
    assert _read_back(xlsx, csvp) == ("NEW", "NEW")
    assert not list(tmp_path.glob("*.bak"))
    assert not list(tmp_path.glob("*.tmp"))


# ---------- #47: 月次上限は refused 契約で伝える ----------

@pytest.fixture()
def isolated_counter(tmp_path, monkeypatch):
    """カウンタを一時ディレクトリへ隔離する（実カウンタを汚さない）。"""
    monkeypatch.setenv("CHOUHYO_USAGE_DIR_FOR_TESTS", str(tmp_path))
    return tmp_path


def test_budget_error_is_operation_refused(isolated_counter):
    """業務的拒否として扱う（cli.main の except OperationRefused に乗る）。"""
    with pytest.raises(api_budget.BudgetExceededError) as ei:
        api_budget.check_and_count(1, cap=0)
    e = ei.value
    assert isinstance(e, OperationRefused)
    assert e.hint and "api_monthly_cap" in e.hint
    # 「送信していない」ことと当月使用数・上限が本文から読み取れる
    assert "送信していない" in str(e)
    assert "当月 0" in str(e) and "上限 0" in str(e)


def test_cli_run_exits_zero_with_refused_event(tmp_path, monkeypatch, capsys,
                                               isolated_counter):
    """CLI 経路の回帰: 上限到達で refused ＋ exit 0（1 だと GUI が誤案内する）。

    実 API は叩けないので、pipeline.run の中で本物の check_and_count を
    上限超過で呼ばせ、生成された例外がそのまま CLI の契約に乗るかを見る。
    """
    from chouhyo_ocr import cli, pipeline

    cfg_path = tmp_path / "config.json"
    cfg_path.write_text(json.dumps({
        "output_dir": str(tmp_path / "out"), "workdir": str(tmp_path / "wd"),
        "log_dir": str(tmp_path / "logs"), "api_monthly_cap": 2}),
        encoding="utf-8")
    api_budget.check_and_count(2, cap=2)          # 枠を使い切らせる

    def fake_run(*a, **kw):
        api_budget.check_and_count(1, 2)          # ここで上限に当たる
        raise AssertionError("上限超過なのに送信処理へ進んだ")

    monkeypatch.setattr(pipeline, "run", fake_run)
    replay = tmp_path / "resp"
    replay.mkdir()

    rc = cli.main(["--config", str(cfg_path), "run", "--input", str(tmp_path),
                   "--replay", str(replay)])

    assert rc == 0, "exit 1 だと GUI が「再度押すと続きから」と誤案内する"
    events = [json.loads(line) for line in capsys.readouterr().out.splitlines()
              if line.strip()]
    refused = [e for e in events if e.get("event") == "refused"]
    assert len(refused) == 1, events
    assert refused[0]["ok"] is False
    assert "送信していない" in refused[0]["error"]
    assert "当月 2" in refused[0]["error"] and "上限 2" in refused[0]["error"]
    assert "api_monthly_cap" in refused[0]["hint"]


# ---------- #52 M-4: リトライを月次カウンタへ乗せる ----------

class _FakeVisionModule:
    def Image(self, content):        # noqa: N802
        return content

    def ImageContext(self, language_hints):   # noqa: N802
        return language_hints


class _ErrorResponse:
    """resp.error.message が立った応答（_NON_RETRYABLE に該当せず再試行される）。"""

    class _Error:
        message = "backend error"
        code = 13

    error = _Error()


class _DeterministicErrorResponse:
    """決定的な応答内エラー（PERMISSION_DENIED）。再送しても結果は変わらない。"""

    class _Error:
        message = "The caller does not have permission"
        code = 7

    error = _Error()


class _CountingClient:
    def __init__(self, response=None, raises=None):
        self.calls = 0
        self._response = response or _ErrorResponse()
        self._raises = raises

    def document_text_detection(self, **kw):
        self.calls += 1
        if self._raises is not None:
            raise self._raises
        return self._response


def _client_with(cap: int, response=None, raises=None):
    from chouhyo_ocr import vision_client
    c = vision_client.RealVisionClient.__new__(vision_client.RealVisionClient)
    c._vision = _FakeVisionModule()
    c._to_dict = lambda x: {}
    c.client = _CountingClient(response, raises)
    c.monthly_cap = cap
    c.BACKOFF_INITIAL = 0        # テストを待たせない（バックオフ自体は別責務）
    return c


def test_each_retry_counts_one_unit(isolated_counter):
    """再試行 N 回なら月次カウンタも N 加算される（旧実装は常に 1）。"""
    from chouhyo_ocr import vision_client

    c = _client_with(cap=100)
    with pytest.raises(vision_client.SendError):
        c.annotate(b"png", "p1")
    assert c.client.calls == vision_client.RealVisionClient.MAX_ATTEMPTS
    assert api_budget.used_this_month() == c.client.calls


def test_retry_stops_at_cap_without_exceeding(isolated_counter):
    """上限を跨ぐリトライは打ち切る（上限を超えた送信を発生させない）。"""
    c = _client_with(cap=3)
    with pytest.raises(api_budget.BudgetExceededError):
        c.annotate(b"png", "p1")
    assert c.client.calls == 3, "上限を超えて送信した"
    assert api_budget.used_this_month() == 3


# ---------- #99: 応答内エラーの種別で再送可否を分ける ----------

def test_deterministic_api_error_stops_after_one_attempt(isolated_counter):
    """決定的な応答内エラー（PERMISSION_DENIED）は再送しない（issue #99）。

    旧実装は種別を見ずに MAX_ATTEMPTS 回まで再送し、ページ1枚あたり
    5ユニットと約15秒を確定で失っていた。カウンタ消費は従来どおり
    「試行の直前に1」なので、1ユニットで止まる。
    """
    from chouhyo_ocr import vision_client

    c = _client_with(cap=100, response=_DeterministicErrorResponse())
    with pytest.raises(vision_client.SendError) as ei:
        c.annotate(b"png", "p1")
    assert c.client.calls == 1, "決定的エラーで再送した"
    assert api_budget.used_this_month() == 1
    assert ei.value.code == "SEND_API_7"


def test_transient_api_error_still_retries(isolated_counter):
    """一時エラー（code=13 INTERNAL）は従来どおり最大 MAX_ATTEMPTS 回。"""
    from chouhyo_ocr import vision_client

    c = _client_with(cap=100)          # _ErrorResponse は code=13
    with pytest.raises(vision_client.SendError):
        c.annotate(b"png", "p1")
    assert c.client.calls == vision_client.RealVisionClient.MAX_ATTEMPTS


class _RenamedPermissionDenied(Exception):
    """名前の集合に載っていないが gRPC ステータスは決定的な例外（issue #99）。"""

    class _Status:
        value = (7, "permission denied")

    grpc_status_code = _Status()


def test_exception_is_classified_by_grpc_status_when_name_is_unknown(isolated_counter):
    """例外側は名前照合に加えて grpc_status_code でも判定する（二段構え）。"""
    from chouhyo_ocr import vision_client

    c = _client_with(cap=100, raises=_RenamedPermissionDenied("denied"))
    with pytest.raises(vision_client.SendError) as ei:
        c.annotate(b"png", "p1")
    assert c.client.calls == 1, "決定的エラーで再送した"
    assert api_budget.used_this_month() == 1
    assert ei.value.code == "SEND__RenamedPermissionDenied"


class _RenamedInternalError(Exception):
    class _Status:
        value = (13, "internal")

    grpc_status_code = _Status()


def test_transient_exception_still_retries(isolated_counter):
    """一時的な gRPC ステータス（13 INTERNAL）は従来どおり再送する。"""
    from chouhyo_ocr import vision_client

    c = _client_with(cap=100, raises=_RenamedInternalError("boom"))
    with pytest.raises(vision_client.SendError):
        c.annotate(b"png", "p1")
    assert c.client.calls == vision_client.RealVisionClient.MAX_ATTEMPTS


def test_status_number_reads_real_google_exceptions():
    """google.api_core 例外の grpc_status_code の形を実物で固定する（issue #99）。

    `.value` が `(番号, 説明)` のタプルで番号自体も IntEnum、という前提で
    `_status_number` を書いている。ライブラリ側が形を変えたらここで落ちる。
    """
    exceptions = pytest.importorskip("google.api_core.exceptions")
    from chouhyo_ocr.vision_client import _status_number

    assert _status_number(exceptions.PermissionDenied("x").grpc_status_code) == 7
    assert _status_number(exceptions.Unauthenticated("x").grpc_status_code) == 16
    assert _status_number(exceptions.InternalServerError("x").grpc_status_code) == 13
    assert _status_number(None) is None
