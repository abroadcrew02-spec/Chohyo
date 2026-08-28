"""異体字・サロゲートペアの保持（要件 §6.4: 﨑・髙 等を含む環境で実行できること）。

転記主義のツールが常用字体へ丸めないこと・xlsx/csv の往復で文字が
壊れないことを、出力の実ファイルで確認する。
"""
import pytest
from openpyxl import load_workbook

from chouhyo_ocr.config import Config
from chouhyo_ocr.paths import app_root
from chouhyo_ocr.render_out import write_outputs
from chouhyo_ocr.render_rows import build_row
from chouhyo_ocr.template import load_template
from chouhyo_ocr.columns import derive_columns

TPL = app_root() / "templates" / "chouhyo-v1.json"
EXOTIC = "山﨑髙𠮷"  # 﨑(U+FA11)・髙(U+9AD9)・𠮷(U+20BB7 サロゲートペア)


@pytest.fixture(scope="module")
def template():
    return load_template(TPL)


def test_exotic_chars_survive_xlsx_and_csv(template, tmp_path):
    cells = {c.field_id: ("", None, c.kind, False) for c in template.cells}
    cells["person_氏名"] = (EXOTIC, 0.95, "text", False)
    cells["family_01_氏名"] = ("渡邊𠮷子", 0.95, "text", False)
    page = {"page_id": "p_﨑0001", "source_file": "帳票_髙橋.png", "page_no": 1,
            "status": "", "unassigned_below_table": 0}
    row = build_row(template, page, cells, {}, Config(unclear_threshold=0.5))
    cols = derive_columns(template)
    xlsx, csvp, _risky = write_outputs(tmp_path, "t", cols, [row])

    ws = load_workbook(xlsx)["output"]
    header = [c.value for c in ws[1]]
    data = [c.value for c in ws[2]]
    assert data[header.index("person_氏名")] == EXOTIC          # 常用字体へ丸めない
    assert data[header.index("family_01_氏名")] == "渡邊𠮷子"
    assert data[header.index("帳票ID")] == "p_﨑0001"           # 日本語ファイル名由来のID
    assert data[header.index("入力ファイル名")] == "帳票_髙橋.png"

    text = csvp.read_text(encoding="utf-8-sig")
    assert EXOTIC in text and "渡邊𠮷子" in text and "帳票_髙橋.png" in text
