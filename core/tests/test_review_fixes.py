"""品質レビュー（2026-08-28）HIGH 指摘の再発防止テスト（issue #11/#13/#14）。

- #11: 金額正規化は normalize 属性で発火し、列名に依存しない（要件 §5.2）
- #13: glob メタ文字を含むファイル名の PDF が展開できる
- #14: config.json の不正値・未知キーは ConfigError で明示的に拒否される

#12/#15（エディタ）は TypeScript 側のため tsc ＋実機で確認する。
"""
import json

import pytest

from chouhyo_ocr import cli
from chouhyo_ocr.config import Config, ConfigError, load_config
from chouhyo_ocr.paths import app_root
from chouhyo_ocr.render_rows import UNCLEAR, build_row
from chouhyo_ocr.template import load_template

TPL = app_root() / "templates" / "chouhyo-v1.json"
CFG = Config(unclear_threshold=0.85, era_threshold=0.06)
RESP = (app_root() / "core" / "workdir" / "responses"
        / "帳票抽出検証用2026-08-24_p0001.json")
PAGE_PNG_EXISTS = (app_root() / "testdata" / "local" / "pages" / "sample-1.png").exists()


# ---------- #11: 金額正規化の発火条件 ----------

def _template_with(tmp_path, rename_to=None, drop_normalize=False):
    t = json.loads(TPL.read_text(encoding="utf-8"))
    for face in t["faces"]:
        for tb in face.get("tables", []):
            for c in tb["columns"]:
                if c["name"] == "金額":
                    if drop_normalize:
                        c.pop("normalize", None)
                    if rename_to:
                        c["name"] = rename_to
    p = tmp_path / "t.json"
    p.write_text(json.dumps(t, ensure_ascii=False), encoding="utf-8")
    return load_template(p)


def _row_value(template, cells, colname):
    row = build_row(
        template,
        {"page_id": "p", "source_file": "s.png", "page_no": 1,
         "status": "", "unassigned_below_table": 0},
        cells, {}, CFG)
    cols = [oc for c in template.cells for oc in c.output_columns()]
    return row.values[cols.index(colname)]


def test_amount_fires_on_normalize_attr_not_column_name(tmp_path):
    """列名を「支払金額」へ改名しても normalize: amount があれば整数化される。"""
    template = _template_with(tmp_path, rename_to="支払金額")
    cells = {c.field_id: ("", None, c.kind, False) for c in template.cells}
    cells["detail_01_支払金額"] = ("1,234", 0.95, "text", False)
    assert _row_value(template, cells, "detail_01_支払金額") == 1234


def test_amount_does_not_fire_without_normalize_attr(tmp_path):
    """normalize が無ければ、列名が「金額」でも正規化しない（属性だけが契約）。"""
    template = _template_with(tmp_path, drop_normalize=True)
    cells = {c.field_id: ("", None, c.kind, False) for c in template.cells}
    cells["detail_01_金額"] = ("1,234", 0.95, "text", False)
    assert _row_value(template, cells, "detail_01_金額") == "1,234"


def test_shipped_template_declares_amount_normalize():
    """同梱テンプレートの金額列は normalize: amount を宣言している。"""
    template = load_template(TPL)
    amount_cells = [c for c in template.cells if c.normalize == "amount"]
    assert len(amount_cells) == 28  # 明細28行 × 金額1列
    assert all(c.field_id.endswith("金額") for c in amount_cells)


def test_normalize_ignored_on_choice_and_subfields(tmp_path):
    """choice・subfields 付きセルの normalize は読み込み時に落とす（再レビュー D-5）。

    エディタで種類を切り替えると隠れた normalize が JSON に残りうる。残っていても
    無害（CellSpec に載らない）で、validate_v1 の件数 28 も汚染されないこと。
    """
    from chouhyo_ocr.columns import validate_v1
    t = json.loads(TPL.read_text(encoding="utf-8"))
    for face in t["faces"]:
        for tb in face.get("tables", []):
            for c in tb["columns"]:
                if c["kind"] == "choice" or c.get("subfields"):
                    c["normalize"] = "amount"  # エディタの隠れ値を装う
    p = tmp_path / "t.json"
    p.write_text(json.dumps(t, ensure_ascii=False), encoding="utf-8")
    template = load_template(p)
    polluted = [c for c in template.cells
                if c.normalize and (c.kind == "choice" or c.subfields)]
    assert polluted == []
    assert len(validate_v1(template)) == 220  # 列数が汚染されない
    from chouhyo_ocr.columns import amount_cell_count
    assert amount_cell_count(template) == 28  # 金額カウントも汚染されない


def test_amount_count_is_reported_not_rejected(tmp_path):
    """normalize の喪失は拒否ではなく件数の**見える化**で捕まえる（2026-08-31）。

    固定数（28）での拒否は決め打ち廃止で外した——表の行数を変える正当な編集
    まで拒否してしまうため。代わりに verify と編集画面の保存結果が
    「金額 N 列」を必ず表示し、N=0 を見た管理者が気づける（N-1/N-4 の
    検知経路をエラーから表示へ移した）。

    「拒否されないこと」は validate_v1 が例外を投げないことで固定する。
    列数そのものの後退検知（GUI 側の読み込み時基準比較・CLI の
    `verify --expect-columns`）は本テストの責務外——
    test_verify_expect_columns_matches_and_detects_shortfall で別途固定する
    （issue #65-1・レビュー指摘 S-5: 金額列の契約と列数比較の契約は別物の
    ため同居させない）。
    """
    from chouhyo_ocr.columns import amount_cell_count, validate_v1
    ok = _template_with(tmp_path)
    assert amount_cell_count(ok) == 28
    broken = _template_with(tmp_path, drop_normalize=True)
    validate_v1(broken)                       # 拒否はしない（例外にならない）
    assert amount_cell_count(broken) == 0     # が、表示で 0 と分かる


def _verify_cfg(tmp_path):
    """verify を cli.main 経由で呼ぶための最小 config.json（issue #65-1）。

    poppler/credentials/api_budget など列数と無関係なチェックの成否は
    環境依存（この tmp_path には資格情報が無いため常に NG になる）ため、
    以下のテストは overall の終了コードではなく「template」チェックの
    イベント自身の ok/error だけを見る。同型の判定方法は
    test_output_columns_stage4.py の _verify_template_event が既に使っている
    （このverify呼び出しは環境非依存で再現できることを同ファイルが示している）。
    """
    p = tmp_path / "config.json"
    p.write_text(json.dumps({
        "output_dir": str(tmp_path / "out"), "workdir": str(tmp_path / "wd"),
        "log_dir": str(tmp_path / "logs"),
    }), encoding="utf-8")
    return p


def _verify_template_event(tmp_path, capsys, template_path, expect_columns=None):
    cfg_path = _verify_cfg(tmp_path)
    args = ["--config", str(cfg_path), "verify", "--template", str(template_path)]
    if expect_columns is not None:
        args += ["--expect-columns", str(expect_columns)]
    cli.main(args)
    events = [json.loads(line) for line in capsys.readouterr().out.splitlines() if line.strip()]
    return next(e for e in events if e.get("check") == "template")


def test_verify_expect_columns_matches_and_detects_shortfall(tmp_path, capsys):
    """`verify --expect-columns` は列数の後退を検知する最後の砦（issue #65-1 穴C）。

    GUI 側の読み込み時基準比較（Editor.tsx の columnDecreaseFor）に加えて、
    CLI 単体でも列数の後退を検知できるようにする段——GUI が基準を取得
    できない場合（invoke 失敗・自動読込失敗）でも、この CLI 経路は独立に
    機能する。一致・不足・省略の3方向を固定する（レビュー指摘 S-5 で
    test_amount_count_is_reported_not_rejected から分離）。
    """
    # 一致（実列数 220 と同じ期待値）: 失敗にしない
    matched = _verify_template_event(tmp_path, capsys, TPL, expect_columns=220)
    assert matched["ok"] is True
    assert matched["columns"] == 220
    assert "error" not in matched

    # 実列数を上回る期待値を渡すと不足として検知する（ok:False）。
    # メッセージに記入値（帳票の値）は含まれない——列数という件数のみ
    short = _verify_template_event(tmp_path, capsys, TPL, expect_columns=221)
    assert short["ok"] is False
    assert "221" in short["error"] and "220" in short["error"]

    # 省略時は従来どおり挙動不変（列数と無関係に ok:True のまま）
    omitted = _verify_template_event(tmp_path, capsys, TPL)
    assert omitted["ok"] is True
    assert "error" not in omitted


# ---------- #13: glob メタ文字を含むファイル名 ----------

def test_expand_handles_bracket_filename(tmp_path):
    """scan[1].pdf（ブラウザの重複既定名）が展開失敗にならない。"""
    from PIL import Image

    from chouhyo_ocr.ingest import expand
    src = tmp_path / "scan[1].pdf"
    Image.new("L", (200, 280), 255).save(src)  # 1ページの最小 PDF
    out = tmp_path / "pages"
    out.mkdir()
    pages = expand(src, dpi=72, out_dir=out)
    assert len(pages) == 1
    assert pages[0].name.startswith("scan[1]-")


def test_run_accepts_single_file_input(tmp_path):
    """run --input は単一ファイルも受ける（フォルダ縛りの解消・issue #19）。"""
    from chouhyo_ocr.ingest import list_inputs
    f = tmp_path / "scan_0001.png"
    f.write_bytes(b"\x89PNG\r\n")
    assert list_inputs(f) == [f]
    bad = tmp_path / "memo.txt"
    bad.write_text("x")
    assert list_inputs(bad) == []


def test_expand_page_cli_returns_png(tmp_path):
    """expand-page が PDF の指定ページを PNG 展開してパスを返す（issue #19）。"""
    import subprocess

    from PIL import Image
    src = tmp_path / "sample.pdf"
    Image.new("L", (200, 280), 255).save(src)
    cfg_file = tmp_path / "config.json"
    cfg_file.write_text(json.dumps({"workdir": str(tmp_path / "wd"),
                                    "log_dir": str(tmp_path / "logs")}),
                        encoding="utf-8")
    from chouhyo_ocr.paths import app_root
    python = app_root() / ".venv" / "Scripts" / "python.exe"
    r = subprocess.run(
        [str(python), "-X", "utf8", "-m", "chouhyo_ocr.cli",
         "--config", str(cfg_file), "expand-page", "--input", str(src)],
        cwd=app_root() / "core", capture_output=True, text=True,
        encoding="utf-8", timeout=120)
    assert r.returncode == 0, r.stderr
    ev = next(json.loads(l) for l in r.stdout.splitlines()
              if l.strip() and "expand_page" in l)
    assert ev["ok"] is True and ev["pages"] == 1
    from pathlib import Path as _P
    assert _P(ev["page_path"]).exists()
    # 絶対パスであること。相対だと GUI 側の cwd で解決されて見つからない
    #（dev 窓で「展開中…」のまま止まった実測原因・2026-08-28）
    assert _P(ev["page_path"]).is_absolute()
    # 業務的な失敗（ページ範囲外）は exit 0 ＋ ok:false で伝える（issue #21）。
    # 非ゼロで返すと Rust の run_core_capture が Err(stdout) を返し、フロントの
    # ok:false 分岐が到達不能になって**生 JSON が画面に出る**（実測）
    r2 = subprocess.run(
        [str(python), "-X", "utf8", "-m", "chouhyo_ocr.cli",
         "--config", str(cfg_file), "expand-page", "--input", str(src),
         "--page", "5"],
        cwd=app_root() / "core", capture_output=True, text=True,
        encoding="utf-8", timeout=120)
    assert r2.returncode == 0, "業務的失敗を非ゼロで返している"
    ev2 = next(json.loads(l) for l in r2.stdout.splitlines()
               if l.strip() and "expand_page" in l)
    assert ev2["ok"] is False and "ページ 5 が無い" in ev2["error"]


def _expand_page_cfg(tmp_path):
    cfg_path = tmp_path / "config.json"
    cfg_path.write_text(json.dumps({"workdir": str(tmp_path / "wd"),
                                    "log_dir": str(tmp_path / "logs")}),
                        encoding="utf-8")
    return cfg_path


def test_expand_page_reason_is_align_on_alignment_failure(tmp_path, capsys):
    """位置合わせ失敗（罫線アンカーが取れない白紙画像）は reason:"align"（起票漏れ分）。

    以前は TemplateError・AlignError・画像不正を全て同じ aligned:false に潰す
    bare except で、GUI 側が失敗種別を区別できなかった。
    """
    from PIL import Image

    from chouhyo_ocr import cli
    src = tmp_path / "sample.pdf"
    Image.new("L", (200, 280), 255).save(src)  # 罫線が無く align_page が失敗する
    cfg_path = _expand_page_cfg(tmp_path)
    r = cli.main(["--config", str(cfg_path), "expand-page", "--input", str(src)])
    assert r == 0
    ev = next(json.loads(l) for l in capsys.readouterr().out.splitlines()
              if l.strip() and "expand_page" in l)
    assert ev["ok"] is True and ev["aligned"] is False
    assert ev["reason"] == "align"
    # 例外メッセージ本文は出さない（パスに入力ファイル名が乗りうる・既存方針）
    assert "error" not in ev


def test_expand_page_reason_is_template_on_broken_template(tmp_path, capsys):
    """テンプレート破損（schema_version 不正）は reason:"template"（align とは区別）。"""
    from PIL import Image

    from chouhyo_ocr import cli
    from chouhyo_ocr.paths import app_root
    src = tmp_path / "sample.pdf"
    Image.new("L", (200, 280), 255).save(src)
    cfg_path = _expand_page_cfg(tmp_path)

    raw = json.loads((app_root() / "templates" / "chouhyo-v1.json")
                     .read_text(encoding="utf-8"))
    raw["schema_version"] = 2
    bad_tpl = tmp_path / "bad.json"
    bad_tpl.write_text(json.dumps(raw, ensure_ascii=False), encoding="utf-8")

    r = cli.main(["--config", str(cfg_path), "expand-page", "--input", str(src),
                  "--template", str(bad_tpl)])
    assert r == 0
    ev = next(json.loads(l) for l in capsys.readouterr().out.splitlines()
              if l.strip() and "expand_page" in l)
    assert ev["ok"] is True and ev["aligned"] is False
    assert ev["reason"] == "template"
    assert "error" not in ev


def test_expand_page_reason_is_size_on_page_size_mismatch(tmp_path, capsys):
    """寸法/向き不一致（Q-H1・PageSizeMismatch）は reason:"size"（N-2）。

    PageSizeMismatch は AlignError のサブクラス。cmd_expand_page の except を
    AlignError より前に置かないと下の分岐に落ちて reason:"align" に化ける
    ——編集画面には「読み取り時に自動補正される」という誤った案内が出ていた
    （実際は run が様式不一致として弾く）。
    """
    from PIL import Image

    from chouhyo_ocr import cli
    src = tmp_path / "sample.pdf"
    # テンプレート（chouhyo-v1.json・2490x3510・比 約0.7094）に対し、200x250
    # は比が 0.8 で相対差 約12.8%（1% 閾値を大きく超える）
    Image.new("L", (200, 250), 255).save(src)
    cfg_path = _expand_page_cfg(tmp_path)
    r = cli.main(["--config", str(cfg_path), "expand-page", "--input", str(src)])
    assert r == 0
    ev = next(json.loads(l) for l in capsys.readouterr().out.splitlines()
              if l.strip() and "expand_page" in l)
    assert ev["ok"] is True and ev["aligned"] is False
    assert ev["reason"] == "size"
    assert "error" not in ev


def test_choice_with_subfields_keeps_row_length(tmp_path):
    """choice＋subfields のテンプレートでも行の値数＝列数が保たれる（issue #26）。

    エディタの種類切替が分割指定を残しても、読み込み時に落として無害化する。
    """
    from chouhyo_ocr.render_rows import build_row
    t = json.loads(TPL.read_text(encoding="utf-8"))
    for face in t["faces"]:
        for tb in face.get("tables", []):
            for c in tb["columns"]:
                if c.get("subfields"):
                    c["kind"] = "choice"  # エディタの切替事故を装う
                    c.pop("normalize", None)
                    # マークの x_offset は列相対ではなく**表原点相対**
                    # （実テンプレート: 列 x_offset=671 に対しマーク 686）。
                    # 列の内側に収める——外に出るテンプレートは load_template が
                    # 拒否するようになったため（レビュー4巡目 #48 のコア側検証）
                    half = max(1, c["width"] // 2)
                    c["choice_marks"] = [
                        {"value": "A", "x_offset": c["x_offset"],
                         "width": half},
                        {"value": "B", "x_offset": c["x_offset"] + half,
                         "width": c["width"] - half}]
    p = tmp_path / "t.json"
    p.write_text(json.dumps(t, ensure_ascii=False), encoding="utf-8")
    template = load_template(p)
    # choice セルの subfields は落ち、output_columns は常に1
    assert all(len(c.output_columns()) == 1
               for c in template.cells if c.kind == "choice")
    n_extract = sum(len(c.output_columns()) for c in template.cells)
    cells = {c.field_id: ("", None, c.kind, False) for c in template.cells}
    row = build_row(
        template,
        {"page_id": "p", "source_file": "s.png", "page_no": 1,
         "status": "", "unassigned_below_table": 0},
        cells, {}, CFG)
    assert len(row.values) == n_extract


def test_write_outputs_rejects_row_length_mismatch(tmp_path):
    """値数≠列数の行は明示例外で拒否される（issue #27・assert 非依存）。"""
    from chouhyo_ocr.render_out import write_outputs
    from chouhyo_ocr.render_rows import Row
    cols = ["要確認セル数", "最低信頼度", "帳票ID", "入力ファイル名",
            "ページ番号", "ステータス", "a", "b", "c"]
    bad = Row(page_id="p1", source_file="s.pdf", page_no=1, status="正常",
              values=["x", "y"], unclear_count=0, min_conf="")  # 3列に2値
    with pytest.raises(ValueError, match="値数"):
        write_outputs(tmp_path, "t", cols, [bad])


def test_store_upserts_sweep_stale_rows(tmp_path):
    """cell/era の総入れ替えで、今回書かなかった残骸が消える（issue #28）。"""
    from chouhyo_ocr.store import Store
    st = Store(tmp_path / "s.db")
    st.upsert_page("p1", "a.pdf", 1, "expanded")
    st.upsert_cells("p1", [("f_old", "旧値", 0.9, "text", 0),
                           ("f_keep", "残す", 0.9, "text", 0)])
    st.upsert_eras("p1", {"era_old": {"昭": 0.5}})
    # テンプレ変更後の再割付を模す: f_old / era_old は今回書かれない
    st.upsert_cells("p1", [("f_keep", "新値", 0.95, "text", 0),
                           ("f_new", "追加", 0.9, "text", 0)])
    st.upsert_eras("p1", {"era_new": {"S": 0.4}})
    cells = st.cells("p1")
    assert set(cells) == {"f_keep", "f_new"}       # f_old が残らない
    assert cells["f_keep"][0] == "新値"
    assert set(st.era_scores("p1")) == {"era_new"}  # 旧選択肢名が残らない
    st.close()


def test_run_reports_stale_pages(tmp_path):
    """入力から消えたファイルの行が残っている場合、run が可視化する（issue #28）。"""
    if not (RESP.exists() and PAGE_PNG_EXISTS):
        pytest.skip("保存済み応答が無い環境")
    import shutil

    from chouhyo_ocr.pipeline import run
    from chouhyo_ocr.vision_client import ReplayClient
    from chouhyo_ocr.paths import app_root
    page_png = app_root() / "testdata" / "local" / "pages" / "sample-1.png"
    cfg = Config(unclear_threshold=0.4, output_dir=str(tmp_path / "o"),
                 workdir=str(tmp_path / "w"), log_dir=str(tmp_path / "l"))
    inp = tmp_path / "in"; inp.mkdir()
    resp = tmp_path / "resp"; resp.mkdir()
    shutil.copy(page_png, inp / "a.png")
    shutil.copy(RESP, resp / "a_p0001.json")
    run(inp, TPL, cfg, ReplayClient(resp))

    (inp / "a.png").unlink()  # a を消し、別内容の b を足す
    # 内容まで同一にすると #46 の改名判定（同じ紙が改名された）に入り、
    # 中間データが b へ付け替わるので stale にならない。ここで見たいのは
    # 「入力から消えたファイルの行が残る」ケースなので、IEND より後ろの
    # パディングで sha1 だけ変えて別の紙にする（画素は同一）
    (inp / "b.png").write_bytes(page_png.read_bytes() + b"\n")
    shutil.copy(RESP, resp / "b_p0001.json")
    events = []
    run(inp, TPL, cfg, ReplayClient(resp), progress=events.append)
    stale = [e for e in events if e.get("event") == "stale_pages"]
    assert stale and stale[0]["count"] == 1 and stale[0]["files"] == ["a.png"]


def test_run_e2e_with_single_file_input(tmp_path):
    """単一ファイル入力で run→render が完走し、通常の1行出力になる（issue #19）。"""
    import shutil

    from chouhyo_ocr.paths import app_root
    from chouhyo_ocr.pipeline import render, run
    from chouhyo_ocr.vision_client import ReplayClient
    resp_src = app_root() / "testdata" / "local" / "s2" / "resp_DOCUMENT_TEXT_DETECTION.json"
    page_png = app_root() / "testdata" / "local" / "pages" / "sample-1.png"
    if not (resp_src.exists() and page_png.exists()):
        pytest.skip("保存済み応答が無い環境")
    cfg = Config(unclear_threshold=0.4, output_dir=str(tmp_path / "out"),
                 workdir=str(tmp_path / "wd"), log_dir=str(tmp_path / "logs"))
    single = tmp_path / "scan_0001.png"
    shutil.copy(page_png, single)
    resp = tmp_path / "resp"; resp.mkdir()
    shutil.copy(resp_src, resp / "scan_0001_p0001.json")

    summary = run(single, TPL, cfg, ReplayClient(resp))  # フォルダでなくファイル
    assert summary.pages == 1 and summary.rows == 1
    _x, _c, rows = render(TPL, cfg, timestamp="sf")
    assert rows[0].source_file == "scan_0001.png"
    assert rows[0].status == "正常"


def test_expand_replaced_pdf_does_not_mix_stale_pages(tmp_path):
    """同名 PDF の差し替え再実行で旧展開分が混ざらない（issue #20 の実測シナリオ）。

    12頁→2頁へ差し替えると、旧実装は残骸と合わせ14頁を返し、ゼロ埋め幅の
    差（12頁=2桁・2頁=1桁）で辞書順ソートの並びも崩れていた。
    """
    from PIL import Image

    from chouhyo_ocr.ingest import expand
    out = tmp_path / "pages"; out.mkdir()
    src = tmp_path / "y.pdf"

    frames = [Image.new("L", (100, 140), 255 - i) for i in range(12)]
    frames[0].save(src, save_all=True, append_images=frames[1:])
    first = expand(src, dpi=36, out_dir=out)
    assert len(first) == 12

    frames2 = [Image.new("L", (100, 140), 250), Image.new("L", (100, 140), 249)]
    frames2[0].save(src, save_all=True, append_images=frames2[1:])
    second = expand(src, dpi=36, out_dir=out)
    assert len(second) == 2, f"残骸が混ざった: {[p.name for p in second]}"
    nos = [int(p.stem.rsplit("-", 1)[1]) for p in second]
    assert nos == [1, 2]  # 数値順


def test_expand_does_not_pick_sibling_stem_pages(tmp_path):
    """a.pdf の展開が a-1.pdf 由来の a-1-1.png を拾わない（再レビュー N-13）。"""
    from PIL import Image

    from chouhyo_ocr.ingest import expand
    src = tmp_path / "a.pdf"
    Image.new("L", (200, 280), 255).save(src)
    out = tmp_path / "pages"
    out.mkdir()
    (out / "a-1-1.png").write_bytes(b"stale")   # a-1.pdf の展開分を装う
    (out / "a-extra.png").write_bytes(b"stale")  # 数字でない接尾辞
    pages = expand(src, dpi=72, out_dir=out)
    assert [p.name for p in pages] == ["a-1.png"]


# ---------- #14: config の検証 ----------

def _write_cfg(tmp_path, data):
    p = tmp_path / "config.json"
    p.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return p


def test_zero_threshold_rejected(tmp_path):
    """〓閾値 0 は転記主義の無効化なので拒否（GUI の空入力事故の最終防衛線）。"""
    p = _write_cfg(tmp_path, {"unclear_threshold": 0})
    with pytest.raises(ConfigError, match="unclear_threshold"):
        load_config(p)


def test_typo_key_rejected(tmp_path):
    """キー名 typo は無言で既定値に落とさず、キー名を示して拒否する。"""
    p = _write_cfg(tmp_path, {"unclear_thresold": 0.9})
    with pytest.raises(ConfigError, match="unclear_thresold"):
        load_config(p)


def test_wrong_type_rejected(tmp_path):
    p = _write_cfg(tmp_path, {"send_limit": "abc"})
    with pytest.raises(ConfigError, match="send_limit"):
        load_config(p)
    p2 = _write_cfg(tmp_path, {"workdir": ""})
    with pytest.raises(ConfigError, match="workdir"):
        load_config(p2)


# ---------- issue #72 (t)・FR-F29・AC-F60: last_template だけ例外を投げない ----------

def test_last_template_default_is_shipped(tmp_path):
    p = _write_cfg(tmp_path, {})
    assert load_config(p).last_template == "shipped"


def test_last_template_accepts_user_prefixed_name(tmp_path):
    p = _write_cfg(tmp_path, {"last_template": "user:帳票B"})
    assert load_config(p).last_template == "user:帳票B"


def test_last_template_invalid_format_falls_back_without_raising(tmp_path):
    """他のキーと違い ConfigError を投げない——形式不正は黙って "shipped" へ
    倒す（08 §3.10 不変条件6・AC-F60: 設定1行で起動不能にしない）。
    """
    for bad in ("bogus", "user:", "shipped:extra", "", 123, None, ["shipped"]):
        p = _write_cfg(tmp_path, {"last_template": bad})
        cfg = load_config(p)  # 例外を投げないこと自体が検証対象
        assert cfg.last_template == "shipped"


def test_last_template_fallback_logs_warning_without_name(tmp_path):
    """フォールバック時に警告ログが残るが、テンプレート名は出さない
    （Q-S1・FR-F50 の方針を踏襲）。値そのものに顧客名を想起させる文字列が
    入っていても、ログにはイベント名以外の情報を一切載せない。

    M-1（2026-09-02 マリン指摘）: フォールバック警告は `load_config()` 単体
    では出ない——`_validate()` は理由を `Config.last_template_fallback_reason`
    に積むだけで、実際に `log.warn()` するのは `cli._load_config_and_init_log`
    （`log.init` 直後の1回のみ）。本番の呼び出し順（load_config → log.init →
    警告出力）と揃えるため、ここも直接 `load_config`+`log.init` を呼ぶのではなく
    `cli._load_config_and_init_log` を通す。
    """
    bad_value = "田中様_申込書テンプレート"  # 形式不正（shipped/user: どちらでもない）
    p = _write_cfg(tmp_path, {"last_template": bad_value,
                               "log_dir": str(tmp_path / "logs")})
    cfg = cli._load_config_and_init_log(p)
    assert cfg.last_template == "shipped"
    app_log = (tmp_path / "logs" / "app.log").read_text(encoding="utf-8")
    assert "config_last_template_fallback" in app_log
    assert bad_value not in app_log


def test_valid_config_passes(tmp_path):
    p = _write_cfg(tmp_path, {"unclear_threshold": 0.9, "send_limit": 50,
                              "workdir": "wd"})
    cfg = load_config(p)
    assert cfg.unclear_threshold == 0.9
    assert cfg.send_limit == 50
    assert cfg.era_threshold == 0.05  # 未指定は既定値


def test_send_limit_zero_is_valid(tmp_path):
    """send_limit=0（送信しないドライラン）は従来どおり通る（再レビュー N-6）。"""
    cfg = load_config(_write_cfg(tmp_path, {"send_limit": 0}))
    assert cfg.send_limit == 0


# ---------- #34: xlsx の数式インジェクション ----------

def test_xlsx_never_writes_formula_from_read_values(tmp_path):
    """読取値が = で始まっても数式セルにならない（値は保持・issue #34）。

    openpyxl は "=" 始まりを無条件に数式型へ昇格させ、number_format="@" は
    効かない。実測では <f> タグが書かれ Excel で計算が走り、同じ行の COUNTIF
    まで巻き込んで壊れた。data_type の明示固定で塞ぐ（値は書き換えない）。
    """
    import zipfile

    from openpyxl import load_workbook

    from chouhyo_ocr.render_out import write_outputs
    from chouhyo_ocr.render_rows import Row
    cols = ["要確認セル数", "最低信頼度", "帳票ID", "入力ファイル名",
            "ページ番号", "ステータス", "備考", "氏名"]
    row = Row(page_id="p1", source_file="s.pdf", page_no=1, status="正常",
              values=["=SUM(A1:A9)", "+1+2"], unclear_count=0, min_conf="0.900")
    xlsx, csvp, _risky = write_outputs(tmp_path, "inj", cols, [row])

    xml = zipfile.ZipFile(xlsx).read("xl/worksheets/sheet1.xml").decode("utf-8")
    # COUNTIF（管理列）は数式のまま。読取値由来の数式が増えていないこと
    assert xml.count("<f>") == 1, "読取値が数式セルとして書かれた"
    assert "SUM(A1:A9)" not in xml.replace("&gt;", ">").split("<f>")[1].split("</f>")[0]

    ws = load_workbook(xlsx)["output"]
    assert ws.cell(row=2, column=7).value == "=SUM(A1:A9)"  # 値は保持
    assert ws.cell(row=2, column=7).data_type == "s"


def test_risky_prefix_detection_has_no_side_effects(tmp_path):
    """危険接頭の検出は出力を1バイトも変えない（D-28・A6）。

    CSV は読取値をそのまま書く。値の書き換え・〓化は転記主義（§5.5）と
    §8-12 の xlsx↔csv 一致に反するため行わない——検出は警告のみ。
    """
    import csv as csvmod

    from chouhyo_ocr.render_out import scan_risky_prefixes, write_outputs
    from chouhyo_ocr.render_rows import Row
    cols = ["要確認セル数", "最低信頼度", "帳票ID", "入力ファイル名",
            "ページ番号", "ステータス", "備考", "氏名"]
    risky_row = Row(page_id="p1", source_file="s.pdf", page_no=1, status="正常",
                    values=["=SUM(A1:A9)", "山田"], unclear_count=0, min_conf="0.9")
    safe_row = Row(page_id="p1", source_file="s.pdf", page_no=1, status="正常",
                   values=["普通の備考", "山田"], unclear_count=0, min_conf="0.9")

    _x1, c1, risky = write_outputs(tmp_path, "r1", cols, [risky_row])
    _x2, c2, safe = write_outputs(tmp_path, "r2", cols, [safe_row])
    assert risky == [("p1", "備考")]   # (page_id, 列名) のみ・値は返さない
    assert safe == []

    # CSV の値は読取値とバイト一致（接頭文字の付加・除去が無い）
    with open(c1, encoding="utf-8-sig", newline="") as f:
        got = list(csvmod.reader(f))[1][6]
    assert got == "=SUM(A1:A9)"
    # 検出は列名だけを見て値を持ち出さない
    assert all(len(t) == 2 and isinstance(t[1], str) for t in
               scan_risky_prefixes(cols, [risky_row]))


# ---------- #31/#32: テンプレート検証の穴（ファジングで検出）----------

def test_duplicate_choice_values_rejected(tmp_path):
    """選択肢の値重複を拒否する（issue #31）。

    重複すると era の候補が1つに潰れ、共通フロア減算で自滅して
    丸印があっても永久に〓になる（実測: decide({'昭': 0.3}) → 未選択）。
    """
    from chouhyo_ocr.template import TemplateError
    t = json.loads(TPL.read_text(encoding="utf-8"))
    for face in t["faces"]:
        for tb in face.get("tables", []):
            for c in tb["columns"]:
                for m in c.get("choice_marks", []):
                    m["value"] = "昭"  # 全部同じ値に
    p = tmp_path / "dup.json"
    p.write_text(json.dumps(t, ensure_ascii=False), encoding="utf-8")
    with pytest.raises(TemplateError, match="選択肢の値が重複"):
        load_template(p)


def test_overlapping_face_rects_rejected(tmp_path):
    """面の切り出し範囲の重なりを拒否する（issue #32・二重転記の防止）。"""
    from chouhyo_ocr.template import TemplateError
    t = json.loads(TPL.read_text(encoding="utf-8"))
    t["faces"][1]["source"]["rect"]["y"] = 100  # front(0..1880) と大きく重なる
    p = tmp_path / "ov.json"
    p.write_text(json.dumps(t, ensure_ascii=False), encoding="utf-8")
    with pytest.raises(TemplateError, match="重なっている"):
        load_template(p)


def test_overlapping_cell_rects_rejected(tmp_path):
    """同一面のセル矩形の重なりを拒否する（issue #24）。

    mapping は定義順の first-hit で解決するため、重なり帯の文字の行き先が
    「テンプレートの記述順」という見えない要素で決まってしまう。
    """
    from chouhyo_ocr.template import TemplateError
    t = json.loads(TPL.read_text(encoding="utf-8"))
    f0 = t["faces"][0]["fields"]
    f0[1]["rect"] = dict(f0[0]["rect"])  # 2つの欄を完全に重ねる
    p = tmp_path / "ov.json"
    p.write_text(json.dumps(t, ensure_ascii=False), encoding="utf-8")
    with pytest.raises(TemplateError, match="欄の矩形が重なっている"):
        load_template(p)


def test_shipped_template_has_no_overlaps():
    """同梱テンプレートに重なりが無い（issue #24 で6組を解消済み）。"""
    template = load_template(TPL)  # 重なりがあれば load 時点で落ちる
    assert len(template.cells) == 194  # 郵便番号1/2 の分離で 192→194（2026-08-31）


def test_cloud_marker_does_not_false_positive(monkeypatch):
    """無関係なフォルダ名を同期フォルダと誤検知しない（レビュー M-10）。

    `resolve()` を素通しに差し替える（issue #79）。`is_cloud_synced_path` は
    先頭で `Path(p).resolve()` を呼ぶが、UNC パスの解決は実際にネットワーク
    へ名前解決に行くため、共有が存在しない・DNS が遅い環境では待たされたり
    失敗したりして、この試験だけが散発的に落ちていた（単独再実行では通る）。

    ここで確かめたいのはパス文字列の判定規則（成分の完全一致・UNC の接頭辞）
    であって、実在するパスの解決結果ではない。差し替えても入力は全て絶対パス
    なので、resolve が返すべき値と一致する。
    """
    from pathlib import Path

    from chouhyo_ocr.paths import is_cloud_synced_path

    monkeypatch.setattr(Path, "resolve", lambda self, strict=False: self)
    sep = chr(92)  # バックスラッシュ（ソース中のエスケープ事故を避ける）
    join = lambda *p: sep.join(p)  # noqa: E731
    assert not is_cloud_synced_path(join("C:", "work", "dropbox_backup"))
    assert not is_cloud_synced_path(join("C:", "data", "my_onedrive_notes"))
    assert is_cloud_synced_path(join("C:", "Users", "u", "Dropbox", "d"))
    assert is_cloud_synced_path(join("C:", "Users", "u", "OneDrive - 会社名", "d"))
    assert is_cloud_synced_path(sep * 2 + join("fileserver", "share", "wd"))  # UNC


def test_out_of_face_rect_rejected(tmp_path):
    """面の範囲外へはみ出した欄を拒否する（レビュー M-20）。

    範囲外の欄は文字が来ず常に〓になるのに、原因が表示されなかった。
    """
    from chouhyo_ocr.template import TemplateError
    t = json.loads(TPL.read_text(encoding="utf-8"))
    t["faces"][0]["fields"][0]["rect"]["y"] = 99999  # 面の下端より遥か下
    p = tmp_path / "oob.json"
    p.write_text(json.dumps(t, ensure_ascii=False), encoding="utf-8")
    with pytest.raises(TemplateError, match="はみ出している"):
        load_template(p)


def test_two_choice_can_reach_undecided():
    """2択でも判定不能に到達できる（レビュー M-4）。

    旧実装は共通フロア減算で second が必ず 0 になり、拮抗しても常に
    どちらかを選んでいた（実測: 乱数20万件で判定不能0件）。
    """
    from chouhyo_ocr import era
    assert era.decide({"A": 0.20, "B": 0.19}, 0.05) == era.UNDECIDED
    assert era.decide({"A": 0.30, "B": 0.05}, 0.05) == "A"     # 明確な差は選ぶ
    assert era.decide({"A": 0.01, "B": 0.01}, 0.05) == era.UNSELECTED
    # 3候補では従来どおりフロアを引く（縦罫線の共通インクを相殺）
    assert era.decide({"A": 0.31, "B": 0.20, "C": 0.20}, 0.05) == "A"


def test_date_cells_exclude_printed_labels():
    """生年月日欄が印字ラベル「年」「月」「日」を巻き込まない（おかゆ指摘）。

    実測（修正前）: person_生年月日_日 が p0001 で "20日"、p0002 で "月20日"。
    矩形の幅較正が印字ラベルを避けきれず、日付列に非数値が混入していた。
    印字ラベルは 年=[1986,2045] 月=[2157,2206] 日=[2324,2380]（面ローカル）。
    """
    import json as _json

    from chouhyo_ocr.mapping import symbols_from_response, to_face_local
    resp_dir = app_root() / "core" / "workdir" / "responses"
    if not (resp_dir / "帳票抽出検証用2026-08-24_p0001.json").exists():
        pytest.skip("保存済み応答が無い環境")
    template = load_template(TPL)
    face = template.face("front")
    for rid in ("p0001", "p0002"):
        resp = _json.loads(
            (resp_dir / f"帳票抽出検証用2026-08-24_{rid}.json").read_text(encoding="utf-8"))
        local = to_face_local(face, symbols_from_response(resp))
        got = {}
        for name in ("年", "月", "日"):
            cell = next(c for c in template.cells
                        if c.field_id == f"person_生年月日_{name}")
            r = cell.rect
            inside = sorted((s for s in local
                             if r.x <= s.x < r.x + r.w and r.y <= s.y < r.y + r.h),
                            key=lambda s: s.x)
            got[name] = "".join(s.text for s in inside)
        assert got == {"年": "7", "月": "7", "日": "20"}, f"{rid}: {got}"
        assert all(v.isdigit() for v in got.values()), f"{rid}: 非数値の混入 {got}"


def test_expand_page_scoped_stale_removal(tmp_path):
    """page 指定の展開が他ページの PNG を巻き添えにしない（レビュー M-4）。"""
    from PIL import Image

    from chouhyo_ocr.ingest import expand
    src = tmp_path / "doc.pdf"
    frames = [Image.new("L", (100, 140), 250), Image.new("L", (100, 140), 249)]
    frames[0].save(src, save_all=True, append_images=frames[1:])
    out = tmp_path / "pages"; out.mkdir()

    p1 = expand(src, dpi=36, out_dir=out, page=1)
    assert len(p1) == 1
    p2 = expand(src, dpi=36, out_dir=out, page=2)
    assert len(p2) == 1
    # 1ページ目の PNG が残っている（旧実装は page 指定を無視して全部消していた）
    assert p1[0].exists(), "page=2 の展開が page=1 の PNG を消した"


def test_expand_page_stale_removal_includes_aligned_png(tmp_path):
    """page 指定の展開が同じページの -aligned.png も stale として消す（#60 M-7）。

    テンプレート編集画面（cli.cmd_expand_page）が editor_pages/ に作る位置
    合わせ済み下地「<stem>-p{page:04d}-aligned.png」は、従来の stale 掃除
    （<stem>-<数字> 完全一致）の対象外で purge するまで永久に残っていた。
    帳票原本の複製（個人情報）が滞留する問題への対応。
    """
    from PIL import Image

    from chouhyo_ocr.ingest import expand
    src = tmp_path / "doc.pdf"
    Image.new("L", (100, 140), 255).save(src)  # 1ページの最小 PDF
    out = tmp_path / "editor_pages"
    out.mkdir()
    # cli.cmd_expand_page が作る名前を模す（page=1 用の固定名）
    stale_aligned = out / f"{src.stem}-p0001-aligned.png"
    stale_aligned.write_bytes(b"stale-aligned")
    other_page_aligned = out / f"{src.stem}-p0002-aligned.png"
    other_page_aligned.write_bytes(b"other-page-aligned")

    expand(src, dpi=36, out_dir=out, page=1)
    assert not stale_aligned.exists(), "旧 -aligned.png が掃除対象外のまま残った"
    # page=1 の展開が page=2 の -aligned.png を巻き添えにしない
    # （レビュー M-4 と同じページ指定スコープの原則）
    assert other_page_aligned.exists(), "page=1 の展開が page=2 の -aligned.png を消した"


# --- detect-grid の防御（レビュー M-9・画像なし／読めない） -------------------

def test_detect_grid_without_image_reports_instead_of_crashing(capsys, tmp_path):
    """--mode ruled で画像を渡さないと、素の例外でなく日本語の指示が出る。"""
    from chouhyo_ocr.cli import main
    assert main(["detect-grid", "--region", "0,0,100,100", "--mode", "ruled"]) == 0
    ev = json.loads(capsys.readouterr().out.strip().splitlines()[-1])
    assert ev["event"] == "detect_grid" and ev["ok"] is False
    assert "画像" in ev["error"]


def test_detect_grid_with_unreadable_image_reports_instead_of_crashing(capsys):
    from chouhyo_ocr.cli import main
    assert main(["detect-grid", "--region", "0,0,100,100", "--mode", "ruled",
                 "--image", "no_such_file.png"]) == 0
    ev = json.loads(capsys.readouterr().out.strip().splitlines()[-1])
    assert ev["ok"] is False and "読み込め" in ev["error"]


# --- 試験ランナーの集計（レビュー LOW: 失敗メッセージの数字を拾わない） ------

def test_run_all_tests_counts_only_summary_lines():
    import re
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "scripts"))
    from run_all_tests import SUMMARY_LINE

    def agg(out):
        got = {}
        for line in out.splitlines():
            s = line.strip().strip("=").strip()
            if not (SUMMARY_LINE.match(s) or "test result:" in line):
                continue
            for n, k in re.findall(r"(\d+) (passed|failed|skipped|error)", line):
                got[k] = got.get(k, 0) + int(n)
        return got

    # pytest -q は = の罫線を付けない（実測 2026-08-28）
    assert agg("...  [100%]\n7 passed in 0.32s\n") == {"passed": 7}
    assert agg("==== 5 passed, 1 failed in 2.0s ====") == {"passed": 5, "failed": 1}
    assert agg("test result: ok. 2 passed; 0 failed; 0 ignored")["passed"] == 2
    # 失敗メッセージ中の "7 passed" は数えない
    assert agg("FAILED t.py::x - AssertionError: expected 7 passed\n"
               "3 passed in 1.0s") == {"passed": 3}
    # 1件も走らなかったときは 0 件と分かる（呼び出し側が FAIL にできる）
    assert agg("no tests ran in 0.10s") == {}
