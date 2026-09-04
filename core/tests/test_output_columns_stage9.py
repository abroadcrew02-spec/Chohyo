"""出力列制御（issue #66 系）段9: 升（1行 × 1列）単位の出力対象外。

要件の正本: docs/design/chouhyo-ocr/05_output_columns_requirements.md
§4.5（FR-3.1〜FR-3.4）・§7.5（AC-3.1〜AC-3.18）。

第1弾（段1〜段8）の `output` は「欄」と「表の列（全行一括）」の2粒度だった。
本段はそこへ3つ目の粒度——`tables[].output_disabled_cells` による升単位の
例外——を足す。設計上の要は次の2点で、テストもこの2点を軸に並べる。

  1. **窓口は `output_cells()` のまま**。升の例外は読み込み時に
     `CellSpec.output` へ畳まれるので、列を作る側（derive_columns）・値を
     作る側（build_row / build_failure_row）・件数を数える側（verify）は
     1行も変わらない。「追従するはず」を実測で固定するのがここ。
  2. **列が優先**。`columns[].output: false` の列では、例外リストに何が
     書いてあっても全行が対象外になる（論理積）。

AC の対応（本ファイル内の呼称 H01〜H18 = 要件書の AC-3.1〜AC-3.18）:
  H01 1つの升だけ対象外にでき、他の行・他の列は残る／定義は消えない
  H02 升を外しても残りの列の相対順序が変わらない
  H03 xlsx/csv のヘッダが derive_columns と一致し、行の値数も一致する
  H04 既存テンプレート無改変で 220/214 のまま（空配列も省略と同値）
  H05 未知の属性名（table 直下・要素内の両方）は読み込み拒否
  H06 実在しない row_no / column は読み込み拒否（table_id・row_no・column を明示）
  H07 同じ升への重複指定は読み込み拒否
  H08 列 false との共存で列が優先（例外リストは他の列にはそのまま効く）
  H09 全升を外して抽出列0になれば validate_v1 が拒否
  H10 verify の output_disabled_cells が升を数え、column_names が導出列と一致
  H11 外した升にだけ記入がある行が空行にならない（＋警告の印が升ごとに付く）
  H12 外す→戻すで API 送信 0・check_reusable が remap を要求・値が同一に復帰
  H13 geometry_hash 不変・template_hash 変化
  H14 COUNTIF 範囲・条件付き書式2本・csv 静的カウントが追従（char_level ON/OFF）
  H15 同一中間データ・同一テンプレートの再出力がバイト一致
  H16 xlsx と csv の抽出対象列のセル値が一致
  H17 外した升の読取値も漏出防止・purge の対象（特別扱いの経路が無い）
  H18 「全升を列挙」と「列ごと false」の2表現が同じ出力になる

core/tests に conftest.py が無いため、共通ヘルパは既存の stage1/stage2 と
同じ内容をこのファイルにも複製する（既存ファイルの流儀に合わせる）。
"""
import csv as csvmod
import dataclasses
import json
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from chouhyo_ocr import api_budget, cli
from chouhyo_ocr.align import geometry_hash, template_hash
from chouhyo_ocr.columns import (META_COLUMNS, derive_columns, excel_column_letter,
                                 extract_columns, validate_v1)
from chouhyo_ocr.config import Config
from chouhyo_ocr.mapping import CellContent, Symbol, assign
from chouhyo_ocr.paths import app_root, template_schema_path
from chouhyo_ocr.pipeline import OperationRefused, remap, render, run
from chouhyo_ocr.render_out import write_outputs
from chouhyo_ocr.render_rows import UNCLEAR, build_row
from chouhyo_ocr.store import Store
from chouhyo_ocr.template import TemplateError, load_template, output_cells
from chouhyo_ocr.vision_client import ReplayClient

TPL = app_root() / "templates" / "chouhyo-v1.json"
RESP = app_root() / "testdata" / "local" / "s2" / "resp_DOCUMENT_TEXT_DETECTION.json"
PAGE_PNG = app_root() / "testdata" / "local" / "pages" / "sample-1.png"
PYTHON = Path(sys.executable)

needs_replay = pytest.mark.skipif(
    not (RESP.exists() and PAGE_PNG.exists()), reason="保存済み応答が無い環境")

CFG = Config(unclear_threshold=0.85, era_threshold=0.06)


# ---------- 共通ヘルパ ----------

def _raw() -> dict:
    return json.loads(TPL.read_text(encoding="utf-8"))


def _write(tmp_path, data, name):
    p = tmp_path / name
    p.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return p


def _table(raw, table_id):
    return next(t for f in raw["faces"] for t in f.get("tables", [])
                if t["table_id"] == table_id)


def _disable_cells(raw, table_id, entries):
    """tables[].output_disabled_cells を設定する（entries: [(row_no, 列名)]）。"""
    t = _table(raw, table_id)
    t["output_disabled_cells"] = [{"row_no": r, "column": c} for r, c in entries]
    return t


def _disable_column(raw, table_id, name):
    col = next(c for c in _table(raw, table_id)["columns"] if c["name"] == name)
    col["output"] = False
    return col


def make_cfg(tmp_path) -> Config:
    return Config(unclear_threshold=0.4, output_dir=str(tmp_path / "out"),
                  workdir=str(tmp_path / "wd"), log_dir=str(tmp_path / "logs"))


def _run_with_template(tmp_path, cfg, tpl_path):
    """実応答（保存済み・課金ゼロ）で1ページ分の run を、指定テンプレートで行う。"""
    inp = tmp_path / "input"; inp.mkdir(parents=True)
    resp = tmp_path / "resp"; resp.mkdir(parents=True)
    shutil.copy(PAGE_PNG, inp / "a.png")
    shutil.copy(RESP, resp / "a_p0001.json")
    run(inp, tpl_path, cfg, ReplayClient(resp))


def page(status="", below=0):
    return {"page_id": "p_0001", "source_file": "s.png", "page_no": 1,
            "status": status, "unassigned_below_table": below}


def _sym(text, rect, dx=5, dy=5, conf=0.98):
    return Symbol(text=text, x=rect.x + dx, y=rect.y + dy, conf=conf)


def _sheet_xml(path):
    z = zipfile.ZipFile(path)
    return z.read("xl/worksheets/sheet1.xml").decode("utf-8")


def _cfg_file(tmp_path, cfg):
    p = tmp_path / "cli_config.json"
    p.write_text(json.dumps({"output_dir": cfg.output_dir, "workdir": cfg.workdir,
                             "log_dir": cfg.log_dir}), encoding="utf-8")
    return p


@pytest.fixture()
def isolated_counter(tmp_path, monkeypatch):
    """API 送信カウンタを一時ディレクトリへ隔離する（P4-2・stage1 と同内容）。"""
    monkeypatch.setenv("CHOUHYO_USAGE_DIR_FOR_TESTS", str(tmp_path / "usage"))
    return tmp_path


def _synth(rows=5, columns=("品目", "備考"), disabled=None):
    """升の指定だけを見るための最小テンプレート（実素材に依存しない）。

    出荷テンプレートに「備考」という表の列が無いため、要件の例示に相当する
    field_id は合成側で用意する。table_id は既存スキーマの pattern
    `^[A-Za-z0-9._-]+$` により ASCII のみなので、例示の「明細」ではなく
    `meisai` を使う（列名 `name` には pattern が無く日本語のまま置ける）。
    """
    cols = [{"name": name, "x_offset": i * 200, "width": 190, "kind": "text"}
            for i, name in enumerate(columns)]
    table = {
        "table_id": "meisai", "row_pitch": 50, "row_height": 40,
        "blocks": [{"origin": {"x": 10, "y": 10}, "rows": rows}],
        "columns": cols,
    }
    if disabled is not None:
        table["output_disabled_cells"] = [{"row_no": r, "column": c}
                                          for r, c in disabled]
    return {
        "schema_version": 1, "template_id": "synth", "render_dpi": 300,
        "image": {"width": 1000, "height": 1000},
        "record": {"pages": 1},
        "faces": [{
            "face_id": "front",
            "source": {"page_offset": 0, "rect": {"x": 0, "y": 0, "w": 1000, "h": 1000}},
            "tables": [table],
        }],
    }


# ========== H01: 1つの升だけを対象外にする ==========

def test_h01_single_cell_excluded_keeps_other_rows_and_definition(tmp_path):
    """`meisai_03_備考` だけが列から消え、同じ列の他の行・同じ行の他の列は残る。"""
    base = load_template(_write(tmp_path, _synth(), "h01_base.json"))
    base_cols = derive_columns(base)
    assert "meisai_03_備考" in base_cols  # 前提: 例示どおりの field_id が作れている

    t = load_template(_write(
        tmp_path, _synth(disabled=[(3, "備考")]), "h01.json"))
    cols = derive_columns(t)

    assert "meisai_03_備考" not in cols
    assert "meisai_02_備考" in cols and "meisai_04_備考" in cols  # 同じ列の他の行は残る
    assert "meisai_03_品目" in cols                            # 同じ行の他の列も残る
    assert len(cols) == len(base_cols) - 1                    # 消えたのは1列だけ

    # 定義（枠・種別）はテンプレートに残る——削除ではなく出力の可否だけが変わる
    cell = next(c for c in t.cells if c.field_id == "meisai_03_備考")
    base_cell = next(c for c in base.cells if c.field_id == "meisai_03_備考")
    assert cell.output is False
    assert cell.rect == base_cell.rect and cell.kind == base_cell.kind
    assert len(t.cells) == len(base.cells)             # 物理セルの数は不変
    assert len(output_cells(t)) == len(t.cells) - 1    # 出力から外れたのは1升


def test_h01_shipped_template_single_cell(tmp_path):
    """出荷テンプレートでも同じ（明細表 3行目の品目1升だけを外す）。"""
    raw = _raw()
    _disable_cells(raw, "detail", [(3, "品目")])
    t = load_template(_write(tmp_path, raw, "h01s.json"))
    cols = derive_columns(t)

    assert "detail_03_品目" not in cols
    assert len(cols) == 219  # 220 - 1（実測値。列単位で外すと 220-28 になる）
    assert sum(1 for c in t.cells
               if c.field_id.endswith("_品目") and not c.output) == 1


# ========== H02: 列順不変 ==========

def test_h02_remaining_column_order_is_unchanged(tmp_path):
    """升を外しても、残った列の並びは元の並びの部分列そのもの。"""
    base_cols = derive_columns(load_template(TPL))
    raw = _raw()
    _disable_cells(raw, "detail", [(3, "品目"), (7, "金額")])
    _disable_cells(raw, "family", [(2, "氏名")])
    cols = derive_columns(load_template(_write(tmp_path, raw, "h02.json")))

    removed = {"detail_03_品目", "detail_07_金額", "family_02_氏名"}
    assert cols == [c for c in base_cols if c not in removed]
    assert cols[:len(META_COLUMNS)] == list(META_COLUMNS)  # 管理6列は先頭のまま


# ========== H03: ヘッダ＝derive_columns・値数一致 ==========

def test_h03_headers_match_derive_columns_and_row_width(tmp_path):
    raw = _raw()
    _disable_cells(raw, "detail", [(1, "品目"), (2, "金額")])
    t = load_template(_write(tmp_path, raw, "h03.json"))
    columns = derive_columns(t)
    n_extract = len(columns) - len(META_COLUMNS)

    # 値側（build_row）が列側（derive_columns）と同じ幅で組み上がる
    cells = {c.field_id: ("", None, c.kind, False) for c in t.cells}
    row = build_row(t, page(), cells, {}, CFG)
    assert len(row.values) == len(row.origins) == n_extract

    xlsx, csvp, _risky = write_outputs(tmp_path / "o", "h03", columns, [row])
    xheader = [c.value for c in load_workbook(xlsx)["output"][1]]
    with open(csvp, encoding="utf-8-sig", newline="") as f:
        cheader = next(csvmod.reader(f))
    assert xheader == columns
    assert cheader == columns


# ========== H04: 既存テンプレ無改変（＋空配列は省略と同値） ==========

def test_h04_unmodified_shipped_template_is_untouched():
    t = load_template(TPL)
    assert len(derive_columns(t)) == 220
    assert len(extract_columns(t)) == 214
    assert all(c.output is True for c in t.cells)
    assert len(output_cells(t)) == len(t.cells) == 194


def test_h04_empty_list_is_equivalent_to_omission(tmp_path):
    """`output_disabled_cells: []` は省略と同値（既定＝全升出力）。"""
    raw = _raw()
    for tid in ("family", "detail"):
        _disable_cells(raw, tid, [])
    t = load_template(_write(tmp_path, raw, "h04.json"))
    assert derive_columns(t) == derive_columns(load_template(TPL))
    assert all(c.output is True for c in t.cells)


# ========== H05: 未知の属性名の拒否 ==========

def test_h05_rejects_unknown_attribute_on_table(tmp_path):
    raw = _raw()
    _table(raw, "detail")["output_disabled_cell"] = []  # 単複のtypo
    with pytest.raises(TemplateError, match="スキーマ検証エラー"):
        load_template(_write(tmp_path, raw, "h05a.json"))


def test_h05_rejects_unknown_attribute_inside_entry(tmp_path):
    raw = _raw()
    _table(raw, "detail")["output_disabled_cells"] = [
        {"row_no": 3, "column": "品目", "col": 4}]  # 要素内の余計なキー
    with pytest.raises(TemplateError, match="スキーマ検証エラー"):
        load_template(_write(tmp_path, raw, "h05b.json"))


def test_h05_rejects_column_index_instead_of_name(tmp_path):
    """列を添字で書く形は型で弾く（列名参照であることをスキーマで固定する）。"""
    raw = _raw()
    _table(raw, "detail")["output_disabled_cells"] = [{"row_no": 3, "column": 4}]
    with pytest.raises(TemplateError, match="スキーマ検証エラー"):
        load_template(_write(tmp_path, raw, "h05c.json"))


def test_h05_row_no_maximum_matches_the_grid_bound_it_stands_for(tmp_path):
    """スキーマの `row_no.maximum` は「表が持ちうる最大行数」を表す。

    1600 という数そのものではなく **`blocks.maxItems` × `blocks[].rows.maximum`
    と一致していること**を主張する。どちらかを緩めたとき、この上限だけが
    古い値で置き去りになるのを検知するため（実在性の検査は load_template が
    別途行うので実害は小さいが、拒否がスキーマ側で起きるか検証側で起きるかが
    静かに入れ替わる）。
    """
    schema = json.loads(template_schema_path().read_text(encoding="utf-8"))
    table = schema["$defs"]["table"]["properties"]
    blocks = table["blocks"]
    row_no = table["output_disabled_cells"]["items"]["properties"]["row_no"]
    assert row_no["maximum"] == blocks["maxItems"] * blocks["items"]["properties"]["rows"]["maximum"]
    assert row_no["minimum"] == 1


# ========== H06: 実在しない row_no / column の拒否 ==========

def test_h06_rejects_row_no_beyond_total_rows(tmp_path):
    raw = _raw()
    _disable_cells(raw, "detail", [(29, "品目")])  # detail は 14+14=28 行
    with pytest.raises(TemplateError) as e:
        load_template(_write(tmp_path, raw, "h06a.json"))
    msg = str(e.value)
    assert "detail" in msg and "29" in msg and "品目" in msg
    assert "28" in msg  # 実際の行数も示す（直し方が分かる文言）


def test_h06_rejects_unknown_column_name(tmp_path):
    raw = _raw()
    _disable_cells(raw, "family", [(2, "備考")])  # どの表にも無い列名
    with pytest.raises(TemplateError) as e:
        load_template(_write(tmp_path, raw, "h06b.json"))
    msg = str(e.value)
    assert "family" in msg and "2" in msg and "備考" in msg


def test_h06_rejects_column_that_exists_in_another_table(tmp_path):
    """列名の検査は**その表の中だけ**で行う（他の表に実在しても拒否する）。

    表をまたいで列名を照合してしまうと、`detail` の列名を `family` に書いた
    指定が通り、どの升にも当たらないまま静かに無視される。
    """
    raw = _raw()
    _disable_cells(raw, "family", [(2, "品目")])  # 品目は detail には在るが family には無い
    with pytest.raises(TemplateError) as e:
        load_template(_write(tmp_path, raw, "h06b2.json"))
    msg = str(e.value)
    assert "family" in msg and "品目" in msg


def test_h06_accepts_last_row_and_rejects_row_zero(tmp_path):
    """境界: 最終行（28）は受理、0 はスキーマの minimum で拒否。"""
    raw = _raw()
    _disable_cells(raw, "detail", [(28, "品目")])
    t = load_template(_write(tmp_path, raw, "h06c.json"))
    assert "detail_28_品目" not in derive_columns(t)

    raw0 = _raw()
    _disable_cells(raw0, "detail", [(0, "品目")])
    with pytest.raises(TemplateError, match="スキーマ検証エラー"):
        load_template(_write(tmp_path, raw0, "h06d.json"))


def test_h06_rejection_happens_even_when_column_is_disabled(tmp_path):
    """列が false でも指定の妥当性検査は飛ばさない。

    飛ばすと、列を出力対象へ戻した瞬間に壊れた指定が生き返る。
    """
    raw = _raw()
    _disable_column(raw, "detail", "品目")
    _disable_cells(raw, "detail", [(99, "品目")])
    with pytest.raises(TemplateError, match="実在しない行"):
        load_template(_write(tmp_path, raw, "h06e.json"))


# ========== H07: 重複の拒否 ==========

def test_h07_rejects_duplicate_entry(tmp_path):
    raw = _raw()
    _disable_cells(raw, "detail", [(3, "品目"), (5, "金額"), (3, "品目")])
    with pytest.raises(TemplateError) as e:
        load_template(_write(tmp_path, raw, "h07.json"))
    msg = str(e.value)
    assert "重複" in msg and "detail" in msg and "3" in msg and "品目" in msg


def test_h07_same_row_different_columns_is_not_duplicate(tmp_path):
    """境界: 同じ行の別の列・別の行の同じ列は重複ではない。"""
    raw = _raw()
    _disable_cells(raw, "detail", [(3, "品目"), (3, "金額"), (4, "品目")])
    t = load_template(_write(tmp_path, raw, "h07b.json"))
    cols = derive_columns(t)
    for name in ("detail_03_品目", "detail_03_金額", "detail_04_品目"):
        assert name not in cols
    assert len(cols) == 217  # 220 - 3


# ========== H08: 列 false との共存（列が優先） ==========

def test_h08_column_false_wins_over_cell_exceptions(tmp_path):
    """列が false なら、その列への升指定は結果を1列も変えない。"""
    only_column = _raw()
    _disable_column(only_column, "detail", "品目")
    t_col = load_template(_write(tmp_path, only_column, "h08a.json"))

    both = _raw()
    _disable_column(both, "detail", "品目")
    _disable_cells(both, "detail", [(3, "品目"), (9, "品目")])
    t_both = load_template(_write(tmp_path, both, "h08b.json"))

    assert derive_columns(t_both) == derive_columns(t_col)
    assert len(derive_columns(t_col)) == 220 - 28  # 列単位で28行ぶん消える
    assert all(not c.output for c in t_both.cells if c.field_id.endswith("_品目"))


def test_h08_exceptions_on_other_columns_still_apply(tmp_path):
    """列 false の列以外への升指定は、そのまま効く（無効化されるのは同じ列だけ）。"""
    raw = _raw()
    _disable_column(raw, "detail", "品目")
    _disable_cells(raw, "detail", [(3, "品目"), (3, "金額")])
    t = load_template(_write(tmp_path, raw, "h08c.json"))
    cols = derive_columns(t)
    assert "detail_03_金額" not in cols
    assert "detail_04_金額" in cols
    assert len(cols) == 220 - 28 - 1


# ========== H09: 全升を外して抽出列0 ==========

def test_h09_all_cells_excluded_is_rejected_by_validate_v1(tmp_path):
    raw = _synth(rows=2, disabled=[(1, "品目"), (1, "備考"),
                                   (2, "品目"), (2, "備考")])
    t = load_template(_write(tmp_path, raw, "h09.json"))
    assert extract_columns(t) == []
    assert len(t.cells) == 4  # 升そのものは消えていない（出力から外れただけ）
    with pytest.raises(TemplateError, match="出力対象"):
        validate_v1(t)


def test_h09_one_remaining_cell_is_accepted(tmp_path):
    """境界: 1升でも残れば拒否しない（拒否は0のときだけ）。"""
    raw = _synth(rows=2, disabled=[(1, "品目"), (1, "備考"), (2, "品目")])
    t = load_template(_write(tmp_path, raw, "h09b.json"))
    assert validate_v1(t) == list(META_COLUMNS) + ["meisai_02_備考"]


# ========== H10: verify の件数と column_names ==========

def test_h10_verify_counts_cells_and_reports_column_names(tmp_path, capsys):
    """`output_disabled_cells` は**物理升**を数え、`columns` は列を数える。

    両者は 1 対 1 ではない。`family` の `生年月日` は subfields で 1 升＝3列
    なので、この2つの数の差が出る組み合わせを1件に混ぜて固定する
    （GUI 側がこの数を「N 欄を出力しません」に使うため・FR-1.9）。
    """
    raw = _raw()
    _disable_column(raw, "detail", "金額")                       # 列単位: 28升 / 28列
    _disable_cells(raw, "detail", [(3, "品目"), (5, "品目")])     # 升単位: 2升 / 2列
    _disable_cells(raw, "family", [(2, "生年月日")])              # 升単位: 1升 / 3列
    tpl = _write(tmp_path, raw, "h10.json")

    cfg_path = tmp_path / "config.json"
    cfg_path.write_text(json.dumps({
        "output_dir": str(tmp_path / "out"), "workdir": str(tmp_path / "wd"),
        "log_dir": str(tmp_path / "logs")}), encoding="utf-8")
    cli.main(["--config", str(cfg_path), "verify", "--template", str(tpl)])

    events = [json.loads(l) for l in capsys.readouterr().out.splitlines() if l.strip()]
    ev = next(e for e in events if e.get("check") == "template")

    t = load_template(tpl)
    assert ev["ok"] is True
    assert ev["output_disabled_cells"] == 31  # 28 + 2 + 1（升の数。列の数ではない）
    assert ev["cells"] == 194                 # 升そのものは減らない
    assert ev["column_names"] == derive_columns(t)
    # 列は 30升ぶんが1列ずつ＋生年月日1升ぶんが3列。31 ではなく 33 列減る
    assert ev["columns"] == len(ev["column_names"]) == 220 - 30 - 3
    for sf in ("年", "月", "日"):
        assert f"family_02_生年月日_{sf}" not in ev["column_names"]
    # 金額列を出力対象から外しても normalize:"amount" の升数は変わらない
    # （amount_cells は template.cells が母集団・出力の可否を見ない）
    assert ev["amount_cells"] == 28


# ========== H11: 外した升にだけ記入がある行 ==========

def test_h11_row_with_only_excluded_cell_filled_is_not_empty_row(tmp_path):
    """升を外した欄だけに記入がある行は空行にならず、同じ行の他の列は〓になる。

    空行判定の母集団は `template.cells`（全升）であり `output` を見ない
    （FR-1.2 の升への延長）。空行と誤判定すると、その行の他の欄が〓では
    なく空文字になり「未記入」と区別できなくなる。
    """
    raw = _raw()
    _disable_cells(raw, "family", [(5, "続柄")])
    t = load_template(_write(tmp_path, raw, "h11.json"))

    zokugara = next(c.rect for c in t.cells if c.field_id == "family_05_続柄")
    result = assign(t.cells, {"front": [_sym("長", zokugara)], "back": []}, t.faces)
    assert ("family", 5) not in result.empty_rows

    cells = {c.field_id: ("", None, c.kind, (c.table_id, c.row_no) in result.empty_rows)
             for c in t.cells}
    content = result.cells["family_05_続柄"]
    cells["family_05_続柄"] = (content.text, content.conf_min, "text", False)

    row = build_row(t, page(), cells, {}, CFG)
    extract_cols = derive_columns(t)[len(META_COLUMNS):]
    assert "family_05_続柄" not in extract_cols
    assert row.values[extract_cols.index("family_05_氏名")] == UNCLEAR
    # 同じ列の別の行（=空行のまま）は空文字。〓との違いが保たれている
    assert row.values[extract_cols.index("family_06_氏名")] == ""


def test_h11_warning_tag_is_applied_per_cell_not_per_column(tmp_path):
    """W-1 の「（出力対象外）」印が升ごとに付く（FR-3.3 の W-1〜W-3 升適用）。

    警告の生成は `CellSpec.output` を直接見ているため、升へ粒度が下りると
    印も自動的に升ごとになる。同じ列の隣り合う2行で印の有無が分かれることを
    実測で固定する——列単位で付いてしまうと、出力に残る側（04行目）の警告まで
    「対象外だから無視してよい」と誤読される。
    """
    base = load_template(TPL)
    rects = {fid: next(c.rect for c in base.cells if c.field_id == fid)
             for fid in ("detail_03_品目", "detail_04_品目")}

    raw = _raw()
    back = next(f for f in raw["faces"] if f["face_id"] == "back")
    for fid, r in rects.items():
        back.setdefault("exclusions", []).append({
            "id": f"cover_{fid}",
            "rect": {"x": r.x + 10, "y": r.y + 10, "w": 20, "h": 20}})
    _disable_cells(raw, "detail", [(3, "品目")])
    t = load_template(_write(tmp_path, raw, "h11w.json"))

    w1 = [w for w in t.warnings if w.startswith("[W-1]")]
    off = [w for w in w1 if "detail_03_品目" in w]
    on = [w for w in w1 if "detail_04_品目" in w]
    assert len(off) == 1 and len(on) == 1  # 両方とも警告そのものは出る（母集団に残る）
    assert off[0].endswith("（出力対象外）")
    assert "（出力対象外）" not in on[0]


# ========== H12: 外す→戻す（送信0・remap 要求・値同一） ==========

@needs_replay
def test_h12_toggle_cell_off_then_on_restores_same_value(tmp_path, isolated_counter):
    cfg = make_cfg(tmp_path)
    _run_with_template(tmp_path, cfg, TPL)
    used_after_run = api_budget.used_this_month()  # Replay は非加算（常に0）

    base_xlsx, _c, _r = render(TPL, cfg, timestamp="base")
    ws = load_workbook(base_xlsx)["output"]
    base_header = [c.value for c in ws[1]]
    base_value = [c.value for c in ws[2]][base_header.index("detail_03_品目")]

    raw = _raw()
    _disable_cells(raw, "detail", [(3, "品目")])
    tpl_off = _write(tmp_path, raw, "h12_off.json")

    # 升の指定は非幾何変更。render をそのまま拒否し remap を名指しする
    with pytest.raises(OperationRefused, match="remap"):
        render(tpl_off, cfg, timestamp="off_attempt")
    assert remap(tpl_off, cfg) == 1
    xlsx_off, _c2, _r2 = render(tpl_off, cfg, timestamp="off")
    header_off = [c.value for c in load_workbook(xlsx_off)["output"][1]]
    assert "detail_03_品目" not in header_off
    assert "detail_04_品目" in header_off

    # 戻す（元テンプレート＝指定なし）→ remap → 再出力で同じ値が復帰する
    assert remap(TPL, cfg) == 1
    xlsx_on, _c3, _r3 = render(TPL, cfg, timestamp="on")
    ws_on = load_workbook(xlsx_on)["output"]
    header_on = [c.value for c in ws_on[1]]
    assert [c.value for c in ws_on[2]][header_on.index("detail_03_品目")] == base_value
    assert api_budget.used_this_month() == used_after_run  # 一連の操作で送信0


# ========== H13: ハッシュの向き ==========

def test_h13_geometry_hash_unchanged_template_hash_changed(tmp_path):
    base = _raw()
    changed = _raw()
    _disable_cells(changed, "detail", [(3, "品目")])
    _write(tmp_path, changed, "h13.json")

    assert geometry_hash(changed) == geometry_hash(base)   # 幾何は動かない
    assert template_hash(changed) != template_hash(base)   # 版は動く（remap 要求）

    # 空配列を書いただけでも template_hash は動く（JSON が違うため）。
    # 出力は変わらないが再利用ガードは保守的側に倒れる——これは仕様
    empty = _raw()
    _disable_cells(empty, "detail", [])
    assert geometry_hash(empty) == geometry_hash(base)
    assert template_hash(empty) != template_hash(base)


# ========== H14: 列範囲に依存する経路の追従 ==========

def _row_with_two_unclear_cells(template, cfg):
    """`detail_03_品目` と `detail_04_品目` の2升だけが〓になる行を組む。

    値は **`build_row` に作らせる**（手組みの `Row` へ `unclear_count` を
    書いて assert すると、自分が入れた数を読み返すだけで「csv の静的カウントが
    導出後の列集合を母集団にしているか」を1つも検証しない）。他の升は
    `is_empty=True` にして空文字へ倒し、〓の発生源を2升に限定する。
    """
    cells = {c.field_id: ("", None, c.kind, True) for c in template.cells}
    for fid in ("detail_03_品目", "detail_04_品目"):
        cells[fid] = ("", None, "text", False)  # 読取値が空 → 欄全体〓
    return build_row(template, page(), cells, {}, cfg)


def test_h14_countif_formatting_and_csv_track_reduced_columns(tmp_path):
    raw = _raw()
    _disable_cells(raw, "detail", [(3, "品目")])
    t = load_template(_write(tmp_path, raw, "h14.json"))
    columns = derive_columns(t)
    assert len(columns) == 219
    first = excel_column_letter(len(META_COLUMNS) + 1)
    last = excel_column_letter(len(columns))
    assert last != excel_column_letter(220)  # 前提: 実際に列が縮んでいる

    # 対象外にしなければ2升とも数える（この2が1へ減ることが③の検証点）
    assert _row_with_two_unclear_cells(load_template(TPL), CFG).unclear_count == 2

    for on in (False, True):
        cfg = dataclasses.replace(CFG, unclear_char_level=on)
        row = _row_with_two_unclear_cells(t, cfg)
        assert len(row.values) == len(columns) - len(META_COLUMNS)
        # 外した升は要確認セル数の母集団から抜ける（04 の「母集団が縮む」）
        assert row.unclear_count == 1

        xlsx, csvp, _risky = write_outputs(
            tmp_path / f"out_{on}", f"h14_{on}", columns, [row],
            unclear_char_level=on)
        sheet = _sheet_xml(xlsx)
        pattern = '"*〓*"' if on else '"〓"'
        assert f'COUNTIF({first}2:{last}2,{pattern})' in sheet   # ①範囲が追従
        assert sheet.count("<cfRule") == (2 if on else 1)        # ②書式は OFF1/ON2 本
        assert f"{first}1:{last}" in sheet                       # ②同じ縮んだ範囲
        text = csvp.read_text(encoding="utf-8-sig")
        assert text.splitlines()[1].split(",")[0] == '"1"'       # ③csv 静的カウント


# ========== H15 / H16: 決定性と xlsx↔csv の一致 ==========

@needs_replay
def test_h15_render_is_byte_identical_on_repeat(tmp_path):
    cfg = make_cfg(tmp_path)
    _run_with_template(tmp_path, cfg, TPL)
    raw = _raw()
    _disable_cells(raw, "detail", [(3, "品目"), (7, "金額")])
    tpl = _write(tmp_path, raw, "h15.json")
    assert remap(tpl, cfg) == 1

    x1, c1, _r1 = render(tpl, cfg, timestamp="h15a")
    x2, c2, _r2 = render(tpl, cfg, timestamp="h15b")
    assert x1.read_bytes() == x2.read_bytes()
    assert c1.read_bytes() == c2.read_bytes()


@needs_replay
def test_h16_xlsx_and_csv_values_match(tmp_path):
    cfg = make_cfg(tmp_path)
    _run_with_template(tmp_path, cfg, TPL)
    raw = _raw()
    _disable_cells(raw, "detail", [(3, "品目"), (7, "金額")])
    _disable_cells(raw, "family", [(2, "生年月日")])  # subfields を持つ升
    tpl = _write(tmp_path, raw, "h16.json")
    assert remap(tpl, cfg) == 1

    t = load_template(tpl)
    n_extract = len(extract_columns(t))
    assert n_extract == 214 - 2 - 3  # 品目1 + 金額1 + 生年月日の3サブ列

    xlsx, csvp, _rows = render(tpl, cfg, timestamp="h16")
    ws = load_workbook(xlsx)["output"]
    xrow = [c.value for c in list(ws.iter_rows(min_row=2))[0]][len(META_COLUMNS):]
    with open(csvp, encoding="utf-8-sig", newline="") as f:
        crow = list(csvmod.reader(f))[1][len(META_COLUMNS):]
    assert len(xrow) == len(crow) == n_extract
    for i, (xv, cv) in enumerate(zip(xrow, crow)):
        assert ("" if xv is None else str(xv)) == cv, f"列{i + len(META_COLUMNS) + 1}"


def test_h16_subfield_cell_drops_all_its_columns(tmp_path):
    """升を外した subfields セルは、由来サブ列が3つとも消える（一部残らない）。"""
    raw = _raw()
    _disable_cells(raw, "family", [(2, "生年月日")])
    cols = set(derive_columns(load_template(_write(tmp_path, raw, "h16b.json"))))
    base = set(derive_columns(load_template(TPL)))
    assert base - cols == {"family_02_生年月日_年", "family_02_生年月日_月",
                           "family_02_生年月日_日"}


# ========== H17: 漏出防止・purge ==========

def test_h17_excluded_cell_value_never_leaks(tmp_path):
    """外した升の読取値も、他の升と同じ漏出防止（repr・Row.values）を受ける。"""
    raw = _raw()
    _disable_cells(raw, "detail", [(3, "品目")])
    t = load_template(_write(tmp_path, raw, "h17.json"))
    rect = next(c.rect for c in t.cells if c.field_id == "detail_03_品目")
    secret = "秘密の品名"
    syms = [Symbol(text=ch, x=rect.x + 5 + i * 20, y=rect.y + 5, conf=0.95)
            for i, ch in enumerate(secret)]
    result = assign(t.cells, {"front": [], "back": syms}, t.faces)
    content = result.cells["detail_03_品目"]
    assert content.text == secret        # 中間データには残る（可逆性）
    assert secret not in repr(content)   # repr には出ない
    assert secret not in repr(CellContent(secret, 0.9))

    cells = {c.field_id: ("", None, c.kind, False) for c in t.cells}
    cells["detail_03_品目"] = (content.text, content.conf_min, "text", False)
    row = build_row(t, page(), cells, {}, CFG)
    assert secret not in repr(row)
    assert secret not in row.values  # 対象外なので抽出列にすら現れない


@needs_replay
def test_h17_purge_removes_excluded_cell_intermediate_data(tmp_path):
    """外した升の読取値も purge で他と同様に消える（特別扱いの経路が無い）。"""
    cfg = make_cfg(tmp_path)
    raw = _raw()
    _disable_cells(raw, "detail", [(3, "品目")])
    tpl = _write(tmp_path, raw, "h17p.json")
    _run_with_template(tmp_path, cfg, tpl)

    wd = Path(cfg.workdir)
    store = Store(wd / "intermediate.sqlite")
    try:
        cells_before = store.cells(store.pages()[0]["page_id"])
    finally:
        store.close()
    assert "detail_03_品目" in cells_before  # 外した升の読取値も中間データに残る

    r = subprocess.run(
        [str(PYTHON), "-X", "utf8", "-m", "chouhyo_ocr.cli",
         "--config", str(_cfg_file(tmp_path, cfg)), "purge", "--yes"],
        cwd=app_root() / "core", capture_output=True, text=True,
        encoding="utf-8", timeout=60)
    assert r.returncode == 0
    assert wd.exists()  # keep-list 方式（#83）で workdir 自体は残る
    assert not (wd / "intermediate.sqlite").exists()


# ========== H18: 2つの表現の等価性 ==========

def test_h18_listing_all_cells_equals_disabling_the_column(tmp_path):
    """「全升を列挙」と「列ごと false」が同じ列構成になる（JSON 直編集の等価性）。

    GUI（編集画面）は当面この属性を書かないため、ここで固定するのは
    **JSON を直接編集する運用者が取りうる2つの書き方**が同じ結果に落ちる
    ことである。GUI 側の直列化契約（どちらの形で書くか）は GUI レーンが
    `buildTemplateJson` 側で別途固定する。
    """
    by_column = _raw()
    _disable_column(by_column, "detail", "品目")
    t_col = load_template(_write(tmp_path, by_column, "h18a.json"))

    by_cells = _raw()
    _disable_cells(by_cells, "detail", [(r, "品目") for r in range(1, 29)])
    t_cells = load_template(_write(tmp_path, by_cells, "h18b.json"))

    assert derive_columns(t_cells) == derive_columns(t_col)
    assert ([c.output for c in t_cells.cells] == [c.output for c in t_col.cells])


@needs_replay
def test_h18_two_representations_produce_identical_bytes(tmp_path):
    """列構成だけでなく、出力ファイルのバイト列まで一致する。"""
    cfg = make_cfg(tmp_path)
    _run_with_template(tmp_path, cfg, TPL)

    by_column = _raw()
    _disable_column(by_column, "detail", "品目")
    tpl_col = _write(tmp_path, by_column, "h18c.json")
    assert remap(tpl_col, cfg) == 1
    x_col, c_col, _r = render(tpl_col, cfg, timestamp="h18col")
    b_col = (x_col.read_bytes(), c_col.read_bytes())

    by_cells = _raw()
    _disable_cells(by_cells, "detail", [(r, "品目") for r in range(1, 29)])
    tpl_cells = _write(tmp_path, by_cells, "h18d.json")
    assert remap(tpl_cells, cfg) == 1
    x_cells, c_cells, _r2 = render(tpl_cells, cfg, timestamp="h18cell")

    assert (x_cells.read_bytes(), c_cells.read_bytes()) == b_col
