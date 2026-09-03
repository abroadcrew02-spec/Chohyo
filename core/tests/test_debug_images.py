"""debug-images（読み取り可視化・開発者モード）のテスト。

実サンプル素材（testdata/local/pages・testdata/local/s2）に依存するため、無い環境では skip
（test_e2e_replay と同じ流儀）。
"""
import shutil

import pytest

from chouhyo_ocr.config import Config
from chouhyo_ocr.paths import app_root
from chouhyo_ocr.pipeline import run
from chouhyo_ocr.template import load_template

RESP = app_root() / "testdata" / "local" / "s2" / "resp_DOCUMENT_TEXT_DETECTION.json"
PAGE_PNG = app_root() / "testdata" / "local" / "pages" / "sample-1.png"
TPL = app_root() / "templates" / "chouhyo-v1.json"

pytestmark = pytest.mark.skipif(
    not (RESP.exists() and PAGE_PNG.exists()), reason="replay 素材が無い環境")


from chouhyo_ocr.vision_client import ReplayClient


@pytest.fixture()
def worked(tmp_path):
    cfg = Config(output_dir=str(tmp_path / "o"), workdir=str(tmp_path / "w"),
                 log_dir=str(tmp_path / "l"))
    inp = tmp_path / "in"; inp.mkdir()
    resp = tmp_path / "resp"; resp.mkdir()
    shutil.copy(PAGE_PNG, inp / "a.png")
    shutil.copy(RESP, resp / "a_p0001.json")
    run(inp, TPL, cfg, ReplayClient(resp))
    return cfg, tmp_path


def test_writes_one_png_per_page(worked, tmp_path):
    from pathlib import Path

    from chouhyo_ocr.debug_images import write_debug_images
    from chouhyo_ocr.store import Store
    cfg, _ = worked
    wd = Path(cfg.workdir)
    out = tmp_path / "dbg"
    store = Store(wd / "intermediate.sqlite")
    try:
        made = write_debug_images(store, load_template(TPL), wd / "aligned",
                                  out, cfg)
    finally:
        store.close()
    assert len(made) == 1
    p = made[0]
    assert p.exists() and p.suffix == ".png"
    # 白紙 PNG ではなく、実際に紙と注記が描かれていること（サイズで近似）
    assert p.stat().st_size > 100_000, "描画がほぼ空（オーバーレイが失敗している）"
    # 出力サイズはテンプレート座標系
    from PIL import Image
    with Image.open(p) as im:
        assert im.size == load_template(TPL).image_size


def test_page_filter(worked, tmp_path):
    from pathlib import Path

    from chouhyo_ocr.debug_images import write_debug_images
    from chouhyo_ocr.store import Store
    cfg, _ = worked
    wd = Path(cfg.workdir)
    store = Store(wd / "intermediate.sqlite")
    try:
        none = write_debug_images(store, load_template(TPL), wd / "aligned",
                                  tmp_path / "dbg2", cfg,
                                  page_ids=["存在しないID"])
    finally:
        store.close()
    assert none == []


# ---------- cli debug-images（5巡目 第3〜4段・#59 H-5・#60 M-1①④） ----------

def _cli_cfg_path(cfg, tmp_path, name="cli_config.json"):
    from chouhyo_ocr.config import save_config
    p = tmp_path / name
    save_config(cfg, p)
    return p


def test_out_rejects_cloud_synced_path(worked, tmp_path):
    """--out が同期フォルダ配下だと拒否する（#59 H-5）。

    読取値・信頼度を焼き込んだ画像は中間データより濃い個人情報で、既定の
    workdir/debug/ は purge・verify の同期検査の対象だが --out は検査の外を
    通っていた。
    """
    from chouhyo_ocr import cli
    cfg, _ = worked
    cfg_path = _cli_cfg_path(cfg, tmp_path)
    synced_out = tmp_path / "OneDrive" / "debug"
    r = cli.main(["--config", str(cfg_path), "debug-images",
                  "--template", str(TPL), "--out", str(synced_out)])
    # 業務的な拒否は exit 0（レビュー差し戻し M-3）。main() の規約コメント
    # （:443-450）・同一コマンド内の page_not_found/no_pages/
    # OperationRefused（いずれも 0）と揃える
    assert r == 0
    assert not synced_out.exists(), "拒否されたのに出力先が作られている"


def test_out_default_is_not_checked(worked, tmp_path):
    """既定（--out 省略・workdir/debug）は同期フォルダ検査の対象外（従来どおり）。"""
    from chouhyo_ocr import cli
    cfg, _ = worked
    cfg_path = _cli_cfg_path(cfg, tmp_path)
    r = cli.main(["--config", str(cfg_path), "debug-images", "--template", str(TPL)])
    assert r == 0


def test_debug_images_refuses_after_template_change(worked, tmp_path, capsys):
    """テンプレート変更後は check_reusable が拒否する（#60 M-1①）。

    通さないと、変わった枠に旧テンプレ割付の〓判定を重ねた嘘の可視化を
    出してしまう（render/remap と同じ整合ゲート）。
    """
    import json

    from chouhyo_ocr import cli
    cfg, _ = worked
    cfg_path = _cli_cfg_path(cfg, tmp_path)
    raw = json.loads(TPL.read_text(encoding="utf-8"))
    fld = next(f for f in raw["faces"][0]["fields"] if f["kind"] == "text")
    fld["rect"]["x"] += 1  # geometry_hash（faces.source/exclusions等）は不変だが
                           # template_hash（全体）は変わる——#25 の検査対象
    changed = tmp_path / "changed.json"
    changed.write_text(json.dumps(raw, ensure_ascii=False), encoding="utf-8")
    r = cli.main(["--config", str(cfg_path), "debug-images", "--template", str(changed)])
    assert r == 0  # 業務的な拒否は exit 0（他コマンドの OperationRefused と同じ契約）
    events = [json.loads(l) for l in capsys.readouterr().out.splitlines() if l.strip()]
    refused = [e for e in events if e.get("event") == "refused"]
    assert refused and "テンプレートが変わっている" in refused[0]["error"]


def test_debug_images_no_pages_reports_reason(tmp_path, capsys):
    """中間データが空のとき ok:false・reason:no_pages（#60 M-1④）。

    従来は ok:true, count:0 固定で、業務的失敗と「元々0件」の区別がつかなかった。
    """
    import json

    from chouhyo_ocr.config import Config
    from chouhyo_ocr import cli
    cfg = Config(output_dir=str(tmp_path / "o"), workdir=str(tmp_path / "w"),
                 log_dir=str(tmp_path / "l"))
    cfg_path = _cli_cfg_path(cfg, tmp_path)
    r = cli.main(["--config", str(cfg_path), "debug-images", "--template", str(TPL)])
    assert r == 0
    events = [json.loads(l) for l in capsys.readouterr().out.splitlines() if l.strip()]
    ev = next(e for e in events if e.get("event") == "debug_images")
    assert ev["ok"] is False and ev["reason"] == "no_pages" and ev["count"] == 0


def test_debug_images_page_not_found_reports_reason(worked, tmp_path, capsys):
    """存在しないページ ID を指定すると、該当なしと明示する（#60 M-1④）。"""
    import json

    from chouhyo_ocr import cli
    cfg, _ = worked
    cfg_path = _cli_cfg_path(cfg, tmp_path)
    r = cli.main(["--config", str(cfg_path), "debug-images",
                  "--template", str(TPL), "--page", "存在しないID"])
    assert r == 0
    events = [json.loads(l) for l in capsys.readouterr().out.splitlines() if l.strip()]
    ev = next(e for e in events if e.get("event") == "debug_images")
    assert ev["ok"] is False and ev["reason"] == "page_not_found" and ev["count"] == 0


def test_debug_images_no_aligned_images_reports_reason(worked, tmp_path, capsys):
    """ページはあるが位置合わせ済み画像が無いとき ok:false・reason:no_aligned_images。"""
    import json
    import shutil
    from pathlib import Path

    from chouhyo_ocr import cli
    cfg, _ = worked
    cfg_path = _cli_cfg_path(cfg, tmp_path)
    shutil.rmtree(Path(cfg.workdir) / "aligned")  # 位置合わせ済み画像を消す
    r = cli.main(["--config", str(cfg_path), "debug-images", "--template", str(TPL)])
    assert r == 0
    events = [json.loads(l) for l in capsys.readouterr().out.splitlines() if l.strip()]
    ev = next(e for e in events if e.get("event") == "debug_images")
    assert ev["ok"] is False and ev["reason"] == "no_aligned_images" and ev["count"] == 0


# ---------- M-1（レビュー差し戻し）: unclear_char_level が debug-images まで届く ----------

def test_write_debug_images_accepts_config_and_reflects_char_level(worked, tmp_path):
    """cli.py が cfg.unclear_threshold（スカラー）だけを渡していたため、
    debug_images 内で Config を組み直すと unclear_char_level が常に既定
    False に落ちていた。cfg 本体を渡す経路になったことを型で確認し、
    ON 時に xlsx の一部〓判定と debug 側の〓シェーディング判定が一致する
    （どちらも conf<閾値をゲートに使うため、ON/OFF に関わらず「この欄は
    〓を含むか」の真偽は揃う）ことを実データで固定する。
    """
    import dataclasses
    from pathlib import Path

    from chouhyo_ocr.debug_images import write_debug_images
    from chouhyo_ocr.pipeline import render
    from chouhyo_ocr.render_rows import unclear_reason
    from chouhyo_ocr.store import Store

    cfg, _ = worked
    cfg_on = dataclasses.replace(cfg, unclear_threshold=0.85, unclear_char_level=True)
    wd = Path(cfg_on.workdir)

    # xlsx 側の事実: family_01_氏名 は一部〓（"上〓諒" 等）、detail_01_品目 は
    # 高信頼のクリーンな値（既存テストで確認済みの実データの性質）
    from openpyxl import load_workbook
    xlsx, _csv, _rows = render(TPL, cfg_on, timestamp="dbg_align")
    wb = load_workbook(xlsx)
    ws = wb["output"]
    header = [c.value for c in ws[1]]
    data = [c.value for c in ws[2]]
    name_val = data[header.index("family_01_氏名")]
    detail_val = data[header.index("detail_01_品目")]
    assert "〓" in name_val and name_val != "〓"  # 一部〓であることの前提確認
    assert "〓" not in str(detail_val)             # クリーンな値であることの前提確認

    # debug-images 側: cfg（Config 本体）を渡してもエラーにならない（M-1 の型修正）
    out = tmp_path / "dbg_m1"
    store = Store(wd / "intermediate.sqlite")
    try:
        made = write_debug_images(store, load_template(TPL), wd / "aligned", out, cfg_on)
        assert made  # ON でもクラッシュせず生成される

        # debug の〓シェーディング判定を直接再現する（unclear_reason は
        # write_debug_images 内部が使うのと同じ関数・同じ cfg）
        pid = store.pages()[0]["page_id"]
        cells = store.cells(pid)
        name_raw, name_conf, _k, _e = cells["family_01_氏名"]
        detail_raw, detail_conf, _k2, _e2 = cells["detail_01_品目"]
    finally:
        store.close()
    assert unclear_reason(name_raw, name_conf, cfg_on) is not None, \
        "xlsxで一部〓の欄なのにdebug側が〓なしと判定した（ずれ）"
    assert unclear_reason(detail_raw, detail_conf, cfg_on) is None, \
        "xlsxでクリーンな欄なのにdebug側が〓ありと誤判定した（ずれ）"


def test_field_origins_come_from_the_stored_cell_origin(worked, tmp_path):
    """#65-6: 由来は中間データの cell.origin をそのまま読む（再計算しない）。

    person_郵便番号1/2 は主が空・参照先採用（実測・test_mapping.py::
    test_person_fields と同じ前提）なので 'fallback' が保存されている。
    可視化がそれと同じ値を返すことを、DB の内容と突き合わせて固定する。
    """
    from pathlib import Path

    from chouhyo_ocr.debug_images import _field_origins
    from chouhyo_ocr.store import Store

    cfg, _ = worked
    wd = Path(cfg.workdir)
    store = Store(wd / "intermediate.sqlite")
    try:
        pages = store.pages()
        assert len(pages) == 1
        pid = pages[0]["page_id"]
        origins = _field_origins(store, pid)
        stored = {fid: o for fid, (_cc, o) in store.cell_extras(pid).items()}
    finally:
        store.close()
    assert origins == stored, "cell.origin 以外の情報源が混ざっている"
    assert origins.get("person_郵便番号1") == "fallback"
    assert origins.get("person_郵便番号2") == "fallback"


def test_conflict_origin_is_drawn_as_forced_unclear(worked, tmp_path):
    """#65-6: origin=='conflict' の欄は専用色の枠で描かれる（凡例にも出る）。

    実サンプルに conflict は出ないため、中間データへ直接 conflict を書いてから
    描画する（render_rows.build_row が同じ origin を見て欄全体〓にする経路と
    対になる表示分岐）。描画結果そのものは PNG なので、ここでは
    「例外なく生成される」ことと、色定数が枠色のどれとも重ならないことを固定する。
    """
    from pathlib import Path

    from chouhyo_ocr import debug_images
    from chouhyo_ocr.debug_images import (COL_CHOICE, COL_CONFLICT, COL_EXCLUDED,
                                          COL_FALLBACK, COL_TEXT, _field_origins,
                                          write_debug_images)
    from chouhyo_ocr.store import Store

    assert COL_CONFLICT not in (COL_TEXT, COL_CHOICE, COL_FALLBACK, COL_EXCLUDED)

    cfg, _ = worked
    wd = Path(cfg.workdir)
    store = Store(wd / "intermediate.sqlite")
    try:
        pid = store.pages()[0]["page_id"]
        target = "family_01_氏名"
        char_confs, _origin = store.cell_extras(pid)[target]
        store.upsert_cell_extras(pid, [(target, char_confs, "conflict")])
        assert _field_origins(store, pid)[target] == "conflict"
        made = write_debug_images(store, load_template(TPL), wd / "aligned",
                                  tmp_path / "dbg_conflict", cfg)
    finally:
        store.close()
    assert len(made) == 1 and made[0].stat().st_size > 100_000
    assert debug_images.COL_CONFLICT == (205, 40, 110)
