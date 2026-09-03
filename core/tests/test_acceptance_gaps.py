"""トレーサビリティ整理（03_test_requirements.md）で特定した Gap の自動テスト。

- TR-G1: remap は幾何セクション変更を拒否する（設計 §6.7・geometry_hash）
- TR-G2: 失敗系でも入力ページ数＝出力行数（要件 §3.4・§8-2）
- TR-G3: 様式不一致（D-15・枠外率）が発火し全〓行になる
- TR-G4: .xlsx と .csv の抽出対象列が全列一致する（§8-12 の全量版）
- TR-G5: purge は --yes なしで拒否（要件 §6.3）
- TR-G6: 資格情報なしの verify は失敗コードで終わる
"""
import json
import shutil
import subprocess

import pytest
from openpyxl import load_workbook

from chouhyo_ocr.config import Config
from chouhyo_ocr.paths import app_root
from chouhyo_ocr.pipeline import OperationRefused, remap, render, run
from chouhyo_ocr.vision_client import ReplayClient

RESP = app_root() / "testdata" / "local" / "s2" / "resp_DOCUMENT_TEXT_DETECTION.json"
PAGE_PNG = app_root() / "testdata" / "local" / "pages" / "sample-1.png"
TPL = app_root() / "templates" / "chouhyo-v1.json"
PYTHON = app_root() / ".venv" / "Scripts" / "python.exe"

pytestmark = pytest.mark.skipif(
    not (RESP.exists() and PAGE_PNG.exists()), reason="保存済み応答が無い環境")


def make_cfg(tmp_path) -> Config:
    return Config(unclear_threshold=0.4, output_dir=str(tmp_path / "out"),
                  workdir=str(tmp_path / "wd"), log_dir=str(tmp_path / "logs"))


def run_one_page(tmp_path, cfg):
    inp = tmp_path / "input"; inp.mkdir()
    resp = tmp_path / "resp"; resp.mkdir()
    shutil.copy(PAGE_PNG, inp / "a.png")
    shutil.copy(RESP, resp / "a_p0001.json")
    return run(inp, TPL, cfg, ReplayClient(resp))


def test_tr_g1_remap_rejects_geometry_change(tmp_path):
    """幾何（source.rect 等）を変えたテンプレートで remap → 拒否して run を促す。"""
    cfg = make_cfg(tmp_path)
    run_one_page(tmp_path, cfg)

    tpl2 = tmp_path / "moved.json"
    t = json.loads(TPL.read_text(encoding="utf-8"))
    t["faces"][0]["source"]["rect"]["h"] += 10   # 幾何セクションの変更
    t["faces"][1]["source"]["rect"]["y"] += 10
    t["faces"][1]["source"]["rect"]["h"] -= 10
    tpl2.write_text(json.dumps(t, ensure_ascii=False), encoding="utf-8")

    with pytest.raises(OperationRefused, match="run"):
        remap(tpl2, cfg)

    # 非幾何の変更（欄の矩形）は remap を通す
    tpl3 = tmp_path / "field_moved.json"
    t3 = json.loads(TPL.read_text(encoding="utf-8"))
    t3["faces"][0]["fields"][0]["rect"]["x"] += 5
    tpl3.write_text(json.dumps(t3, ensure_ascii=False), encoding="utf-8")
    assert remap(tpl3, cfg) == 1


def test_tr_g2_failure_rows_preserved(tmp_path):
    """正常1・送信失敗1・展開失敗1 → 出力は3行（行数維持）＋ステータス書き分け。"""
    cfg = make_cfg(tmp_path)
    inp = tmp_path / "input"; inp.mkdir()
    resp = tmp_path / "resp"; resp.mkdir()
    base = PAGE_PNG.read_bytes()
    (inp / "ok.png").write_bytes(base + b"\x01")
    (inp / "noresp.png").write_bytes(base + b"\x02")   # 応答なし → 送信失敗
    (inp / "broken.pdf").write_bytes(b"%PDF-1.4 broken")  # 展開失敗
    shutil.copy(RESP, resp / "ok_p0001.json")

    summary = run(inp, TPL, cfg, ReplayClient(resp))
    assert summary.pages == 3
    assert summary.rows == 3                       # 要件 §3.4

    _x, _c, rows = render(TPL, cfg, timestamp="g2")
    by_file = {r.source_file: r for r in rows}
    assert by_file["ok.png"].status == "正常"
    assert by_file["noresp.png"].status == "送信失敗"
    assert by_file["broken.pdf"].status == "展開失敗"
    assert by_file["noresp.png"].unclear_count == 214   # 全〓
    assert by_file["broken.pdf"].unclear_count == 214


def test_tr_g3_format_mismatch_all_unclear(tmp_path):
    """symbol を欄の無い場所へ集めた応答 → 様式不一致（D-15）＋全〓行。"""
    cfg = make_cfg(tmp_path)
    inp = tmp_path / "input"; inp.mkdir()
    resp = tmp_path / "resp"; resp.mkdir()
    shutil.copy(PAGE_PNG, inp / "a.png")

    d = json.loads(RESP.read_text(encoding="utf-8"))
    for page in d["fullTextAnnotation"]["pages"]:
        for b in page["blocks"]:
            for p in b["paragraphs"]:
                for w in p["words"]:
                    for s in w["symbols"]:
                        # front の欄外・除外外の帯（x=300, y=1700 付近）へ集める
                        s["boundingBox"]["vertices"] = [
                            {"x": 300, "y": 1700}, {"x": 310, "y": 1700},
                            {"x": 310, "y": 1712}, {"x": 300, "y": 1712}]
    (resp / "a_p0001.json").write_text(json.dumps(d, ensure_ascii=False),
                                       encoding="utf-8")
    run(inp, TPL, cfg, ReplayClient(resp))
    _x, _c, rows = render(TPL, cfg, timestamp="g3")
    assert rows[0].status == "様式不一致"
    assert rows[0].unclear_count == 214


def test_tr_g4_xlsx_csv_extract_columns_identical(tmp_path):
    """§8-12: 抽出対象の全列（現行 214）のセル値が .xlsx と .csv で全列一致する。"""
    import csv as csvmod
    cfg = make_cfg(tmp_path)
    run_one_page(tmp_path, cfg)
    xlsx, csvp, _rows = render(TPL, cfg, timestamp="g4")

    ws = load_workbook(xlsx)["output"]
    xrow = [c.value for c in list(ws.iter_rows(min_row=2))[0]][6:]
    with open(csvp, encoding="utf-8-sig", newline="") as f:
        crow = list(csvmod.reader(f))[1][6:]
    assert len(xrow) == len(crow) == 214
    for i, (xv, cv) in enumerate(zip(xrow, crow)):
        # xlsx の空セルは None・数値は int で読める。CSV 文字列へ正規化して比較
        norm = "" if xv is None else str(xv)
        assert norm == cv, f"列{i + 7} で不一致"


def test_tr_g5_purge_requires_yes(tmp_path):
    cfg_file = tmp_path / "config.json"
    wd = tmp_path / "wd"; wd.mkdir(); (wd / "x.txt").write_text("x")
    cfg_file.write_text(json.dumps({"workdir": str(wd),
                                    "log_dir": str(tmp_path / "logs")}),
                        encoding="utf-8")
    base = [str(PYTHON), "-X", "utf8", "-m", "chouhyo_ocr.cli",
            "--config", str(cfg_file), "purge"]
    r1 = subprocess.run(base, cwd=app_root() / "core", capture_output=True,
                        text=True, encoding="utf-8", timeout=60)
    assert r1.returncode == 1 and wd.exists() and (wd / "x.txt").exists()  # 拒否・消えない
    r2 = subprocess.run(base + ["--yes"], cwd=app_root() / "core",
                        capture_output=True, text=True, encoding="utf-8",
                        timeout=60)
    # 明示時は中身が消える（要件 §6.3）。workdir 自体は keep-list 方式（#83）
    # では cred.dpapi を残す余地のために残る——ここには cred.dpapi が無いので
    # 空フォルダとして残る
    assert r2.returncode == 0 and wd.exists() and not (wd / "x.txt").exists()


def test_tr_g6_verify_fails_without_credentials(tmp_path):
    import os
    cfg_file = tmp_path / "config.json"
    cfg_file.write_text(json.dumps({"workdir": str(tmp_path / "wd"),
                                    "output_dir": str(tmp_path / "out"),
                                    "log_dir": str(tmp_path / "logs")}),
                        encoding="utf-8")
    env = {k: v for k, v in os.environ.items()
           if k != "GOOGLE_APPLICATION_CREDENTIALS"}
    r = subprocess.run(
        [str(PYTHON), "-X", "utf8", "-m", "chouhyo_ocr.cli",
         "--config", str(cfg_file), "verify"],
        cwd=app_root() / "core", capture_output=True, text=True,
        encoding="utf-8", timeout=120, env=env)
    assert r.returncode == 1
    ev = [json.loads(l) for l in r.stdout.splitlines() if l.strip()]
    cred = next(e for e in ev if e.get("check") == "credentials")
    assert cred["ok"] is False and cred["state"] == "missing"
