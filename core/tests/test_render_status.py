"""render 段の失敗ステータス配線（issue #80・T-20〜T-29）。

08 §2.4.3 が「未配線」と注記していた `row_build_failed` を DB
（page.status / page.status_reason）と進捗イベントへ繋ぐ。あわせて
`row_build_bug`（コード欠陥の疑い）を分け、status は `様式不一致` ではなく
新設の `出力失敗` を使う。

このファイルで最も重要なのは T-24 —— render の失敗で `page.state` を
`failed` へ落とすと、次の run の todo に入って送信済みページが再送され、
利用者が二重に課金される。state を動かさないことをテストで固定する。

差し替え対象は `chouhyo_ocr.pipeline.build_row`（pipeline が
`from .render_rows import ... build_row` で自分の名前空間に取り込んでいる
ため、pipeline 側の名前を差し替えれば効く）。
test_format_check_pipeline.py の `pipeline_mod.assign` 差し替えと同じ流儀。
"""
import csv as csv_mod
import shutil

import pytest
from openpyxl import load_workbook

from chouhyo_ocr import logging_safe, render_rows
from chouhyo_ocr.config import Config
from chouhyo_ocr.paths import app_root
from chouhyo_ocr.pipeline import render, run
from chouhyo_ocr.pipeline_errors import OperationRefused
from chouhyo_ocr.store import Store
from chouhyo_ocr.vision_client import ReplayClient

TPL = app_root() / "templates" / "chouhyo-v1.json"
RESP = app_root() / "workdir" / "s2" / "resp_DOCUMENT_TEXT_DETECTION.json"
PAGE_PNG = app_root() / "workdir" / "pages" / "sample-1.png"
PAGE2_PNG = app_root() / "workdir" / "pages" / "sample-2.png"

needs_replay = pytest.mark.skipif(
    not (RESP.exists() and PAGE_PNG.exists()),
    reason="保存済み応答・サンプル画像が無い環境")


def _cfg(tmp_path) -> Config:
    return Config(unclear_threshold=0.4,
                  output_dir=str(tmp_path / "out"), workdir=str(tmp_path / "wd"),
                  log_dir=str(tmp_path / "logs"))


def _prepare(tmp_path, pages=("sample-1.png",)):
    """入力フォルダと replay 素材を用意して run を1回通す（送信は 0 円）。"""
    inp = tmp_path / "input"; inp.mkdir()
    replay = tmp_path / "responses"; replay.mkdir()
    src = {"sample-1.png": PAGE_PNG, "sample-2.png": PAGE2_PNG}
    for name in pages:
        shutil.copy(src[name], inp / name)
        shutil.copy(RESP, replay / f"{name[:-4]}_p0001.json")
    logging_safe.init(str(tmp_path / "logs"))
    cfg = _cfg(tmp_path)
    summary = run(inp, TPL, cfg, ReplayClient(replay))
    return inp, replay, cfg, summary


def _pages(cfg):
    with Store(cfg.workdir + "/intermediate.sqlite") as store:
        return {r["page_id"]: dict(r) for r in store.pages()}


def _boom(exc):
    def _f(*a, **kw):
        raise exc
    return _f


def _render_allowing_total_failure(cfg, timestamp, progress=None):
    """done ページが1枚しかない環境では「全滅」＝既存の OperationRefused に
    当たる（レビュー M-1・T-28 がその挙動そのものを固定する）。

    DB への書き込みも進捗イベントも拒否より前に済んでいるので、ここでは
    拒否を飲んで、そのあとの状態を検査する。
    """
    kw = {} if progress is None else {"progress": progress}
    try:
        return render(TPL, cfg, timestamp=timestamp, **kw)
    except OperationRefused:
        return None


# --- T-20 / T-21: 例外の種類で理由コードが割れる（決定13・許可リスト方式） ---

@needs_replay
def test_t20_data_error_writes_render_failed_status(tmp_path, monkeypatch):
    """T-20: build_row が ValueError（中間データの壊れ方として説明がつく）
    → page.status が「出力失敗」・status_reason が row_build_failed。
    """
    _inp, _replay, cfg, _s = _prepare(tmp_path)
    import chouhyo_ocr.pipeline as pipeline_mod
    monkeypatch.setattr(pipeline_mod, "build_row", _boom(ValueError("broken cell")))
    _render_allowing_total_failure(cfg, "t20")

    page = _pages(cfg)["sample-1_p0001"]
    assert page["status"] == "出力失敗"
    assert page["status_reason"] == "row_build_failed"
    # 様式の問題として記録しない（06 §7・利用者がテンプレートを疑う）
    assert page["status"] != render_rows.STATUS_FORMAT_MISMATCH


@needs_replay
def test_t21_type_error_is_treated_as_code_bug(tmp_path, monkeypatch):
    """T-21: TypeError は許可リストに無い＝コード欠陥の疑いへ倒す
    （fail-closed）。ログに row_build_bug とスタックが残る。
    """
    _inp, _replay, cfg, _s = _prepare(tmp_path)
    import chouhyo_ocr.pipeline as pipeline_mod
    monkeypatch.setattr(pipeline_mod, "build_row",
                        _boom(TypeError("unsupported operand")))
    _render_allowing_total_failure(cfg, "t21")

    page = _pages(cfg)["sample-1_p0001"]
    assert page["status"] == "出力失敗"
    assert page["status_reason"] == "row_build_bug"

    app_log = (tmp_path / "logs" / "app.log").read_text(encoding="utf-8")
    assert "row_build_bug" in app_log
    assert "row_build_bug_total" in app_log
    # 記入値をログへ出さない（§8.1）。例外メッセージ本文は残さず型名だけ
    assert "unsupported operand" not in app_log


# --- T-22 / T-23 / T-25: クリアの範囲（決定14） ---

@needs_replay
def test_t22_success_clears_the_mark_render_itself_left(tmp_path, monkeypatch):
    """T-22: 原因が直って再 render すると、印（status/status_reason）が
    剥がれて行が「正常」に戻る。08 §2.4.3 が未配線の理由に挙げていた残留。
    """
    _inp, _replay, cfg, _s = _prepare(tmp_path)
    import chouhyo_ocr.pipeline as pipeline_mod
    monkeypatch.setattr(pipeline_mod, "build_row", _boom(ValueError("broken")))
    _render_allowing_total_failure(cfg, "t22a")
    assert _pages(cfg)["sample-1_p0001"]["status"] == "出力失敗"

    monkeypatch.undo()
    _x, _c, rows = render(TPL, cfg, timestamp="t22b")
    page = _pages(cfg)["sample-1_p0001"]
    assert page["status"] == ""
    assert page["status_reason"] == ""
    assert rows[0].status == render_rows.STATUS_OK


@needs_replay
def test_t23_render_does_not_touch_marks_from_other_paths(tmp_path, monkeypatch):
    """T-23: run が付けた status（様式不一致 / frame_lines・state=failed）は
    render を通しても変わらない。render がクリアしてよいのは自分の語彙だけ。
    """
    _inp, _replay, cfg, _s = _prepare(tmp_path)
    db = cfg.workdir + "/intermediate.sqlite"
    with Store(db) as store:
        store.set_status("sample-1_p0001",
                         render_rows.STATUS_FORMAT_MISMATCH, reason="frame_lines")
        store.set_state("sample-1_p0001", "failed")

    render(TPL, cfg, timestamp="t23")

    page = _pages(cfg)["sample-1_p0001"]
    assert page["status"] == render_rows.STATUS_FORMAT_MISMATCH
    assert page["status_reason"] == "frame_lines"
    assert page["state"] == "failed"


@needs_replay
def test_t25_status_round_trips_across_repeated_renders(tmp_path, monkeypatch):
    """T-25: 失敗 → 成功 → 失敗 と 3 回 render しても印が正しく往復する。"""
    _inp, _replay, cfg, _s = _prepare(tmp_path)
    import chouhyo_ocr.pipeline as pipeline_mod
    pid = "sample-1_p0001"

    monkeypatch.setattr(pipeline_mod, "build_row", _boom(KeyError("missing")))
    _render_allowing_total_failure(cfg, "r1")
    assert _pages(cfg)[pid]["status_reason"] == "row_build_failed"

    monkeypatch.undo()
    render(TPL, cfg, timestamp="r2")
    assert _pages(cfg)[pid]["status"] == ""

    monkeypatch.setattr(pipeline_mod, "build_row", _boom(RuntimeError("bug")))
    _render_allowing_total_failure(cfg, "r3")
    assert _pages(cfg)[pid]["status_reason"] == "row_build_bug"


# --- T-24: 再課金しない（最重要） ---

@needs_replay
def test_t24_render_failure_does_not_resend_on_next_run(tmp_path, monkeypatch):
    """T-24（Go 条件1）: render が失敗しても state は done のまま。
    次の run で todo に入らず api_calls が増えない＝送信済みページを
    再送しない（＝再課金しない）。

    state を failed へ落とすと `todo = [p for p in all_pages if p["state"]
    not in ("done","skipped_duplicate")]` に入り、この不変条件が壊れる。
    """
    inp, replay, cfg, first = _prepare(tmp_path)
    assert first.api_calls == 1

    import chouhyo_ocr.pipeline as pipeline_mod
    monkeypatch.setattr(pipeline_mod, "build_row", _boom(ValueError("broken")))
    _render_allowing_total_failure(cfg, "t24")
    page = _pages(cfg)["sample-1_p0001"]
    assert page["status"] == "出力失敗"
    assert page["state"] == "done", "render の失敗で state を動かしてはいけない"

    monkeypatch.undo()
    events: list[dict] = []
    second = run(inp, TPL, cfg, ReplayClient(replay), events.append)
    assert second.api_calls == 0, "送信済みページが再送された（再課金）"
    start_ev = next(e for e in events if e.get("event") == "start")
    assert start_ev.get("todo", 0) == 0
    assert second.reused_pages == 1


# --- T-26 / T-28: バッチを止めない・全滅は拒否 ---

@pytest.mark.skipif(not (RESP.exists() and PAGE_PNG.exists() and PAGE2_PNG.exists()),
                    reason="サンプル画像2枚が無い環境")
def test_t26_one_broken_page_does_not_lose_the_other(tmp_path, monkeypatch):
    """T-26（issue #39 の維持）: 2 ページ中 1 ページだけ失敗しても、
    もう 1 ページの行は正常に出て xlsx も生成される。
    """
    _inp, _replay, cfg, _s = _prepare(tmp_path, pages=("sample-1.png", "sample-2.png"))
    import chouhyo_ocr.pipeline as pipeline_mod
    orig = pipeline_mod.build_row

    def _selective(template, page, *a, **kw):
        if page["page_id"] == "sample-1_p0001":
            raise ValueError("broken")
        return orig(template, page, *a, **kw)

    monkeypatch.setattr(pipeline_mod, "build_row", _selective)
    xlsx, csvp, rows = render(TPL, cfg, timestamp="t26")

    assert xlsx.exists() and csvp.exists()
    assert len(rows) == 2
    by_id = {r.page_id: r for r in rows}
    assert by_id["sample-1_p0001"].status == "出力失敗"
    assert by_id["sample-2_p0001"].status != "出力失敗"
    pages = _pages(cfg)
    assert pages["sample-2_p0001"]["status"] == ""


@needs_replay
def test_t28_all_done_pages_failing_still_refuses(tmp_path, monkeypatch):
    """T-28: done ページが全滅なら従来どおり OperationRefused（文言も維持）。"""
    _inp, _replay, cfg, _s = _prepare(tmp_path)
    import chouhyo_ocr.pipeline as pipeline_mod
    monkeypatch.setattr(pipeline_mod, "build_row", _boom(ValueError("broken")))
    with pytest.raises(OperationRefused) as ei:
        render(TPL, cfg, timestamp="t28")
    assert "すべてで行の組み立てに失敗した" in str(ei.value)


# --- T-27: 進捗イベント（決定15） ---

@needs_replay
def test_t27_progress_events_do_not_move_the_progress_bar(tmp_path, monkeypatch):
    """T-27: render_page_failed が失敗ページ数だけ・render_summary が 1 回。
    `event: "page"` は 1 件も出さない —— GUI は page で進捗バーを +1 する
    ので、run 末尾の render が page を出すとバーが二重に進む。
    """
    _inp, _replay, cfg, _s = _prepare(tmp_path)
    import chouhyo_ocr.pipeline as pipeline_mod
    monkeypatch.setattr(pipeline_mod, "build_row", _boom(ValueError("broken")))
    events: list[dict] = []
    _render_allowing_total_failure(cfg, "t27", progress=events.append)

    failed = [e for e in events if e["event"] == "render_page_failed"]
    assert len(failed) == 1
    assert failed[0]["page_id"] == "sample-1_p0001"
    assert failed[0]["status"] == "出力失敗"
    assert failed[0]["reason_code"] == "row_build_failed"
    # 記入値・列名・face_id を進捗へ載せない（§8.1）
    assert set(failed[0]) == {"event", "page_id", "status", "reason_code"}
    assert [e for e in events if e["event"] == "page"] == []


@needs_replay
def test_t27b_render_summary_is_emitted_once_on_success(tmp_path):
    """T-27 続き: 成功時は render_summary が 1 回だけ出て、失敗の内訳が 0。"""
    _inp, _replay, cfg, _s = _prepare(tmp_path)
    events: list[dict] = []
    render(TPL, cfg, timestamp="t27b", progress=events.append)
    summaries = [e for e in events if e["event"] == "render_summary"]
    assert len(summaries) == 1
    assert summaries[0]["row_build_failed"] == 0
    assert summaries[0]["row_build_bug"] == 0
    assert summaries[0]["pages"] == 1
    assert [e for e in events if e["event"] == "page"] == []


# --- T-29: 決定16（busy_timeout）の裏取り ---

def test_t29_store_sets_busy_timeout(tmp_path):
    """T-29 前段: Store が busy_timeout を立てている（既定 0 = 即例外）。"""
    with Store(tmp_path / "s.sqlite") as store:
        (v,) = store.con.execute("PRAGMA busy_timeout").fetchone()
    assert v >= 5000


@needs_replay
def test_t29_run_tail_render_writes_status_without_lock_error(tmp_path, monkeypatch):
    """T-29: run は自分の Store を開いたまま末尾で _render_locked を呼び、
    そこが page.status を書く。同一 DB への別接続なので、busy_timeout が
    無いとロック衝突が即 `database is locked` になる。

    ※このテストは衝突を確実に再現させるものではない（タイミング依存）。
    「run 経路から render の書き込みが通り、status が DB に残る」ことを
    固定するのが目的。
    """
    inp = tmp_path / "input"; inp.mkdir()
    shutil.copy(PAGE_PNG, inp / "sample-1.png")
    replay = tmp_path / "responses"; replay.mkdir()
    shutil.copy(RESP, replay / "sample-1_p0001.json")
    logging_safe.init(str(tmp_path / "logs"))
    cfg = _cfg(tmp_path)

    import chouhyo_ocr.pipeline as pipeline_mod
    monkeypatch.setattr(pipeline_mod, "build_row", _boom(ValueError("broken")))
    events: list[dict] = []
    try:
        run(inp, TPL, cfg, ReplayClient(replay), events.append)
    except OperationRefused:
        pass  # 全滅拒否は既存挙動。書き込み自体はその前に済んでいる

    page = _pages(cfg)["sample-1_p0001"]
    assert page["status"] == "出力失敗"
    assert page["state"] == "done"
    assert [e for e in events if e["event"] == "render_page_failed"]


# --- Go 条件3: _FAILURE_STATUSES への値追加が #62 の〓集計を変えない ---

def test_new_status_is_a_failure_status_like_the_old_one():
    """`出力失敗` が失敗系集合に入り、compose_status の扱いが
    `様式不一致` と同じであること（全〓行になる側）。
    """
    assert render_rows.STATUS_RENDER_FAILED in render_rows._FAILURE_STATUSES
    for below, processed in ((0, False), (0, True), (5, True)):
        assert (render_rows.compose_status(render_rows.STATUS_RENDER_FAILED,
                                           below, processed)
                == render_rows.compose_status(render_rows.STATUS_FORMAT_MISMATCH,
                                              below, processed)
                .replace(render_rows.STATUS_FORMAT_MISMATCH,
                         render_rows.STATUS_RENDER_FAILED))


@pytest.mark.parametrize("char_level", [False, True])
def test_unclear_aggregation_is_unchanged_by_the_new_status(tmp_path, char_level):
    """Go 条件3: #62 の〓集計3経路（xlsx の COUNTIF・条件付き書式・csv）の
    件数が `出力失敗` の追加で変わらない。文字単位〓（#62）の ON/OFF 両方で見る
    ——COUNTIF のワイルドカード化と2本目の条件付き書式はこのゲートの中にある。

    〓の数はセルの値から数えるもので、ステータス文字列には依存しない。
    同じページを `様式不一致` と `出力失敗` の2通りで出力し、ステータス列の
    語だけを置き換えれば xlsx・csv とも完全に一致することで確認する。
    """
    from chouhyo_ocr.columns import META_COLUMNS
    from chouhyo_ocr.render_out import write_outputs

    columns = ["氏名", "住所", *META_COLUMNS]
    n = len(columns) - len(META_COLUMNS)

    def _rows(status):
        return [render_rows.Row("p1", "a.png", 1, status,
                                [render_rows.UNCLEAR] * n, n, "", ("",) * n)]

    old_x, old_c, _ = write_outputs(tmp_path / "old", "t", columns,
                                    _rows(render_rows.STATUS_FORMAT_MISMATCH),
                                    unclear_char_level=char_level)
    new_x, new_c, _ = write_outputs(tmp_path / "new", "t", columns,
                                    _rows(render_rows.STATUS_RENDER_FAILED),
                                    unclear_char_level=char_level)

    def _swap(v):
        return (v.replace(render_rows.STATUS_RENDER_FAILED,
                          render_rows.STATUS_FORMAT_MISMATCH)
                if isinstance(v, str) else v)

    old_ws = load_workbook(old_x)["output"]
    new_ws = load_workbook(new_x)["output"]
    assert [[c.value for c in r] for r in old_ws.iter_rows()] == \
           [[_swap(c.value) for c in r] for r in new_ws.iter_rows()]
    # 条件付き書式（〓の色分け）の本数も変わらない
    assert (len(list(old_ws.conditional_formatting))
            == len(list(new_ws.conditional_formatting)))

    with open(old_c, encoding="utf-8-sig", newline="") as f:
        old_csv = list(csv_mod.reader(f))
    with open(new_c, encoding="utf-8-sig", newline="") as f:
        new_csv = [[_swap(v) for v in row] for row in csv_mod.reader(f)]
    assert old_csv == new_csv
