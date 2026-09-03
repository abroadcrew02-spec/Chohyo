"""出力列制御 MVP（issue #66）第2弾: 列の並べ替えの core 受入テスト。

QA裏取り実測（scratchpad/qa_reorder_probe.py・2026-09-01・トワ「全通過確認済み」）
を正式な pytest テストへ移植したもの。手順は実測スクリプトと同一（出荷テンプレを
tmp へ複製 → 配列順だけ入替 → derive_columns / render / remap で観測）。

対象 AC（05_output_columns_requirements.md §7.3）:
  AC-2.1  面内の欄どうしの並べ替えで列順が変わり、220列・管理6列先頭・
          列名集合は不変
  AC-2.3  表の内部列の並べ替えが全行へ一括反映される（家族欄=60列維持）
  AC-2.6  geometry_hash 不変・template_hash 変化・render が remap を
          名指しして拒否・remap 成立・API 送信カウンタ増分0
  AC-2.11 並べ替え後の同一中間データでの再出力がバイト一致する（NFR-04）
  AC-2.12 xlsx ヘッダ＝並べ替え後 derive_columns・xlsx↔csv のヘッダ列名の
          並び一致・抽出列セル値の不一致0件
  AC-2.13 JSON 直接編集の配列順がそのまま列順になる（NFR-03・core視点）

GUI（並べ替えUI・上下ボタン・Undo統合・出力列リストのハイライト等）は対象外。
本ファイルが固定するのは「配列順を変えたときに core がどう振る舞うか」の
契約のみ——並べ替え UI の実装そのものは検証しない。
"""
import csv as csvmod
import json
import re
import shutil

import pytest
from openpyxl import load_workbook

from chouhyo_ocr import align, api_budget
from chouhyo_ocr.columns import META_COLUMNS, derive_columns
from chouhyo_ocr.config import Config
from chouhyo_ocr.paths import app_root
from chouhyo_ocr.pipeline import OperationRefused, remap, render, run
from chouhyo_ocr.template import load_template
from chouhyo_ocr.vision_client import ReplayClient

TPL = app_root() / "templates" / "chouhyo-v1.json"
RESP = app_root() / "workdir" / "s2" / "resp_DOCUMENT_TEXT_DETECTION.json"
PAGE_PNG = app_root() / "workdir" / "pages" / "sample-1.png"

needs_replay = pytest.mark.skipif(
    not (RESP.exists() and PAGE_PNG.exists()), reason="保存済み応答が無い環境")


# ---------- 共通ヘルパ（core/tests に conftest.py が無いため各ファイルで複製） ----------

def _raw() -> dict:
    return json.loads(TPL.read_text(encoding="utf-8"))


def _write(tmp_path, data, name):
    p = tmp_path / name
    p.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return p


def _swapped_fields_template(tmp_path):
    """front 面の単発欄（fields 配列）の先頭2つを入れ替えたテンプレート。

    AC-2.1・AC-2.6・AC-2.11・AC-2.12・AC-2.13 に共通の素材（QA実測と同一手順）。
    戻り値は (テンプレートパス, 入れ替えたfield_id A, B)。
    """
    raw = _raw()
    f0 = raw["faces"][0]["fields"]
    a, b = f0[0]["field_id"], f0[1]["field_id"]
    f0[0], f0[1] = f0[1], f0[0]
    return _write(tmp_path, raw, "swap.json"), a, b


@pytest.fixture()
def isolated_counter(tmp_path, monkeypatch):
    """API 送信カウンタを一時ディレクトリへ隔離する（実カウンタを汚さない・
    test_api_budget.py の isolated_counter fixture と同内容を複製）。
    """
    monkeypatch.setenv("CHOUHYO_USAGE_DIR_FOR_TESTS", str(tmp_path / "usage"))
    return tmp_path


# ---------- AC-2.1 ----------

def test_ac_2_1_single_field_swap_reorders_columns_keeps_set_and_meta(tmp_path):
    """AC-2.1: 面内の欄どうしの並べ替え（fields 配列順の入替）で列順が変わり、
    220列・管理6列先頭・列名集合（増減も改名もない）は不変。
    """
    base_cols = derive_columns(load_template(TPL))
    tpl_sw, a, b = _swapped_fields_template(tmp_path)
    new_cols = derive_columns(load_template(tpl_sw))

    assert len(new_cols) == len(base_cols) == 220
    assert new_cols[:6] == list(META_COLUMNS)
    assert set(new_cols) == set(base_cols)
    assert new_cols != base_cols

    ia_old, ib_old = base_cols.index(a), base_cols.index(b)
    ia_new, ib_new = new_cols.index(a), new_cols.index(b)
    assert (ia_old < ib_old) != (ia_new < ib_new), (
        f"入れ替えが列順へ反映されていない（旧 {ia_old}<{ib_old} / 新 {ia_new}<{ib_new}）")


# ---------- AC-2.3 ----------

def test_ac_2_3_table_column_swap_reflects_in_all_rows_keeps_column_count(tmp_path):
    """AC-2.3: 表の内部列の並べ替え（tables[].columns 配列順の入替）が
    全行へ一括反映され、表由来の出力列数（家族欄=60列: 10行×6列）は
    不変（1回の並べ替えが行ごとにバラバラに効くことはない）。
    """
    base_cols = derive_columns(load_template(TPL))
    raw = _raw()
    tbl = next(t for f in raw["faces"] for t in f.get("tables", []) if len(t["columns"]) >= 2)
    assert tbl["table_id"] == "family"  # 前提: front 最初の表（QA実測と同じ経路）
    c0, c1 = tbl["columns"][0]["name"], tbl["columns"][1]["name"]
    tbl["columns"][0], tbl["columns"][1] = tbl["columns"][1], tbl["columns"][0]
    tpl_sw = _write(tmp_path, raw, "swap_col.json")
    cols2 = derive_columns(load_template(tpl_sw))

    tid = tbl["table_id"]
    rows_old = [c for c in base_cols if c.startswith(tid + "_")]
    rows_new = [c for c in cols2 if c.startswith(tid + "_")]
    assert len(rows_old) == len(rows_new) == 60  # 表由来の列数は不変

    def per_row(cols, tid):
        d: dict[str, list[str]] = {}
        for c in cols:
            m = re.match(rf"^{re.escape(tid)}_(\d+)_(.*)$", c)
            if m:
                d.setdefault(m.group(1), []).append(m.group(2))
        return d

    old_pr, new_pr = per_row(base_cols, tid), per_row(cols2, tid)
    nrows = len(new_pr)
    assert nrows == 10
    flipped = sum(1 for v in new_pr.values()
                 if c1 in v and c0 in v and v.index(c1) < v.index(c0))
    was = sum(1 for v in old_pr.values()
             if c1 in v and c0 in v and v.index(c1) < v.index(c0))
    assert flipped == nrows and was == 0, (
        f"表 {tid}: 1回の並べ替えが全{nrows}行に反映されていない"
        f"（入替後 {flipped}/{nrows}行・入替前 {was}/{nrows}行）")


# ---------- AC-2.6 ----------

def test_ac_2_6_geometry_hash_unchanged_but_template_hash_changes_on_reorder(tmp_path):
    """AC-2.6（前半）: 並べ替え（配列順の変更）は幾何セクションに触れないため
    geometry_hash は不変。template_hash は変化する（remap 要求の根拠）。
    """
    tpl_sw, _a, _b = _swapped_fields_template(tmp_path)
    gh_base = align.geometry_hash(json.loads(TPL.read_text(encoding="utf-8")))
    gh_swap = align.geometry_hash(json.loads(tpl_sw.read_text(encoding="utf-8")))
    th_base = align.template_hash(json.loads(TPL.read_text(encoding="utf-8")))
    th_swap = align.template_hash(json.loads(tpl_sw.read_text(encoding="utf-8")))
    assert gh_base == gh_swap
    assert th_base != th_swap


@needs_replay
def test_ac_2_6_render_requires_remap_and_reorder_costs_no_resend(tmp_path, isolated_counter):
    """AC-2.6（後半）: 並べ替えテンプレでの render は remap を名指しして拒否する。
    remap は成立し（幾何が不変なので拒否されない）、API 送信カウンタの増分は0
    （isolated_counter で実測・NFR-01 の並べ替え版: 配列順の変更だけでは
    課金が発生しない）。
    """
    tpl_sw, _a, _b = _swapped_fields_template(tmp_path)
    cfg = Config(unclear_threshold=0.4, output_dir=str(tmp_path / "out"),
                workdir=str(tmp_path / "wd"), log_dir=str(tmp_path / "logs"))
    inp = tmp_path / "input"; inp.mkdir()
    resp = tmp_path / "resp"; resp.mkdir()
    shutil.copy(PAGE_PNG, inp / "a.png")
    shutil.copy(RESP, resp / "a_p0001.json")
    run(inp, TPL, cfg, ReplayClient(resp))
    used0 = api_budget.used_this_month()

    with pytest.raises(OperationRefused, match="remap"):
        render(tpl_sw, cfg, timestamp="try")
    assert remap(tpl_sw, cfg) == 1
    assert api_budget.used_this_month() == used0


# ---------- AC-2.11 / AC-2.12 ----------

@needs_replay
def test_ac_2_11_and_ac_2_12_byte_identical_rerender_and_xlsx_csv_match_after_reorder(
        tmp_path):
    """AC-2.11: 並べ替え後・同一中間データでの再出力がバイト一致する（NFR-04・
    §6.2 決定性）。**列構成を変えた状態を新たな固定条件にする**——既存 T-29 は
    列不変を前提にした資産で本機能により無効化されている（05 §8 の方針）。
    このテストがその置き換え先として、並べ替え後の状態を新たな固定条件にする。

    AC-2.12: xlsx ヘッダが並べ替え後の derive_columns と一致し、xlsx↔csv の
    ヘッダ列名の並び・抽出列セル値が全一致する（受入基準12の回帰）。
    """
    tpl_sw, _a, _b = _swapped_fields_template(tmp_path)
    new_cols = derive_columns(load_template(tpl_sw))

    cfg = Config(unclear_threshold=0.4, output_dir=str(tmp_path / "out"),
                workdir=str(tmp_path / "wd"), log_dir=str(tmp_path / "logs"))
    inp = tmp_path / "input"; inp.mkdir()
    resp = tmp_path / "resp"; resp.mkdir()
    shutil.copy(PAGE_PNG, inp / "a.png")
    shutil.copy(RESP, resp / "a_p0001.json")
    run(inp, TPL, cfg, ReplayClient(resp))
    assert remap(tpl_sw, cfg) == 1

    x1, c1p, _r1 = render(tpl_sw, cfg, timestamp="r1")
    x2, c2p, _r2 = render(tpl_sw, cfg, timestamp="r2")
    assert x1.read_bytes() == x2.read_bytes()    # AC-2.11: xlsx バイト一致
    assert c1p.read_bytes() == c2p.read_bytes()  # AC-2.11: csv バイト一致

    ws = load_workbook(x1)["output"]
    header = [c.value for c in list(ws.iter_rows(min_row=1))[0]]
    xrow = [c.value for c in list(ws.iter_rows(min_row=2))[0]][len(META_COLUMNS):]
    with open(c1p, encoding="utf-8-sig", newline="") as fh:
        rows = list(csvmod.reader(fh))
    chead, crow = rows[0][len(META_COLUMNS):], rows[1][len(META_COLUMNS):]

    assert header == new_cols                           # AC-2.12: xlsx ヘッダ＝derive_columns
    assert header[len(META_COLUMNS):] == chead           # AC-2.12: xlsx↔csv ヘッダ列名の並び一致
    mism = [i for i, (xv, cv) in enumerate(zip(xrow, crow))
           if ("" if xv is None else str(xv)) != cv]
    assert not mism, f"{len(xrow)}列中 不一致 {len(mism)}件"  # AC-2.12: セル値の不一致0件


# ---------- AC-2.13 ----------

def test_ac_2_13_json_direct_edit_array_order_becomes_column_order(tmp_path):
    """AC-2.13: JSON 直接編集で配列順を入れ替えた結果が画面（GUI）経由と一致する
    （NFR-03）。GUI は本ファイルの対象外のため、ここでは「core は JSON 経由か
    画面経由かをそもそも区別しない（テンプレートの配列順だけを見る）」という
    core 側の契約——JSON 直接編集の配列順がそのまま列順になる（AC-2.1 と
    同一経路）ことを固定する。画面側の等価性そのものは GUI テストが別途担う。
    """
    base_cols = derive_columns(load_template(TPL))
    tpl_sw, _a, _b = _swapped_fields_template(tmp_path)
    new_cols = derive_columns(load_template(tpl_sw))
    assert new_cols != base_cols               # 並べ替えが反映されている
    assert set(new_cols) == set(base_cols)     # 増減・改名は無い（並べ替えのみ）
