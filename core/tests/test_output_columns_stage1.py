"""出力列制御 MVP（issue #66）段1: core の `output` 属性一式の受入基準を固定する。

要件の正本: docs/design/chouhyo-ocr/05_output_columns_requirements.md §7.2。
段1 の範囲は付録 C のとおり: schema 定義／template.py の読み込み・検証／
output_cells() 1関数への集約／columns.py・render_rows.py の3経路／
validate_v1 の抽出列0拒否。

このファイルで固定する AC:
  - AC-1.1  output:false の欄が列に出ず、テンプレート側の定義（rect・kind・
            choice_marks・normalize）は残る
  - AC-1.2  subfields を持つ表の列を対象外にすると、由来サブ列が全行一括で
            消える（一部だけ残らない）
  - AC-1.3  対象外を解除して再出力しても API 送信は増えず、check_reusable が
            remap を要求し、当該列が対象外化前と同一の値で復帰する
  - AC-1.9  抽出対象列が0になるテンプレートは validate_v1 で拒否される
  - AC-1.11 （core 部分）COUNTIF 範囲・条件付き書式2本・csv 静的カウントが
            導出後の列数へ追従する（unclear_char_level の ON/OFF 両方）
  - AC-1.13 output を持たない出荷テンプレートは全欄 output=True・220/214列
  - AC-1.14 未定義の属性（例: outputt）を持つテンプレートは読み込み拒否
  - AC-1.17 対象外欄を1つ含む状態でも再出力がバイト一致する（決定性）
  - AC-1.19 対象外適用後も xlsx と csv の抽出対象列のセル値が一致する
            （列数非依存の突合）
  - AC-1.20 Row.origins が values と同時にスキップされ、対象外でない他の
            セルの由来色は生き続ける

段2 に送るもの（このファイルの対象外）: W-3 警告文の印・対象外欄由来カウンタの
run/remap 両配線・debug_images._field_origins との2経路一致（AC-1.4〜1.8,
1.10, 1.12, 1.20 の debug_images 側）。
"""
import csv as csvmod
import json
import shutil
import zipfile

import pytest
from openpyxl import load_workbook

from chouhyo_ocr import api_budget
from chouhyo_ocr.columns import (META_COLUMNS, derive_columns, excel_column_letter,
                                 extract_columns, validate_v1)
from chouhyo_ocr.config import Config
from chouhyo_ocr.paths import app_root
from chouhyo_ocr.pipeline import OperationRefused, remap, render, run
from chouhyo_ocr.render_out import write_outputs
from chouhyo_ocr.render_rows import UNCLEAR, Row, build_row
from chouhyo_ocr.template import TemplateError, load_template, output_cells
from chouhyo_ocr.vision_client import ReplayClient

TPL = app_root() / "templates" / "chouhyo-v1.json"
RESP = app_root() / "workdir" / "s2" / "resp_DOCUMENT_TEXT_DETECTION.json"
PAGE_PNG = app_root() / "workdir" / "pages" / "sample-1.png"

needs_replay = pytest.mark.skipif(
    not (RESP.exists() and PAGE_PNG.exists()), reason="保存済み応答が無い環境")


# ---------- 共通ヘルパ ----------

def _raw() -> dict:
    return json.loads(TPL.read_text(encoding="utf-8"))


def _write(tmp_path, data, name):
    p = tmp_path / name
    p.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return p


def _disable_field(raw, field_id):
    fld = next(f for f in raw["faces"][0]["fields"] if f["field_id"] == field_id)
    fld["output"] = False
    return fld


def _disable_table_column(raw, table_id, col_name):
    tbl = next(t for f in raw["faces"] for t in f.get("tables", []) if t["table_id"] == table_id)
    col = next(c for c in tbl["columns"] if c["name"] == col_name)
    col["output"] = False
    return col


def make_cfg(tmp_path) -> Config:
    return Config(unclear_threshold=0.4, output_dir=str(tmp_path / "out"),
                  workdir=str(tmp_path / "wd"), log_dir=str(tmp_path / "logs"))


def setup_done(tmp_path, cfg):
    """実応答（保存済み・課金ゼロ）で1ページ分の run を済ませる。"""
    inp = tmp_path / "input"; inp.mkdir()
    resp = tmp_path / "resp"; resp.mkdir()
    shutil.copy(PAGE_PNG, inp / "a.png")
    shutil.copy(RESP, resp / "a_p0001.json")
    run(inp, TPL, cfg, ReplayClient(resp))
    return inp, resp


def _sheet_xml(path):
    z = zipfile.ZipFile(path)
    return z.read("xl/worksheets/sheet1.xml").decode("utf-8")


@pytest.fixture()
def isolated_counter(tmp_path, monkeypatch):
    """API 送信カウンタを一時ディレクトリへ隔離する（P4-2: test_api_budget.py
    の isolated_counter fixture を流用する方針・conftest.py が無いため同内容
    をここに複製する。実カウンタを汚さない）。
    """
    monkeypatch.setenv("CHOUHYO_USAGE_DIR_FOR_TESTS", str(tmp_path / "usage"))
    return tmp_path


def page(status="", below=0):
    return {"page_id": "p_0001", "source_file": "s.png", "page_no": 1,
            "status": status, "unassigned_below_table": below}


CFG = Config(unclear_threshold=0.85, era_threshold=0.06)


# ---------- AC-1.1 ----------

def test_ac_1_1_output_false_excludes_column_but_retains_definition(tmp_path):
    baseline = load_template(TPL)
    base_by_id = {c.field_id: c for c in baseline.cells}

    raw = _raw()
    choice_fld = next(f for f in raw["faces"][0]["fields"] if f["kind"] == "choice")
    choice_fld["output"] = False
    choice_id = choice_fld["field_id"]
    _disable_field(raw, "person_電話番号")
    _disable_table_column(raw, "detail", "金額")

    t = load_template(_write(tmp_path, raw, "ac11.json"))
    cols = derive_columns(t)

    assert choice_id not in cols
    assert "person_電話番号" not in cols
    assert not any(c.startswith("detail_") and c.endswith("_金額") for c in cols)

    # 定義（rect・kind・choice_marks・normalize）は変わらず残る
    for fid in (choice_id, "person_電話番号"):
        cell = next(c for c in t.cells if c.field_id == fid)
        base = base_by_id[fid]
        assert cell.output is False
        assert cell.rect == base.rect
        assert cell.kind == base.kind
        assert cell.choice_marks == base.choice_marks
        assert cell.normalize == base.normalize

    detail_amount_cells = [c for c in t.cells
                           if c.table_id == "detail" and c.field_id.endswith("_金額")]
    assert len(detail_amount_cells) == 28  # 明細28行分、欄自体は消えない
    assert all(c.output is False for c in detail_amount_cells)
    assert all(c.normalize == "amount" for c in detail_amount_cells)
    base_amount = {c.field_id: c for c in baseline.cells
                   if c.table_id == "detail" and c.field_id.endswith("_金額")}
    for c in detail_amount_cells:
        assert c.rect == base_amount[c.field_id].rect


def test_output_cells_collects_only_output_true(tmp_path):
    """T3: output_cells() が対象外判定の唯一の窓口であることの直接検証。"""
    raw = _raw()
    _disable_field(raw, "person_電話番号")
    t = load_template(_write(tmp_path, raw, "oc.json"))

    oc = output_cells(t)
    assert all(c.output for c in oc)
    assert "person_電話番号" not in {c.field_id for c in oc}
    assert len(oc) == len(t.cells) - 1


# ---------- AC-1.2 ----------

def test_ac_1_2_subfields_disappear_together_not_partially(tmp_path):
    baseline_cols = set(derive_columns(load_template(TPL)))

    raw = _raw()
    _disable_table_column(raw, "family", "生年月日")
    t = load_template(_write(tmp_path, raw, "ac12.json"))
    cols = set(derive_columns(t))

    removed = baseline_cols - cols
    expected_removed = {
        c for c in baseline_cols
        if c.startswith("family_") and c.split("_", 2)[-1] in
        ("生年月日_年", "生年月日_月", "生年月日_日")}
    assert len(expected_removed) == 30  # 10行 × 3サブフィールド
    assert removed == expected_removed  # 過不足なく一致（一部だけ残らない）

    dob_cells = [c for c in t.cells
                if c.table_id == "family" and c.field_id.endswith("_生年月日")]
    assert len(dob_cells) == 10
    assert all(c.output is False for c in dob_cells)
    assert all(c.subfields == ("年", "月", "日") for c in dob_cells)
    # output_columns() 自体は変わらず3列名を返す（弾くのは output_cells 側）
    assert all(len(c.output_columns()) == 3 for c in dob_cells)


# ---------- AC-1.9 ----------

def test_ac_1_9_rejects_template_with_zero_output_columns(tmp_path):
    raw = _raw()
    for face in raw["faces"]:
        for fld in face.get("fields", []):
            fld["output"] = False
        for tb in face.get("tables", []):
            for col in tb["columns"]:
                col["output"] = False
    t = load_template(_write(tmp_path, raw, "all_off.json"))
    assert extract_columns(t) == []
    with pytest.raises(TemplateError, match="出力対象"):
        validate_v1(t)


def test_ac_1_9_accepts_when_at_least_one_column_remains(tmp_path):
    """境界: 1欄でも output:true が残れば拒否しない（拒否は0のときだけ）。"""
    raw = _raw()
    for face in raw["faces"]:
        for fld in face.get("fields", []):
            fld["output"] = False
        for tb in face.get("tables", []):
            for col in tb["columns"]:
                col["output"] = False
    # detail の品目だけ残す
    detail = next(t for f in raw["faces"] for t in f.get("tables", []) if t["table_id"] == "detail")
    next(c for c in detail["columns"] if c["name"] == "品目")["output"] = True
    t = load_template(_write(tmp_path, raw, "one_on.json"))
    cols = validate_v1(t)
    assert len(cols) == len(META_COLUMNS) + 28  # 品目 × 28行


# ---------- AC-1.13 ----------

def test_ac_1_13_unmodified_template_all_cells_output_true():
    t = load_template(TPL)
    assert all(c.output is True for c in t.cells)
    assert len(derive_columns(t)) == 220
    assert len(extract_columns(t)) == 214
    assert len(output_cells(t)) == len(t.cells)


# ---------- AC-1.14 ----------

def test_ac_1_14_rejects_unknown_field_attribute(tmp_path):
    raw = _raw()
    raw["faces"][0]["fields"][0]["outputt"] = False  # 属性名のtypo
    with pytest.raises(TemplateError, match="スキーマ検証エラー"):
        load_template(_write(tmp_path, raw, "typo_field.json"))


def test_ac_1_14_rejects_unknown_table_column_attribute(tmp_path):
    raw = _raw()
    tbl = next(t for f in raw["faces"] for t in f.get("tables", []) if t["table_id"] == "family")
    tbl["columns"][0]["outputt"] = False
    with pytest.raises(TemplateError, match="スキーマ検証エラー"):
        load_template(_write(tmp_path, raw, "typo_col.json"))


# ---------- AC-1.20（render_rows レベル。debug_images 側は段2） ----------

def test_ac_1_20_excluded_cell_origin_and_value_skip_together(tmp_path):
    raw = _raw()
    _disable_field(raw, "person_電話番号")
    t = load_template(_write(tmp_path, raw, "ac120.json"))

    cells = {c.field_id: ("", None, c.kind, False) for c in t.cells}
    cells["person_氏名"] = ("テスト太郎", 0.95, "text", False)
    cells["person_郵便番号1"] = ("262-0032", 0.95, "text", False)
    extras = {"person_郵便番号1": ("", "fallback")}

    row = build_row(t, page(), cells, {}, CFG, extras=extras)

    extract_cols = derive_columns(t)[len(META_COLUMNS):]
    assert "person_電話番号" not in extract_cols
    # values と origins は常に同じ長さ・同じ列集合に対応する（対で欠落する）
    assert len(row.values) == len(row.origins) == len(extract_cols)

    by_col = dict(zip(extract_cols, row.origins))
    assert by_col["person_郵便番号1"] == "fallback"  # 対象外でないセルの由来色は生きている


# ---------- AC-1.11（core 部分。write_outputs を直接叩き、実応答に依存しない） ----------

def test_ac_1_11_countif_and_formatting_track_reduced_columns(tmp_path):
    raw = _raw()
    _disable_field(raw, "person_電話番号")
    t = load_template(_write(tmp_path, raw, "ac111.json"))
    columns = derive_columns(t)
    assert len(columns) == 219  # 220 - 1（person_電話番号を除いた実測）
    n_extract = len(columns) - len(META_COLUMNS)
    first = excel_column_letter(len(META_COLUMNS) + 1)
    last = excel_column_letter(len(columns))
    assert last != excel_column_letter(220)  # 前提: 実際に列が縮んでいる

    values = [UNCLEAR, "旭〓市"] + ["普通の値"] * (n_extract - 2)
    row = Row(page_id="p1", source_file="s.png", page_no=1, status="正常",
             values=values, unclear_count=2, min_conf="0.900",
             origins=("",) * n_extract)

    for on in (False, True):
        xlsx, csvp, _risky = write_outputs(
            tmp_path / f"out_{on}", f"t_{on}", columns, [row], unclear_char_level=on)
        sheet = _sheet_xml(xlsx)
        # ①COUNTIF 範囲が縮んだ列数（first〜last）へ追従
        pattern = '&quot;*〓*&quot;' if on else '&quot;〓&quot;'
        assert (f'COUNTIF({first}2:{last}2,{pattern})' in sheet
               or f'COUNTIF({first}2:{last}2,"{"*〓*" if on else "〓"}")' in sheet)
        # ②条件付き書式: OFF=1本・ON=2本。いずれも同じ縮んだ範囲を指す
        assert sheet.count("<cfRule") == (2 if on else 1)
        assert f"{first}1:{last}" in sheet
        # ③csv 側の静的カウントは xlsx 側と同じ値（③④の一致）
        text = csvp.read_text(encoding="utf-8-sig")
        assert text.splitlines()[1].split(",")[0] == '"2"'


# ---------- AC-1.3（実応答・remap 経由の復帰。API 送信なしを isolated_counter で実測） ----------

@needs_replay
def test_ac_1_3_toggle_off_then_on_via_remap_restores_same_value(tmp_path, isolated_counter):
    cfg = make_cfg(tmp_path)
    setup_done(tmp_path, cfg)
    used_after_run = api_budget.used_this_month()  # Replay は非加算（常に0）

    baseline_xlsx, _c, _rows = render(TPL, cfg, timestamp="base")
    base_header = [c.value for c in load_workbook(baseline_xlsx)["output"][1]]
    base_data = [c.value for c in load_workbook(baseline_xlsx)["output"][2]]
    baseline_phone = base_data[base_header.index("person_電話番号")]

    raw = _raw()
    _disable_field(raw, "person_電話番号")
    tpl_off = _write(tmp_path, raw, "off.json")

    # 対象外化（非幾何変更）は render をそのまま拒否し remap を名指しする
    with pytest.raises(OperationRefused, match="remap"):
        render(tpl_off, cfg, timestamp="off_attempt")

    assert remap(tpl_off, cfg) == 1
    xlsx_off, _c2, _rows_off = render(tpl_off, cfg, timestamp="off")
    assert api_budget.used_this_month() == used_after_run  # 送信カウンタ増分0

    header_off = [c.value for c in load_workbook(xlsx_off)["output"][1]]
    assert "person_電話番号" not in header_off

    # 解除（output 省略=true の元テンプレへ戻す）→ remap → 再出力
    assert remap(TPL, cfg) == 1
    xlsx_on, _c3, _rows_on = render(TPL, cfg, timestamp="on")
    header_on = [c.value for c in load_workbook(xlsx_on)["output"][1]]
    data_on = [c.value for c in load_workbook(xlsx_on)["output"][2]]
    assert "person_電話番号" in header_on
    assert data_on[header_on.index("person_電話番号")] == baseline_phone
    assert api_budget.used_this_month() == used_after_run  # 一連の操作で送信は一切増えない


# ---------- AC-1.17（決定性） ----------

@needs_replay
def test_ac_1_17_render_determinism_with_output_false_field(tmp_path):
    cfg = make_cfg(tmp_path)
    setup_done(tmp_path, cfg)
    raw = _raw()
    _disable_field(raw, "person_電話番号")
    tpl_off = _write(tmp_path, raw, "det.json")
    assert remap(tpl_off, cfg) == 1

    x1, c1, _r1 = render(tpl_off, cfg, timestamp="d1")
    x2, c2, _r2 = render(tpl_off, cfg, timestamp="d2")
    assert x1.read_bytes() == x2.read_bytes()
    assert c1.read_bytes() == c2.read_bytes()


# ---------- AC-1.19（xlsx↔csv・列数非依存の突合） ----------

@needs_replay
def test_ac_1_19_xlsx_csv_values_match_with_output_false_field(tmp_path):
    cfg = make_cfg(tmp_path)
    setup_done(tmp_path, cfg)
    raw = _raw()
    _disable_field(raw, "person_電話番号")
    tpl_off = _write(tmp_path, raw, "match.json")
    assert remap(tpl_off, cfg) == 1

    t = load_template(tpl_off)
    n_extract = len(extract_columns(t))
    assert n_extract == 213  # 214 - 1（決め打ちではなく導出結果として確認）

    xlsx, csvp, _rows = render(tpl_off, cfg, timestamp="g19")
    ws = load_workbook(xlsx)["output"]
    xrow = [c.value for c in list(ws.iter_rows(min_row=2))[0]][len(META_COLUMNS):]
    with open(csvp, encoding="utf-8-sig", newline="") as f:
        crow = list(csvmod.reader(f))[1][len(META_COLUMNS):]
    assert len(xrow) == len(crow) == n_extract
    for i, (xv, cv) in enumerate(zip(xrow, crow)):
        norm = "" if xv is None else str(xv)
        assert norm == cv, f"列{i + len(META_COLUMNS) + 1} で不一致"
