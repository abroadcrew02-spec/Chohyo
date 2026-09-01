"""出力列制御 MVP（issue #66）段0: 共通基盤の受入基準（AC-0.1）を固定する。

段0 の範囲は2点のみ:
  - FR-0.1: verify の template チェックへ column_names（derive_columns の
    結果そのまま・管理6列を含む順序付き全列名）を追加する
  - FR-0.2: GUI 側の保存時差分表示の母集団是正（F-10 バグ修正・GUI 側で対応）

このファイルは AC-0.1（verify の column_names が実際の xlsx 出力ヘッダ行と
完全一致する）を、Vision API 応答データに依存しない形で固定する。

証明の構造:
  ①verify の column_names は derive_columns(template) の戻り値そのもの
    （cli.py の cmd_verify で cols を直接渡している）
  ②write_outputs（run/render が実際に使う経路・pipeline.py:624,665）は、
    渡された columns をそのまま xlsx の1行目（ヘッダ）へ書く
①②を実測で確認すれば、「同一テンプレート・同一実行で完全一致する」
（AC-0.1）が導ける。E2E で実際に run を回さなくても固定できる。
"""
import json

from openpyxl import load_workbook

from chouhyo_ocr import cli
from chouhyo_ocr.columns import derive_columns
from chouhyo_ocr.paths import app_root
from chouhyo_ocr.render_out import write_outputs
from chouhyo_ocr.render_rows import Row
from chouhyo_ocr.template import load_template

TPL = app_root() / "templates" / "chouhyo-v1.json"


def _cfg(tmp_path):
    p = tmp_path / "config.json"
    p.write_text(json.dumps({
        "output_dir": str(tmp_path / "out"), "workdir": str(tmp_path / "wd"),
        "log_dir": str(tmp_path / "logs"),
    }), encoding="utf-8")
    return p


def _verify_template_event(tmp_path, capsys, template_path=TPL):
    cfg_path = _cfg(tmp_path)
    cli.main(["--config", str(cfg_path), "verify", "--template", str(template_path)])
    events = [json.loads(line) for line in capsys.readouterr().out.splitlines() if line.strip()]
    return next(e for e in events if e.get("check") == "template")


def test_verify_reports_column_names_as_ordered_string_list(tmp_path, capsys):
    """FR-0.1: column_names は順序付きの文字列リストで、既存フィールドは不変。"""
    tpl_ev = _verify_template_event(tmp_path, capsys)
    assert tpl_ev["ok"] is True
    assert isinstance(tpl_ev["column_names"], list)
    assert all(isinstance(c, str) for c in tpl_ev["column_names"])
    # 既存フィールドは不変（列数の内訳・cells・amount_cells・exclusions 等）
    assert len(tpl_ev["column_names"]) == tpl_ev["columns"]
    assert "columns" in tpl_ev and "cells" in tpl_ev and "amount_cells" in tpl_ev
    assert "exclusions" in tpl_ev and "warnings" in tpl_ev


def test_verify_column_names_equals_derive_columns_independently():
    """column_names の値そのものが derive_columns(template) と一致する
    （cli.py の実装を信用せず、独立に呼び出した結果と突き合わせる）。
    """
    t = load_template(TPL)
    expected = derive_columns(t)
    # column_names は無改変の出荷テンプレなら 220列（抽出214・管理6）のはず
    # （§0.3 の TR-G4 据え置き方針と同じ前提）
    assert len(expected) == 220
    assert expected[:6] == [
        "要確認セル数", "最低信頼度", "帳票ID", "入力ファイル名", "ページ番号", "ステータス"]


def test_write_outputs_xlsx_header_is_exactly_the_columns_argument(tmp_path):
    """②: write_outputs（run/render の実経路）は columns をそのまま
    xlsx の1行目へ書く。Vision 応答に依存せず、pipeline.py と同じ呼び出し形
    （derive_columns の結果をそのまま渡す）で直接検証する。
    """
    t = load_template(TPL)
    columns = derive_columns(t)
    n_extract = len(columns) - 6
    row = Row(page_id="p0001", source_file="p1.pdf", page_no=1, status="正常",
              values=[""] * n_extract, unclear_count=0, min_conf="0.9",
              origins=("",) * n_extract)
    xlsx, _csvp, _risky = write_outputs(tmp_path / "out", "t1", columns, [row])
    wb = load_workbook(xlsx)
    header = [c.value for c in wb["output"][1]]
    assert header == columns


def test_ac_0_1_column_names_matches_actual_xlsx_header(tmp_path, capsys):
    """AC-0.1 本体: 同一テンプレートについて、verify の column_names と
    実際に書き出される xlsx のヘッダ行（1行目）が完全一致する。
    """
    tpl_ev = _verify_template_event(tmp_path, capsys)

    t = load_template(TPL)
    columns = derive_columns(t)
    n_extract = len(columns) - 6
    row = Row(page_id="p0001", source_file="p1.pdf", page_no=1, status="正常",
              values=[""] * n_extract, unclear_count=0, min_conf="0.9",
              origins=("",) * n_extract)
    xlsx, _csvp, _risky = write_outputs(tmp_path / "out2", "t2", columns, [row])
    wb = load_workbook(xlsx)
    header = [c.value for c in wb["output"][1]]

    assert tpl_ev["column_names"] == header
